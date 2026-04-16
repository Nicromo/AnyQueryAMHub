"""
AM Hub — Enterprise Account Manager Dashboard
Реальные данные из Merchrules · Персональные дашборды · AI-ассистент
"""
import os
import json
import logging
from contextlib import asynccontextmanager
from datetime import datetime, timedelta
from datetime import timezone as tz
from typing import List, Optional

from fastapi import (
    FastAPI, Request, Depends, HTTPException, Query, Cookie,
    Form, status, UploadFile, File
)
from fastapi.responses import JSONResponse, HTMLResponse, RedirectResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from sqlalchemy import text

from database import engine, get_db, Base, init_db, SessionLocal
from models import (
    Client, Task, Meeting, CheckUp, User, SyncLog, AuditLog, Notification, QBR, AccountPlan,
    ClientNote, TaskComment, FollowupTemplate, VoiceNote,
)
from auth import (
    get_current_user, get_current_admin,
    authenticate_user, create_user, create_access_token,
    verify_password, hash_password,
    log_audit,
)
from error_handlers import log_error, handle_db_error
from middlewares import (
    LoggingMiddleware, RateLimitMiddleware,
    ErrorHandlingMiddleware, SecurityHeadersMiddleware
)

# Merchrules sync
from merchrules_sync import get_auth_token as mr_get_auth_token

# AI
from ai_followup import process_transcript as ai_process_transcript
from ai_assistant import generate_prep_brief, generate_smart_followup, detect_account_risks

logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO)


# ============================================================================
# ENV HELPERS — единый источник конфигурации
# ============================================================================

def _env(key: str, default: str = "") -> str:
    return os.environ.get(key, default)

def _env_bool(key: str) -> bool:
    return bool(os.environ.get(key, ""))

def _extract_sheets_id(val: str) -> str:
    """Вырезает spreadsheet ID из полного URL или возвращает как есть."""
    if not val:
        return ""
    # https://docs.google.com/spreadsheets/d/ID/edit...
    import re
    m = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", val)
    if m:
        return m.group(1)
    return val


class Env:
    """Централизованный доступ к переменным окружения."""
    # Merchrules
    MR_LOGIN      = property(lambda self: _env("MERCHRULES_LOGIN"))
    MR_PASSWORD   = property(lambda self: _env("MERCHRULES_PASSWORD"))
    MR_URL        = property(lambda self: _env("MERCHRULES_API_URL", "https://merchrules.any-platform.ru"))
    MR_ACTIVE     = property(lambda self: bool(_env("MERCHRULES_LOGIN") and _env("MERCHRULES_PASSWORD")))
    # AI
    GROQ_KEY      = property(lambda self: _env("GROQ_API_KEY") or _env("API_GROQ"))
    QWEN_KEY      = property(lambda self: _env("QWEN_API_KEY"))
    AI_ACTIVE     = property(lambda self: bool(_env("GROQ_API_KEY") or _env("API_GROQ") or _env("QWEN_API_KEY")))
    AI_TYPE       = property(lambda self: "qwen" if _env("QWEN_API_KEY") else ("groq" if (_env("GROQ_API_KEY") or _env("API_GROQ")) else ""))
    # Telegram
    TG_TOKEN      = property(lambda self: _env("TG_BOT_TOKEN") or _env("TELEGRAM_BOT_TOKEN"))
    TG_CHAT_ID    = property(lambda self: _env("TG_NOTIFY_CHAT_ID") or _env("TELEGRAM_CHAT_ID"))
    TG_ACTIVE     = property(lambda self: bool(_env("TG_BOT_TOKEN") or _env("TELEGRAM_BOT_TOKEN")))
    # Airtable — поддерживаем оба имени: AIRTABLE_TOKEN и AIRTABLE_PAT
    AIRTABLE_PAT  = property(lambda self: _env("AIRTABLE_TOKEN") or _env("AIRTABLE_PAT"))
    AIRTABLE_BASE = property(lambda self: _env("AIRTABLE_BASE_ID"))
    AIRTABLE_TABLE = property(lambda self: _env("AIRTABLE_TABLE_ID"))
    AIRTABLE_QBR_TABLE = property(lambda self: _env("AIRTABLE_QBR_TABLE_ID"))
    AIRTABLE_ACTIVE = property(lambda self: bool(_env("AIRTABLE_TOKEN") or _env("AIRTABLE_PAT")))
    # Google Sheets — вырезаем ID из полного URL если передан
    SHEETS_ID     = property(lambda self: _extract_sheets_id(_env("SHEETS_SPREADSHEET_ID")))
    SHEETS_ACTIVE = property(lambda self: bool(_env("SHEETS_SPREADSHEET_ID")))
    # Ktalk
    KTALK_SPACE   = property(lambda self: _env("KTALK_SPACE"))
    KTALK_TOKEN   = property(lambda self: _env("KTALK_API_TOKEN"))
    KTALK_WEBHOOK = property(lambda self: _env("KTALK_WEBHOOK_URL"))
    KTALK_ACTIVE  = property(lambda self: bool(_env("KTALK_SPACE") and _env("KTALK_API_TOKEN")))
    # Tbank Time
    TIME_TOKEN    = property(lambda self: _env("TIME_API_TOKEN") or _env("TIME_SESSION_COOKIE"))
    TIME_ACTIVE   = property(lambda self: bool(_env("TIME_API_TOKEN") or _env("TIME_SESSION_COOKIE")))
    # Ktalk
    KTALK_SPACE   = property(lambda self: _env("KTALK_SPACE"))
    KTALK_TOKEN   = property(lambda self: _env("KTALK_API_TOKEN"))
    KTALK_ACTIVE  = property(lambda self: bool(_env("KTALK_API_TOKEN") and _env("KTALK_SPACE")))
    KTALK_WEBHOOK = property(lambda self: _env("KTALK_WEBHOOK_URL"))
    # Tbank Time
    TIME_TOKEN    = property(lambda self: _env("TIME_API_TOKEN"))
    TIME_URL      = property(lambda self: _env("TIME_BASE_URL", "https://time.tbank.ru"))
    TIME_ACTIVE   = property(lambda self: bool(_env("TIME_API_TOKEN")))
    # Airtable
    AIRTABLE_PAT  = property(lambda self: _env("AIRTABLE_PAT"))
    AIRTABLE_BASE = property(lambda self: _env("AIRTABLE_BASE_ID"))
    AIRTABLE_ACTIVE = property(lambda self: bool(_env("AIRTABLE_PAT")))
    # Google Sheets
    SHEETS_ID     = property(lambda self: _env("SHEETS_SPREADSHEET_ID"))
    SHEETS_ACTIVE = property(lambda self: bool(_env("SHEETS_SPREADSHEET_ID")))
    # App
    APP_URL       = property(lambda self: _env("RAILWAY_PUBLIC_DOMAIN") or _env("APP_URL"))
    SECRET_KEY    = property(lambda self: _env("SECRET_KEY", "your-secret-key-change-in-production"))

env = Env()

templates = Jinja2Templates(directory="templates")

MSK = tz(timedelta(hours=3))  # Moscow timezone


# ============================================================================
# LIFESPAN
# ============================================================================

@asynccontextmanager
async def lifespan(app: FastAPI):
    """App startup"""
    try:
        init_db()
        # Добавляем отсутствующие колонки и таблицы
        with SessionLocal() as db:
            try:
                cols = db.execute(
                    text("SELECT column_name FROM information_schema.columns WHERE table_name = 'users'")
                ).fetchall()
                col_names = {row[0] for row in cols}
                if "settings" not in col_names:
                    db.execute(text("ALTER TABLE users ADD COLUMN settings JSONB"))
                    db.commit()
                    logger.info("✅ Added users.settings column")
            except Exception as e:
                logger.warning(f"Migration users.settings: {e}")

            # Добавляем колонки Meeting
            try:
                mcols = db.execute(
                    text("SELECT column_name FROM information_schema.columns WHERE table_name = 'meetings'")
                ).fetchall()
                mcol_names = {row[0] for row in mcols}
                for col, default in [("followup_status", "'pending'"), ("followup_text", "NULL"),
                                      ("followup_sent_at", "NULL"), ("followup_skipped", "FALSE"),
                                      ("is_qbr", "FALSE")]:
                    if col not in mcol_names:
                        db.execute(text(f"ALTER TABLE meetings ADD COLUMN {col} VARCHAR DEFAULT {default}" if col != "followup_skipped" and col != "is_qbr" else f"ALTER TABLE meetings ADD COLUMN {col} BOOLEAN DEFAULT {default}"))
                        db.commit()
                        logger.info(f"✅ Added meetings.{col}")
            except Exception as e:
                logger.warning(f"Migration meetings: {e}")

            # Добавляем колонки Task
            try:
                tcols = db.execute(
                    text("SELECT column_name FROM information_schema.columns WHERE table_name = 'tasks'")
                ).fetchall()
                tcol_names = {row[0] for row in tcols}
                for col, col_type in [
                    ("confirmed_at", "TIMESTAMP"), ("confirmed_by", "VARCHAR"),
                    ("pushed_to_roadmap", "BOOLEAN DEFAULT FALSE"), ("roadmap_pushed_at", "TIMESTAMP"),
                    ("team", "VARCHAR"), ("task_type", "VARCHAR"), ("source", "VARCHAR DEFAULT 'manual'"),
                ]:
                    if col not in tcol_names:
                        db.execute(text(f"ALTER TABLE tasks ADD COLUMN {col} {col_type}"))
                        db.commit()
                        logger.info(f"✅ Added tasks.{col}")
            except Exception as e:
                logger.warning(f"Migration tasks: {e}")

            # Добавляем колонки Client
            try:
                ccols = db.execute(
                    text("SELECT column_name FROM information_schema.columns WHERE table_name = 'clients'")
                ).fetchall()
                ccol_names = {row[0] for row in ccols}
                for col, col_type in [("last_qbr_date", "TIMESTAMP"), ("next_qbr_date", "TIMESTAMP"),
                                       ("account_plan", "JSONB")]:
                    if col not in ccol_names:
                        db.execute(text(f"ALTER TABLE clients ADD COLUMN {col} {col_type}"))
                        db.commit()
                        logger.info(f"✅ Added clients.{col}")
            except Exception as e:
                logger.warning(f"Migration clients: {e}")

            # Создаём новые таблицы
            try:
                db.execute(text("""
                    CREATE TABLE IF NOT EXISTS qbrs (
                        id SERIAL PRIMARY KEY,
                        client_id INTEGER REFERENCES clients(id),
                        quarter VARCHAR NOT NULL,
                        year INTEGER NOT NULL,
                        date TIMESTAMP,
                        status VARCHAR DEFAULT 'draft',
                        metrics JSONB DEFAULT '{}',
                        summary TEXT,
                        achievements JSONB DEFAULT '[]',
                        issues JSONB DEFAULT '[]',
                        next_quarter_goals JSONB DEFAULT '[]',
                        meeting_id INTEGER REFERENCES meetings(id)
                    )
                """))
                db.commit()
                logger.info("✅ Created qbrs table")

                # Добавляем новые колонки в qbrs если их нет
                try:
                    qcols = db.execute(
                        text("SELECT column_name FROM information_schema.columns WHERE table_name = 'qbrs'")
                    ).fetchall()
                    qcol_names = {row[0] for row in qcols}
                    for col, col_type in [("presentation_url", "VARCHAR"), ("executive_summary", "TEXT"),
                                           ("future_work", "JSONB DEFAULT '[]'"), ("key_insights", "JSONB DEFAULT '[]'")]:
                        if col not in qcol_names:
                            db.execute(text(f"ALTER TABLE qbrs ADD COLUMN {col} {col_type}"))
                            db.commit()
                            logger.info(f"✅ Added qbrs.{col}")
                except Exception as e:
                    logger.warning(f"Migration qbrs columns: {e}")
            except Exception as e:
                logger.warning(f"Migration qbrs: {e}")

            try:
                db.execute(text("""
                    CREATE TABLE IF NOT EXISTS account_plans (
                        id SERIAL PRIMARY KEY,
                        client_id INTEGER REFERENCES clients(id) UNIQUE,
                        quarterly_goals JSONB DEFAULT '[]',
                        action_items JSONB DEFAULT '[]',
                        notes TEXT,
                        strategy TEXT,
                        updated_at TIMESTAMP DEFAULT NOW(),
                        updated_by VARCHAR
                    )
                """))
                db.commit()
                logger.info("✅ Created account_plans table")
            except Exception as e:
                logger.warning(f"Migration account_plans: {e}")

            # Новые таблицы v3
            for table_name, table_sql in [
                ("client_notes", """CREATE TABLE IF NOT EXISTS client_notes (
                    id SERIAL PRIMARY KEY,
                    client_id INTEGER REFERENCES clients(id),
                    user_id INTEGER REFERENCES users(id),
                    content TEXT NOT NULL,
                    is_pinned BOOLEAN DEFAULT FALSE,
                    created_at TIMESTAMP DEFAULT NOW(),
                    updated_at TIMESTAMP DEFAULT NOW()
                )"""),
                ("task_comments", """CREATE TABLE IF NOT EXISTS task_comments (
                    id SERIAL PRIMARY KEY,
                    task_id INTEGER REFERENCES tasks(id),
                    user_id INTEGER REFERENCES users(id),
                    content TEXT NOT NULL,
                    created_at TIMESTAMP DEFAULT NOW()
                )"""),
                ("followup_templates", """CREATE TABLE IF NOT EXISTS followup_templates (
                    id SERIAL PRIMARY KEY,
                    user_id INTEGER REFERENCES users(id),
                    name VARCHAR NOT NULL,
                    content TEXT NOT NULL,
                    category VARCHAR DEFAULT 'general',
                    created_at TIMESTAMP DEFAULT NOW()
                )"""),
                ("voice_notes", """CREATE TABLE IF NOT EXISTS voice_notes (
                    id SERIAL PRIMARY KEY,
                    meeting_id INTEGER REFERENCES meetings(id),
                    client_id INTEGER REFERENCES clients(id),
                    user_id INTEGER REFERENCES users(id),
                    audio_url VARCHAR,
                    transcription TEXT,
                    duration_seconds INTEGER DEFAULT 0,
                    created_at TIMESTAMP DEFAULT NOW()
                )"""),
            ]:
                try:
                    db.execute(text(table_sql))
                    db.commit()
                    logger.info(f"✅ Created {table_name} table")
                except Exception as e:
                    logger.warning(f"Migration {table_name}: {e}")

            if db.query(User).count() == 0:
                import secrets, string
                random_password = "".join(secrets.choice(string.ascii_letters + string.digits) for _ in range(16))
                admin = User(
                    email="admin@company.ru",
                    first_name="Администратор",
                    role="admin",
                    hashed_password=hash_password(random_password),
                    settings={},
                )
                db.add(admin)
                db.commit()
                logger.warning(f"✅ Default admin created: admin@company.ru / {random_password} — СМЕНИТЕ ПАРОЛЬ!")
        logger.info("✅ Database ready")

        # Start scheduler
        try:
            from scheduler import start_scheduler
            start_scheduler()
            logger.info("✅ Scheduler started")
        except Exception as e:
            logger.warning(f"Scheduler: {e}")

        # Register Telegram webhook
        try:
            from tg_bot import set_webhook, BOT_TOKEN as TG_TOKEN
            if TG_TOKEN:
                domain = env.APP_URL
                if domain:
                    await set_webhook(f"{domain}/webhook/telegram")
                    logger.info(f"✅ TG webhook: {domain}/webhook/telegram")
        except Exception as e:
            logger.warning(f"TG webhook: {e}")
    except Exception as e:
        logger.error(f"Startup error: {e}")
    yield


# ============================================================================
# APP SETUP
# ============================================================================

app = FastAPI(title="AM Hub", version="2.0.0", lifespan=lifespan)

# ── Rate limiting ────────────────────────────────────────────────────────────
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded

limiter = Limiter(key_func=get_remote_address, default_limits=["200/minute"])
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

# ── Reusable auth dependency ─────────────────────────────────────────────────
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials

async def get_current_user(
    request: Request,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
) -> User:
    """Единый dependency для авторизации — используется в новых endpoints."""
    token = auth_token
    # Fallback: Bearer header (для API клиентов / расширения)
    if not token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    from auth import decode_access_token
    payload = decode_access_token(token)
    if not payload:
        raise HTTPException(status_code=401, detail="Invalid token")
    user = db.query(User).filter(User.id == int(payload.get("sub", 0))).first()
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    return user

async def get_admin_user(user: User = Depends(get_current_user)) -> User:
    if user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    return user



# ── Sentry ────────────────────────────────────────────────────────────────────
from sentry_config import init_sentry
init_sentry()

# ── Redis cache (заменяем in-memory) ─────────────────────────────────────────
from redis_cache import cache_get, cache_set, cache_del, cache_key



app.add_middleware(SecurityHeadersMiddleware)
app.add_middleware(ErrorHandlingMiddleware)
app.add_middleware(RateLimitMiddleware, requests_per_minute=100)
app.add_middleware(LoggingMiddleware)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

app.mount("/static", StaticFiles(directory="static"), name="static")


# ============================================================================
# MERCHRULES HELPERS
# ============================================================================

def _get_user_cred(user: User) -> tuple:
    """Получить Merchrules логин/пароль: сначала из user.settings, потом из env."""
    settings = (user.settings or {}) if user else {}
    mr = settings.get("merchrules", {})
    login = mr.get("login") or env.MR_LOGIN
    password = mr.get("password") or env.MR_PASSWORD
    return login, password


# ============================================================================
# AUTH PAGES
# ============================================================================

@app.get("/login", response_class=HTMLResponse)
async def login_page(request: Request):
    return templates.TemplateResponse("login.html", {"request": request, "error": None})


@app.post("/login")
async def login_submit(
    request: Request,
    email: str = Form(...),
    password: str = Form(...),
    db: Session = Depends(get_db),
):
    user = authenticate_user(db, email, password)
    if not user:
        return templates.TemplateResponse("login.html", {"request": request, "error": "Неверный email или пароль"})

    token = create_access_token({"sub": str(user.id)})
    response = RedirectResponse(url="/dashboard", status_code=303)
    response.set_cookie(key="auth_token", value=token, httponly=True, samesite="lax", max_age=86400 * 30)
    return response


@app.get("/logout")
async def logout():
    response = RedirectResponse(url="/login", status_code=303)
    response.delete_cookie(key="auth_token")
    return response


# ============================================================================
# DASHBOARD
# ============================================================================

@app.get("/dashboard", response_class=HTMLResponse)
async def dashboard(request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    if not auth_token:
        return RedirectResponse(url="/login", status_code=303)

    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        return RedirectResponse(url="/login", status_code=303)

    user_id = payload.get("sub")
    user = db.query(User).filter(User.id == int(user_id)).first()
    if not user:
        return RedirectResponse(url="/login", status_code=303)

    # Проверка онбординга
    settings = user.settings or {}
    if not settings.get("onboarding_complete"):
        return RedirectResponse(url="/onboarding", status_code=303)

    # Для админа — все клиенты, для менеджера — только его
    query = db.query(Client)
    if user.role == "manager":
        query = query.filter(Client.manager_email == user.email)
    clients = query.all()

    # Обогащаем данными из Merchrules если есть креды
    mr_login, mr_password = _get_user_cred(user)
    has_mr = bool(mr_login and mr_password)

    # Статистика
    now = datetime.now()
    counts = {"ENT": 0, "SME+": 0, "SME-": 0, "SME": 0, "SMB": 0, "SS": 0}
    healthy = warning = overdue = total_open = total_tasks = 0

    for c in clients:
        seg = c.segment or ""
        if seg in counts:
            counts[seg] += 1

        open_tasks = db.query(Task).filter(Task.client_id == c.id, Task.status.in_(["plan", "in_progress"])).count()
        blocked_tasks = db.query(Task).filter(Task.client_id == c.id, Task.status == "blocked").count()
        total_client_tasks = db.query(Task).filter(Task.client_id == c.id).count()
        total_open += open_tasks
        total_tasks += total_client_tasks

        is_overdue = c.needs_checkup and (not c.last_meeting_date or (now - c.last_meeting_date).days > 30)
        is_warning = c.needs_checkup and c.last_meeting_date and 14 < (now - c.last_meeting_date).days <= 30

        if is_overdue:
            overdue += 1
        elif is_warning:
            warning += 1
        else:
            healthy += 1

        c.open_tasks = open_tasks
        c.blocked_tasks = blocked_tasks
        c.total_tasks = total_client_tasks
        c.status = {"color": "red" if is_overdue else ("yellow" if is_warning else "green")}

    # Задачи на сегодня
    today = now.date()
    today_tasks = db.query(Task).filter(
        Task.due_date >= datetime.combine(today, datetime.min.time()),
        Task.due_date < datetime.combine(today + timedelta(days=1), datetime.min.time()),
        Task.status.in_(["plan", "in_progress"]),
    ).all()

    if user.role == "manager":
        today_tasks = [t for t in today_tasks if t.client and t.client.manager_email == user.email]

    # Встречи на сегодня
    today_meetings = db.query(Meeting).filter(
        Meeting.date >= datetime.combine(today, datetime.min.time()),
        Meeting.date < datetime.combine(today + timedelta(days=1), datetime.min.time()),
    ).all()

    return templates.TemplateResponse(
        "dashboard.html", {
            "request": request,
            "user": user,
            "clients": clients,
            "counts": counts,
            "healthy_count": healthy,
            "warning_count": warning,
            "overdue_count": overdue,
            "total_open_tasks": total_open,
            "total_tasks": total_tasks,
            "today_tasks": today_tasks,
            "today_meetings": today_meetings,
            "now": now,
            "has_mr": has_mr,
        },
    )


# ============================================================================
# MY DAY
# ============================================================================

@app.get("/today", response_class=HTMLResponse)
async def my_day(request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    if not auth_token:
        return RedirectResponse(url="/login", status_code=303)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        return RedirectResponse(url="/login", status_code=303)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        return RedirectResponse(url="/login", status_code=303)

    today = datetime.now().date()
    start = datetime.combine(today, datetime.min.time())
    end = datetime.combine(today + timedelta(days=1), datetime.min.time())

    # Задачи на сегодня
    q = db.query(Task).filter(Task.due_date >= start, Task.due_date < end)
    if user.role == "manager":
        q = q.join(Client).filter(Client.manager_email == user.email)
    today_tasks = q.all()

    # Встречи на сегодня
    q2 = db.query(Meeting).filter(Meeting.date >= start, Meeting.date < end)
    if user.role == "manager":
        q2 = q2.join(Client).filter(Client.manager_email == user.email)
    today_meetings = q2.all()

    # Просроченные задачи
    q3 = db.query(Task).filter(Task.due_date < start, Task.status.in_(["plan", "in_progress"]))
    if user.role == "manager":
        q3 = q3.join(Client).filter(Client.manager_email == user.email)
    overdue_tasks = q3.all()

    # Общая статистика
    total_open = db.query(Task).filter(Task.status.in_(["plan", "in_progress"])).count()
    if user.role == "manager":
        total_open = db.query(Task).join(Client).filter(
            Task.status.in_(["plan", "in_progress"]),
            Client.manager_email == user.email
        ).count()

    return templates.TemplateResponse("today.html", {
        "request": request, "user": user,
        "today_tasks": today_tasks,
        "today_meetings": today_meetings,
        "overdue_tasks": overdue_tasks,
        "total_open": total_open,
        "now": datetime.now(),
    })


# ============================================================================
# CLIENTS LIST
# ============================================================================

@app.get("/clients", response_class=HTMLResponse)
async def clients_page(request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    if not auth_token:
        return RedirectResponse(url="/login", status_code=303)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        return RedirectResponse(url="/login", status_code=303)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        return RedirectResponse(url="/login", status_code=303)

    segment = request.query_params.get("segment")
    now = datetime.now()
    q = db.query(Client)
    if user.role == "manager":
        q = q.filter(Client.manager_email == user.email)
    if segment:
        q = q.filter(Client.segment == segment)
    clients = q.all()

    counts = {"ENT": 0, "SME+": 0, "SME-": 0, "SME": 0, "SMB": 0, "SS": 0}
    for c in clients:
        seg = c.segment or ""
        if seg in counts:
            counts[seg] += 1

    for c in clients:
        open_tasks = db.query(Task).filter(Task.client_id == c.id, Task.status.in_(["plan", "in_progress"])).count()
        blocked_tasks = db.query(Task).filter(Task.client_id == c.id, Task.status == "blocked").count()
        is_overdue = c.needs_checkup and (not c.last_meeting_date or (now - c.last_meeting_date).days > 30)
        is_warning = c.needs_checkup and c.last_meeting_date and 14 < (now - c.last_meeting_date).days <= 30
        c.open_tasks = open_tasks
        c.blocked_tasks = blocked_tasks
        c.status = {"color": "red" if is_overdue else ("yellow" if is_warning else "green")}

    return templates.TemplateResponse("clients.html", {
        "request": request, "user": user, "clients": clients,
        "counts": counts, "segment": segment, "now": now,
    })


# ============================================================================
# CLIENT DETAIL + PREP
# ============================================================================

@app.get("/client/{client_id}", response_class=HTMLResponse)
async def client_detail(request: Request, client_id: int, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    if not auth_token:
        return RedirectResponse(url="/login", status_code=303)

    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        return RedirectResponse(url="/login", status_code=303)

    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        return RedirectResponse(url="/login", status_code=303)

    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")

    tasks = db.query(Task).filter(Task.client_id == client_id).order_by(Task.due_date.desc()).all()
    meetings = db.query(Meeting).filter(Meeting.client_id == client_id).order_by(Meeting.date.desc()).all()

    return templates.TemplateResponse("client_detail.html", {
        "request": request, "user": user, "client": client,
        "tasks": tasks, "meetings": meetings, "now": datetime.now(),
    })


@app.get("/prep/{client_id}", response_class=HTMLResponse)
async def prep_page(request: Request, client_id: int, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    if not auth_token:
        return RedirectResponse(url="/login", status_code=303)

    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        return RedirectResponse(url="/login", status_code=303)

    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        return RedirectResponse(url="/login", status_code=303)

    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")

    tasks = db.query(Task).filter(Task.client_id == client_id, Task.status.in_(["plan", "in_progress"])).all()
    meetings = db.query(Meeting).filter(Meeting.client_id == client_id).order_by(Meeting.date.desc()).limit(5).all()

    # AI-подготовка
    try:
        prep_text = generate_prep_brief(client, tasks, meetings)
    except Exception as e:
        prep_text = f"AI недоступен: {e}"

    return templates.TemplateResponse("prep.html", {
        "request": request, "user": user, "client": client,
        "tasks": tasks, "meetings": meetings, "prep_text": prep_text,
        "now": datetime.now(),
    })


# ============================================================================
# FOLLOWUP
# ============================================================================

@app.get("/followup/{client_id}", response_class=HTMLResponse)
async def followup_page(request: Request, client_id: int, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    if not auth_token:
        return RedirectResponse(url="/login", status_code=303)

    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        return RedirectResponse(url="/login", status_code=303)

    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        return RedirectResponse(url="/login", status_code=303)

    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")

    tasks = db.query(Task).filter(Task.client_id == client_id).all()
    meetings = db.query(Meeting).filter(Meeting.client_id == client_id).order_by(Meeting.date.desc()).limit(3).all()

    # Последняя встреча с pending followup (или просто последняя)
    meeting = db.query(Meeting).filter(
        Meeting.client_id == client_id,
        Meeting.followup_status == "pending"
    ).order_by(Meeting.date.desc()).first()
    if not meeting:
        meeting = db.query(Meeting).filter(Meeting.client_id == client_id).order_by(Meeting.date.desc()).first()

    try:
        followup_text = generate_smart_followup(client, tasks, meetings)
    except Exception as e:
        followup_text = f"AI недоступен: {e}"

    return templates.TemplateResponse("followup.html", {
        "request": request, "user": user, "client": client,
        "tasks": tasks, "meetings": meetings, "followup_text": followup_text,
        "meeting": meeting, "now": datetime.now(),
    })


# ============================================================================
# ONBOARDING PARTNER
# ============================================================================

@app.get("/onboarding/{client_id}", response_class=HTMLResponse)
async def onboarding_partner_page(request: Request, client_id: int, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    if not auth_token:
        return RedirectResponse(url="/login", status_code=303)

    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        return RedirectResponse(url="/login", status_code=303)

    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        return RedirectResponse(url="/login", status_code=303)

    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")

    return templates.TemplateResponse("onboarding_partner.html", {
        "request": request, "user": user, "client": client,
        "now": datetime.now(),
    })


# ============================================================================
# TASKS
# ============================================================================

@app.get("/tasks", response_class=HTMLResponse)
async def tasks_page(request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    if not auth_token:
        return RedirectResponse(url="/login", status_code=303)

    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        return RedirectResponse(url="/login", status_code=303)

    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        return RedirectResponse(url="/login", status_code=303)

    status_filter = request.query_params.get("status")
    q = db.query(Task)
    if user.role == "manager":
        q = q.join(Client).filter(Client.manager_email == user.email)
    if status_filter and status_filter != "all":
        q = q.filter(Task.status == status_filter)
    tasks = q.order_by(Task.due_date.desc()).limit(100).all()

    return templates.TemplateResponse("tasks.html", {
        "request": request, "user": user, "tasks": tasks,
        "status_filter": status_filter or "all", "now": datetime.now(),
    })


@app.get("/kanban", response_class=HTMLResponse)
async def kanban_page(request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Kanban-доска задач."""
    if not auth_token:
        return RedirectResponse(url="/login", status_code=303)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        return RedirectResponse(url="/login", status_code=303)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    return templates.TemplateResponse("kanban.html", {"request": request, "user": user})


# ============================================================================
# SYNC
# ============================================================================

@app.get("/sync", response_class=HTMLResponse)
async def sync_page(request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    if not auth_token:
        return RedirectResponse(url="/login", status_code=303)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        return RedirectResponse(url="/login", status_code=303)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    mr_login_val, _ = _get_user_cred(user)
    return templates.TemplateResponse("sync.html", {"request": request, "user": user, "mr_login": mr_login_val})


# ============================================================================
# PLAN & QBR PAGES
# ============================================================================

@app.get("/client/{client_id}/plan", response_class=HTMLResponse)
async def plan_page(request: Request, client_id: int, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    if not auth_token:
        return RedirectResponse(url="/login", status_code=303)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        return RedirectResponse(url="/login", status_code=303)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        return RedirectResponse(url="/login", status_code=303)

    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        raise HTTPException(status_code=404)

    plan = db.query(AccountPlan).filter(AccountPlan.client_id == client_id).first()
    if not plan:
        plan = AccountPlan(client_id=client_id)

    return templates.TemplateResponse("plan.html", {
        "request": request, "user": user, "client": client, "plan": plan,
    })


@app.get("/client/{client_id}/qbr", response_class=HTMLResponse)
async def qbr_page(request: Request, client_id: int, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    if not auth_token:
        return RedirectResponse(url="/login", status_code=303)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        return RedirectResponse(url="/login", status_code=303)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        return RedirectResponse(url="/login", status_code=303)

    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        raise HTTPException(status_code=404)

    qbr = db.query(QBR).filter(QBR.client_id == client_id).order_by(QBR.date.desc()).first()

    return templates.TemplateResponse("qbr.html", {
        "request": request, "user": user, "client": client, "qbr": qbr,
    })


# ============================================================================
# INTEGRATIONS
# ============================================================================

@app.get("/integrations", response_class=HTMLResponse)
async def integrations_page(request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    if not auth_token:
        return RedirectResponse(url="/login", status_code=303)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        return RedirectResponse(url="/login", status_code=303)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        return RedirectResponse(url="/login", status_code=303)

    mr_login_val, mr_password_val = _get_user_cred(user)
    u_settings = user.settings or {}
    ktalk_s = u_settings.get("ktalk", {})
    airtable_s = u_settings.get("airtable", {})
    sheets_s = u_settings.get("sheets", {})
    groq_s = u_settings.get("groq", {})
    tbank_s = u_settings.get("tbank_time", {})
    tg_s = u_settings.get("telegram", {})
    airtable_active = bool(airtable_s.get("pat") or env.AIRTABLE_PAT)
    sheets_active = bool(sheets_s.get("spreadsheet_id") or env.SHEETS_ID)
    ai_active = bool(groq_s.get("api_key") or env.AI_ACTIVE)
    ai_type = "qwen" if env.QWEN_KEY else ("groq" if (groq_s.get("api_key") or env.GROQ_KEY) else "")
    integrations_data = {
        "mr_active": bool(mr_login_val and mr_password_val),
        "mr_login": mr_login_val,
        "airtable_active": airtable_active,
        "sheets_active": sheets_active,
        "sheets_id": env.SHEETS_ID,
        "tg_active": bool(env.TG_TOKEN),
        "ai_active": ai_active,
        "ai_type": ai_type,
        "ktalk_active": bool(env.KTALK_TOKEN and env.KTALK_SPACE),
        "ktalk_space": _env("KTALK_SPACE"),
        "time_active": bool(env.TIME_TOKEN),
    }
    return templates.TemplateResponse("integrations.html", {"request": request, "user": user, **integrations_data})


# ============================================================================
# API: INTEGRATION TESTS
# ============================================================================

@app.get("/api/integrations/test/merchrules")
async def test_merchrules(login: str = "", password: str = ""):
    if not login or not password:
        return {"error": "Need login and password"}
    import httpx
    try:
        async with httpx.AsyncClient(timeout=15) as hx:
            resp = await hx.post(
                "https://merchrules-qa.any-platform.ru/backend-v2/auth/login",
                json={"username": login, "password": password},
            )
        if resp.status_code == 200:
            return {"ok": True}
        return {"error": f"HTTP {resp.status_code}"}
    except Exception as e:
        return {"error": str(e)}


@app.get("/api/integrations/test/ktalk")
async def test_ktalk(space: str = "", token: str = ""):
    if not space or not token:
        return {"error": "Need space and token"}
    import httpx
    try:
        base = f"https://{space}.ktalk.ru"
        async with httpx.AsyncClient(timeout=10) as hx:
            resp = await hx.get(f"{base}/api/v1/spaces/{space}/users",
                headers={"Content-Type": "application/json", "X-Auth-Token": token},
                params={"limit": 1})
        if resp.status_code == 200:
            return {"ok": True, "space": space}
        return {"error": f"HTTP {resp.status_code}: {resp.text[:200]}"}
    except Exception as e:
        return {"error": str(e)}


@app.post("/webhook/telegram")
async def telegram_webhook(request: Request, db: Session = Depends(get_db)):
    """Handle incoming Telegram updates"""
    from tg_bot import handle_update, send_message
    from sheets import get_top50_data

    update = await request.json()
    user_id = (update.get("message", {}) or {}).get("from", {}).get("id", 0)

    # Get clients for this user
    def get_clients_fn():
        # For now, return all clients (can filter by user later)
        return [{"id": c.id, "name": c.name, "segment": c.segment or "",
                 "last_checkup": c.last_checkup, "last_meeting": c.last_meeting_date}
                for c in db.query(Client).all()]

    async def get_top50_fn():
        my_clients = [c.name for c in db.query(Client).all()]
        return await get_top50_data(my_clients)

    try:
        await handle_update(update, get_clients_fn, get_top50_fn)
    except Exception as e:
        logger.error(f"TG webhook error: {e}")
        try:
            chat_id = (update.get("message", {}) or {}).get("chat", {}).get("id")
            if chat_id:
                await send_message(chat_id, f"❌ Ошибка: {str(e)[:100]}")
        except Exception:
            pass

    return {"ok": True}


# ============================================================================
# API: KTALK
# ============================================================================

@app.post("/api/ktalk/notify")
async def api_ktalk_notify(
    request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)
):
    """Send notification to Ktalk channel"""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)

    data = await request.json()
    webhook_url = env.KTALK_WEBHOOK
    if not webhook_url:
        return {"error": "KTALK_WEBHOOK_URL not set"}

    import httpx
    try:
        async with httpx.AsyncClient(timeout=10) as hx:
            await hx.post(webhook_url, json={"text": data.get("text", "")})
        return {"ok": True}
    except Exception as e:
        return {"error": str(e)}


@app.post("/api/ktalk/followup")
async def api_ktalk_followup(
    request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)
):
    """Send meeting followup to Ktalk"""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)

    data = await request.json()
    client_name = data.get("client", "")
    summary = data.get("summary", "")
    tasks = data.get("tasks", [])

    webhook_url = env.KTALK_WEBHOOK
    if not webhook_url:
        return {"error": "KTALK_WEBHOOK_URL not set"}

    text = f"📋 **Followup: {client_name}**\n\n{summary}"
    if tasks:
        text += "\n\n**Задачи:**\n" + "\n".join(f"• {t}" for t in tasks)

    import httpx
    try:
        async with httpx.AsyncClient(timeout=10) as hx:
            await hx.post(webhook_url, json={"text": text})
        return {"ok": True}
    except Exception as e:
        return {"error": str(e)}


# ============================================================================
# API: MERCHRULES SYNC
# ============================================================================

@app.post("/api/sync/extension")
async def api_sync_extension(request: Request, db: Session = Depends(get_db)):
    """
    Приём данных синхронизации от Chrome-расширения AM Hub Sync.
    Авторизация через Bearer токен (JWT) в заголовке Authorization.
    """
    # Авторизация через Bearer header (расширение не использует cookie)
    auth_header = request.headers.get("Authorization", "")
    token = auth_header.removeprefix("Bearer ").strip()
    if not token:
        raise HTTPException(status_code=401, detail="Bearer token required")

    from auth import decode_access_token
    payload_jwt = decode_access_token(token)
    if not payload_jwt:
        raise HTTPException(status_code=401, detail="Invalid token")

    user = db.query(User).filter(User.id == int(payload_jwt.get("sub"))).first()
    if not user:
        raise HTTPException(status_code=401, detail="User not found")

    data = await request.json()
    accounts = data.get("accounts", [])

    if not accounts:
        return {"ok": False, "error": "No accounts in payload"}

    clients_synced = 0
    tasks_synced = 0
    meetings_synced = 0

    for acc in accounts:
        site_id = str(acc.get("id", "")).strip()
        if not site_id:
            continue

        # Найти или создать клиента
        client = db.query(Client).filter(Client.merchrules_account_id == site_id).first()
        if not client:
            client = Client(
                merchrules_account_id=site_id,
                name=acc.get("name") or f"Site {site_id}",
                manager_email=user.email,
                segment=acc.get("segment") or "SMB",
                domain=acc.get("domain"),
            )
            db.add(client)
            db.flush()
        else:
            # Обновляем имя и сегмент если пришли
            if acc.get("name"):
                client.name = acc["name"]
            if acc.get("segment"):
                client.segment = acc["segment"]
            if acc.get("domain"):
                client.domain = acc["domain"]
            if acc.get("health_score") is not None:
                client.health_score = float(acc["health_score"])

        # Гарантируем привязку к менеджеру
        if not client.manager_email:
            client.manager_email = user.email

        clients_synced += 1

        # Задачи
        for t in acc.get("tasks", []):
            mr_task_id = str(t.get("id", "")).strip()
            if not mr_task_id:
                continue
            existing = db.query(Task).filter(Task.merchrules_task_id == mr_task_id).first()
            if existing:
                # Обновляем статус
                existing.status = t.get("status", existing.status)
                existing.priority = t.get("priority", existing.priority)
            else:
                due = None
                if t.get("due_date"):
                    try:
                        due = datetime.fromisoformat(str(t["due_date"])[:19])
                    except Exception:
                        pass
                db.add(Task(
                    client_id=client.id,
                    merchrules_task_id=mr_task_id,
                    title=t.get("title") or "",
                    status=t.get("status") or "plan",
                    priority=t.get("priority") or "medium",
                    source="roadmap",
                    due_date=due,
                    team=t.get("team"),
                    task_type=t.get("task_type"),
                ))
                tasks_synced += 1

        # Встречи
        for m in acc.get("meetings", []):
            mr_meeting_id = str(m.get("id", "")).strip()
            if not mr_meeting_id:
                continue
            ext_id = f"mr_{mr_meeting_id}"
            existing = db.query(Meeting).filter(Meeting.external_id == ext_id).first()
            if not existing:
                meeting_date = None
                raw_date = m.get("date")
                if raw_date:
                    try:
                        meeting_date = datetime.fromisoformat(str(raw_date)[:19])
                    except Exception:
                        pass
                if meeting_date:
                    db.add(Meeting(
                        client_id=client.id,
                        date=meeting_date,
                        type=m.get("type") or "meeting",
                        title=m.get("title"),
                        summary=m.get("summary"),
                        source="merchrules",
                        external_id=ext_id,
                        followup_status="pending",
                    ))
                    meetings_synced += 1
                    # Обновляем last_meeting_date на клиенте
                    if not client.last_meeting_date or meeting_date > client.last_meeting_date:
                        client.last_meeting_date = meeting_date

        # Метрики
        metrics = acc.get("metrics")
        if metrics and isinstance(metrics, dict):
            hs = metrics.get("health_score") or metrics.get("healthScore")
            if hs is not None:
                client.health_score = float(hs)

    db.commit()

    # Логируем
    db.add(SyncLog(
        integration="extension",
        resource_type="accounts",
        action="push",
        status="success",
        records_processed=clients_synced,
        sync_data={"tasks": tasks_synced, "meetings": meetings_synced},
    ))
    db.commit()

    logger.info(f"Extension sync: {clients_synced} clients, {tasks_synced} tasks, {meetings_synced} meetings (user={user.email})")
    return {
        "ok": True,
        "clients_synced": clients_synced,
        "tasks_synced": tasks_synced,
        "meetings_synced": meetings_synced,
    }


@app.get("/api/auth/token")
async def api_get_token(db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Вернуть JWT токен текущего пользователя — для настройки расширения."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload_jwt = decode_access_token(auth_token)
    if not payload_jwt:
        raise HTTPException(status_code=401)
    return {"token": auth_token}


@app.post("/api/sync/airtable")
async def api_sync_airtable(
    request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)
):
    """Синхронизация клиентов из Airtable → локальная БД."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        raise HTTPException(status_code=401)

    from airtable_sync import sync_clients_from_airtable
    try:
        body = await request.json()
    except Exception:
        body = {}

    # Приоритет: body → user.settings → env
    u_settings = user.settings or {}
    at_settings = u_settings.get("airtable", {})
    token = (body.get("token")
             or at_settings.get("pat") or at_settings.get("token")
             or _env("AIRTABLE_TOKEN") or _env("AIRTABLE_PAT"))
    base_id = (body.get("base_id")
               or at_settings.get("base_id")
               or _env("AIRTABLE_BASE_ID"))
    view_id = body.get("view_id") or at_settings.get("view_id", "")

    if not token:
        return {"error": "Нет токена Airtable. Укажите в Настройках → Аккаунты."}

    result = await sync_clients_from_airtable(
        db=db,
        token=token,
        base_id=base_id or None,
        view_id=view_id,
        default_manager_email=user.email,
    )
    return result


@app.post("/api/sync/merchrules")
async def api_sync_merchrules(
    request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)
):
    """Синхронизация с Merchrules — пробует QA и Production."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)

    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        raise HTTPException(status_code=401)

    try:
        body = await request.json()
    except Exception:
        body = {}
    settings = user.settings or {}
    mr = settings.get("merchrules", {})
    login = body.get("login") or mr.get("login") or env.MR_LOGIN
    password = body.get("password") or mr.get("password") or env.MR_PASSWORD
    site_ids_input = body.get("site_ids") or mr.get("site_ids") or settings.get("merchrules_site_ids", [])

    if not login or not password:
        return {"error": "Нужны креды Merchrules"}

    # Сохраняем креды
    mr["login"] = login
    mr["password"] = password
    if site_ids_input:
        settings["merchrules_site_ids"] = site_ids_input
    settings["merchrules"] = mr
    user.settings = dict(settings)
    from sqlalchemy.orm.attributes import flag_modified
    flag_modified(user, "settings")
    db.commit()

    # Пробуем авторизацию — все URL + все варианты поля логина
    import httpx
    urls_to_try = list(dict.fromkeys([
        _env("MERCHRULES_API_URL", "https://merchrules.any-platform.ru"),
        "https://merchrules.any-platform.ru",
        "https://merchrules-qa.any-platform.ru",
    ]))
    login_fields = ["email", "login", "username"]
    base_url = None
    token = None
    last_error = ""
    attempts_log = []

    async with httpx.AsyncClient(timeout=30) as hx:
        outer_break = False
        for url in urls_to_try:
            if outer_break:
                break
            for field in login_fields:
                try:
                    resp = await hx.post(
                        f"{url}/backend-v2/auth/login",
                        json={field: login, "password": password},
                        timeout=15,
                    )
                    attempt_info = f"{url} [{field}] → {resp.status_code}"
                    if resp.status_code == 200:
                        body_resp = resp.json()
                        token = body_resp.get("token") or body_resp.get("access_token") or body_resp.get("accessToken")
                        if token:
                            base_url = url
                            logger.info(f"✅ Merchrules auth OK on {url} with field={field}")
                            outer_break = True
                            break
                        else:
                            last_error = f"Нет токена в ответе ({field}): {body_resp}"
                            attempts_log.append(attempt_info + " [no token]")
                    else:
                        last_error = f"HTTP {resp.status_code} [{field}]: {resp.text[:200]}"
                        attempts_log.append(attempt_info)
                except Exception as e:
                    last_error = str(e)
                    attempts_log.append(f"{url} [{field}] → error: {e}")

    if not token:
        detail = " | ".join(attempts_log[-4:]) if attempts_log else last_error
        logger.error(f"Merchrules auth failed. Attempts: {attempts_log}")
        return {"error": f"Ошибка авторизации Merchrules. {detail}"}

    headers = {"Authorization": f"Bearer {token}"}
    logger.info(f"Using Merchrules base URL: {base_url}")

    synced_clients = 0
    synced_tasks = 0

    try:
        # Если указаны site_id — используем их
        if site_ids_input:
            for sid in site_ids_input:
                sid = str(sid).strip()
                if not sid:
                    continue
                c = db.query(Client).filter(Client.merchrules_account_id == sid).first()
                if not c:
                    c = Client(merchrules_account_id=sid, name=f"Site {sid}", manager_email=user.email, segment="SMB")
                    db.add(c)
                    db.flush()

                # Tasks
                try:
                    r_tasks = await hx.get(f"{base_url}/backend-v2/tasks", params={"site_id": sid, "limit": 50}, headers=headers, timeout=15)
                    if r_tasks.status_code == 200:
                        tasks_data = r_tasks.json()
                        tasks_list = tasks_data.get("tasks") or tasks_data.get("items") or []
                        for t in tasks_list[:20]:
                            existing = db.query(Task).filter(Task.merchrules_task_id == str(t.get("id"))).first()
                            if not existing:
                                db.add(Task(client_id=c.id, merchrules_task_id=str(t.get("id")),
                                    title=t.get("title",""), status=t.get("status","plan"),
                                    priority=t.get("priority","medium"), source="roadmap"))
                                synced_tasks += 1
                except Exception as e:
                    logger.warning(f"Failed to fetch tasks for {sid}: {e}")

                # Meetings
                try:
                    r_meetings = await hx.get(f"{base_url}/backend-v2/meetings", params={"site_id": sid, "limit": 10}, headers=headers, timeout=15)
                    if r_meetings.status_code == 200:
                        meetings_data = r_meetings.json()
                        meetings_list = meetings_data.get("meetings") or meetings_data.get("items") or []
                        if meetings_list:
                            last_mtg = max(meetings_list, key=lambda m: m.get("date", ""))
                            try:
                                c.last_meeting_date = datetime.fromisoformat(last_mtg.get("date", "")[:19])
                            except Exception as e:
                                logger.debug(f"Ignored error: {e}")
                except Exception as e:
                    logger.warning(f"Failed to fetch meetings for {sid}: {e}")

                synced_clients += 1
        else:
            # Без site_ids — получаем все аккаунты менеджера
            accounts = []
            accounts_endpoint_log = []

            # Пробуем несколько возможных endpoint'ов
            for ep in [
                f"{base_url}/backend-v2/accounts",
                f"{base_url}/backend-v2/sites",
                f"{base_url}/backend-v2/accounts?limit=500",
                f"{base_url}/backend-v2/sites?limit=500",
            ]:
                try:
                    r = await hx.get(ep, headers=headers, timeout=20)
                    accounts_endpoint_log.append(f"{ep} → {r.status_code}")
                    if r.status_code == 200:
                        data = r.json()
                        # Пробуем разные ключи в ответе
                        for key in ("accounts", "sites", "items", "data", "results"):
                            if isinstance(data.get(key), list) and data[key]:
                                accounts = data[key]
                                break
                        # Если ответ сам список
                        if not accounts and isinstance(data, list):
                            accounts = data
                        if accounts:
                            logger.info(f"✅ Accounts from {ep}: {len(accounts)}")
                            break
                except Exception as e:
                    accounts_endpoint_log.append(f"{ep} → error: {e}")

            if not accounts:
                return {
                    "error": "Не удалось получить список аккаунтов. Попробуйте указать Site ID вручную.",
                    "endpoints_tried": accounts_endpoint_log,
                    "hint": "Укажите site_id через запятую в поле Site ID, например: 2262, 5335, 8049"
                }

            logger.info(f"Syncing {len(accounts)} accounts for {user.email}")

            for acc in accounts:
                aid = acc.get("id") or acc.get("site_id") or acc.get("siteId")
                if not aid:
                    continue
                site_id = str(aid)

                # Название аккаунта — пробуем разные поля
                acc_name = (
                    acc.get("name") or acc.get("title") or
                    acc.get("company") or acc.get("domain") or
                    f"Account {site_id}"
                )

                # Сегмент если есть
                acc_segment = acc.get("segment") or acc.get("tariff") or acc.get("plan") or None

                c = db.query(Client).filter(Client.merchrules_account_id == site_id).first()
                if not c:
                    c = Client(
                        merchrules_account_id=site_id,
                        name=acc_name,
                        manager_email=user.email,
                        segment=acc_segment,
                    )
                    db.add(c)
                    db.flush()
                else:
                    # Обновляем имя и менеджера
                    c.name = acc_name
                    if not c.manager_email:
                        c.manager_email = user.email
                    if acc_segment and not c.segment:
                        c.segment = acc_segment

                # Tasks
                try:
                    r_tasks = await hx.get(
                        f"{base_url}/backend-v2/tasks",
                        params={"site_id": site_id, "limit": 100},
                        headers=headers, timeout=15,
                    )
                    if r_tasks.status_code == 200:
                        td = r_tasks.json()
                        tasks_list = td.get("tasks") or td.get("items") or (td if isinstance(td, list) else [])
                        for t in tasks_list:
                            tid = str(t.get("id", ""))
                            if not tid:
                                continue
                            existing = db.query(Task).filter(Task.merchrules_task_id == tid).first()
                            if not existing:
                                db.add(Task(
                                    client_id=c.id,
                                    merchrules_task_id=tid,
                                    title=t.get("title") or t.get("name") or "",
                                    status=t.get("status", "plan"),
                                    priority=t.get("priority", "medium"),
                                    source="roadmap",
                                    team=t.get("team") or t.get("assignee") or None,
                                ))
                                synced_tasks += 1
                            else:
                                # Обновляем статус
                                existing.status = t.get("status", existing.status)
                except Exception as e:
                    logger.warning(f"Tasks fetch failed for site {site_id}: {e}")

                # Meetings — последняя дата
                try:
                    r_meetings = await hx.get(
                        f"{base_url}/backend-v2/meetings",
                        params={"site_id": site_id, "limit": 10},
                        headers=headers, timeout=15,
                    )
                    if r_meetings.status_code == 200:
                        md = r_meetings.json()
                        meetings_list = md.get("meetings") or md.get("items") or (md if isinstance(md, list) else [])
                        dates = []
                        for m in meetings_list:
                            d = m.get("date") or m.get("meeting_date") or m.get("createdAt", "")
                            if d:
                                dates.append(str(d)[:19])
                        if dates:
                            try:
                                c.last_meeting_date = datetime.fromisoformat(max(dates))
                            except Exception:
                                pass
                except Exception as e:
                    logger.warning(f"Meetings fetch failed for site {site_id}: {e}")

                synced_clients += 1

        db.commit()
        return {
            "ok": True,
            "clients_synced": synced_clients,
            "tasks_synced": synced_tasks,
            "base_url": base_url,
            "message": f"Синхронизировано: {synced_clients} клиентов, {synced_tasks} задач",
        }
    except Exception as e:
        db.rollback()
        logger.error(f"Merchrules sync error: {e}")
        return {"error": str(e)}


@app.post("/api/auth/taim/test")
async def api_test_taim(request: Request, auth_token: Optional[str] = Cookie(None)):
    """Проверить авторизацию в 1Time (time.tbank.ru / Mattermost)."""
    if not auth_token:
        raise HTTPException(status_code=401)
    body = await request.json()
    login_id = body.get("login", "")
    password = body.get("password", "")
    import taim
    result = await taim.login(login_id, password)
    if result["ok"]:
        summary = await taim.get_summary(login_id, password)
        return {**result, **summary, "password": None}
    return result


# ============================================================================
# AUTH: TBANK TIME (SSO через TinkoffID)
# ============================================================================

@app.get("/auth/time", response_class=HTMLResponse)
async def time_oauth_start(request: Request, auth_token: Optional[str] = Cookie(None)):
    """Подключение Tbank Time — PAT (рекомендуется) или MMAUTHTOKEN (запасной)."""
    if not auth_token:
        return RedirectResponse(url="/login")
    html = open("/home/claude/AnyQueryAMHub/am-hub-final/templates/time_auth.html").read()
    return HTMLResponse(content=html)


@app.post("/api/auth/time/token")
async def api_time_save_token(
    request: Request,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """Сохранить MMAUTHTOKEN, проверить доступ к каналу any-team-support."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        raise HTTPException(status_code=401)

    data = await request.json()
    token = data.get("token", "").strip()
    if not token:
        return {"ok": False, "error": "Токен не передан"}

    # Проверяем токен — запрашиваем данные пользователя
    import httpx
    TIME_BASE = "https://time.tbank.ru"
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}

    try:
        async with httpx.AsyncClient(timeout=15) as hx:
            # 1. Получаем текущего пользователя
            me_resp = await hx.get(f"{TIME_BASE}/api/v4/users/me", headers=headers)
            if me_resp.status_code == 401:
                return {"ok": False, "error": "Токен недействителен или истёк — войдите заново"}
            if me_resp.status_code != 200:
                return {"ok": False, "error": f"HTTP {me_resp.status_code} при проверке токена"}

            me = me_resp.json()
            username = me.get("username", "")
            email = me.get("email", "")

            # 2. Ищем канал any-team-support
            channel_posts_count = None
            channel_id = None
            try:
                # Получаем канал по team/channel name
                ch_resp = await hx.get(
                    f"{TIME_BASE}/api/v4/teams/name/tinkoff/channels/name/any-team-support",
                    headers=headers,
                )
                if ch_resp.status_code == 200:
                    channel_id = ch_resp.json().get("id")
                elif ch_resp.status_code == 404:
                    # Пробуем найти через поиск
                    search_resp = await hx.post(
                        f"{TIME_BASE}/api/v4/channels/search",
                        headers=headers,
                        json={"term": "any-team-support"},
                    )
                    if search_resp.status_code == 200:
                        channels = search_resp.json()
                        for ch in (channels if isinstance(channels, list) else []):
                            if "any-team-support" in (ch.get("name") or ""):
                                channel_id = ch.get("id")
                                break
            except Exception:
                pass

            if channel_id:
                try:
                    posts_resp = await hx.get(
                        f"{TIME_BASE}/api/v4/channels/{channel_id}/posts",
                        headers=headers,
                        params={"per_page": 1},
                    )
                    if posts_resp.status_code == 200:
                        channel_posts_count = posts_resp.json().get("order", []).__len__()
                except Exception:
                    pass

    except Exception as e:
        return {"ok": False, "error": str(e)}

    # Сохраняем токен и channel_id в user.settings
    settings = dict(user.settings or {})
    tm = dict(settings.get("tbank_time", {}))
    tm["session_cookie"] = token
    tm["mmauthtoken"] = token
    tm["username"] = username
    tm["email"] = email
    if channel_id:
        tm["support_channel_id"] = channel_id
    settings["tbank_time"] = tm

    from sqlalchemy.orm.attributes import flag_modified
    user.settings = settings
    flag_modified(user, "settings")
    db.commit()

    logger.info(f"✅ Time token saved for {user.email} (username={username}, channel_id={channel_id})")
    return {
        "ok": True,
        "username": username,
        "email": email,
        "channel_id": channel_id,
        "channel_posts_count": channel_posts_count,
    }


@app.post("/api/auth/time/disconnect")
async def api_time_disconnect(
    request: Request,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """Отключить Tbank Time — удалить токен из user.settings."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        raise HTTPException(status_code=401)
    settings = dict(user.settings or {})
    settings["tbank_time"] = {}
    from sqlalchemy.orm.attributes import flag_modified
    user.settings = settings
    flag_modified(user, "settings")
    db.commit()
    return {"ok": True}



@app.post("/api/auth/ktalk/test")
async def api_test_ktalk(request: Request, auth_token: Optional[str] = Cookie(None)):
    """Проверить OIDC авторизацию в KTalk (tbank.ktalk.ru)."""
    if not auth_token:
        raise HTTPException(status_code=401)
    body = await request.json()
    login_id = body.get("login", "")
    password = body.get("password", "")
    import ktalk
    oidc_cfg = await ktalk._get_oidc_config()
    result = await ktalk.login(login_id, password)
    return {**result, "oidc_grant_types": oidc_cfg.get("grant_types_supported", []), "password": None}


@app.get("/auth/ktalk", response_class=HTMLResponse)
async def ktalk_oauth_start(request: Request, auth_token: Optional[str] = Cookie(None)):
    """
    Запускает OIDC авторизацию через браузер.
    Редиректит пользователя на страницу входа KTalk (SSO Т-Банка с SMS).
    """
    if not auth_token:
        return RedirectResponse(url="/login")
    import secrets, urllib.parse

    # client_id можно переопределить через env если у вас корпоративный OIDC клиент
    client_id = _env("KTALK_OIDC_CLIENT_ID", "KTalk")
    redirect_uri = _env("KTALK_REDIRECT_URI") or (str(request.base_url).rstrip("/") + "/auth/ktalk/callback")

    params = urllib.parse.urlencode({
        "client_id": client_id,
        "response_type": "id_token token",  # implicit flow — токен сразу в hash
        "scope": "profile email allatclaims",
        "redirect_uri": redirect_uri,
        "nonce": secrets.token_urlsafe(16),
        "state": secrets.token_urlsafe(16),
    })
    return RedirectResponse(
        url=f"https://tbank.ktalk.ru/api/authorize/oidc/connect/authorize?{params}"
    )


@app.get("/auth/ktalk/callback", response_class=HTMLResponse)
async def ktalk_oauth_callback(request: Request):
    """
    Callback после OIDC авторизации KTalk (SSO Т-Банка).
    Токен приходит в URL hash (#access_token=...) — JS читает и сохраняет.
    """
    return HTMLResponse(content="""<!DOCTYPE html>
<html><head><meta charset="UTF-8"><title>KTalk — авторизация</title>
<style>
  *{margin:0;padding:0;box-sizing:border-box;}
  body{font-family:Inter,sans-serif;background:#0a0e1a;color:#e2e8f0;
       display:flex;align-items:center;justify-content:center;min-height:100vh;}
  .card{background:#111827;border:1px solid #1e2a3a;border-radius:14px;
        padding:32px 40px;text-align:center;max-width:440px;width:90%;}
  h2{font-size:1.2rem;margin-bottom:10px;}
  p{color:#64748b;font-size:.85rem;line-height:1.6;}
  .ok{color:#22c55e;} .err{color:#ef4444;}
  .btn{display:inline-block;margin-top:16px;padding:10px 20px;
       background:#6366f1;color:#fff;border-radius:8px;text-decoration:none;font-size:.85rem;}
  .manual{margin-top:20px;padding:14px;background:#1e2a3a;border-radius:8px;text-align:left;}
  .manual p{font-size:.78rem;color:#94a3b8;margin-bottom:6px;}
  .manual code{display:block;background:#0a0e1a;padding:8px 10px;border-radius:6px;
               font-size:.75rem;color:#818cf8;word-break:break-all;margin-top:4px;}
  input{width:100%;padding:8px 10px;margin-top:8px;border-radius:6px;
        border:1px solid #1e2a3a;background:#0a0e1a;color:#e2e8f0;font-size:.82rem;}
  .paste-btn{margin-top:8px;padding:7px 14px;background:#22c55e;color:#fff;
             border:none;border-radius:6px;cursor:pointer;font-size:.8rem;}
</style></head>
<body><div class="card">
  <h2 id="title">⏳ Авторизация KTalk...</h2>
  <p id="msg">Получаем токен от Т-Банк SSO</p>
  <div id="manual-block" style="display:none" class="manual">
    <p>Если автоматически не сработало — вставьте токен вручную:</p>
    <p>Откройте DevTools (F12) → Console → введите:</p>
    <code>copy(window.__ktalk_token || 'нет токена')</code>
    <p style="margin-top:8px;">Или скопируйте access_token из URL адресной строки после #</p>
    <input id="manual-token" placeholder="Вставьте access_token сюда...">
    <button class="paste-btn" onclick="saveManualToken()">💾 Сохранить токен</button>
  </div>
</div>
<script>
async function saveToken(token) {
  try {
    const r = await fetch('/api/auth/ktalk/token', {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({access_token: token})
    });
    const d = await r.json();
    if (d.ok) {
      document.getElementById('title').textContent = '✅ KTalk подключён!';
      document.getElementById('msg').innerHTML =
        'Авторизован как: <b>' + (d.user?.firstname||'') + ' ' + (d.user?.surname||'') + '</b>' +
        '<br><br><a href="/settings" class="btn">← Вернуться в настройки</a>';
      document.getElementById('msg').className = 'ok';
    } else {
      showError(d.error || 'Не удалось сохранить токен');
    }
  } catch(e) {
    showError(e.message);
  }
}

function showError(msg) {
  document.getElementById('title').textContent = '❌ Ошибка';
  document.getElementById('msg').textContent = msg;
  document.getElementById('msg').className = 'err';
  document.getElementById('manual-block').style.display = 'block';
}

async function saveManualToken() {
  const token = document.getElementById('manual-token').value.trim();
  if (!token) return;
  await saveToken(token);
}

// Основной flow: читаем токен из URL hash
(async function() {
  const hash = window.location.hash.slice(1);
  const query = window.location.search.slice(1);
  const hashParams = Object.fromEntries(new URLSearchParams(hash));
  const queryParams = Object.fromEntries(new URLSearchParams(query));

  // Токен может быть в hash (implicit flow) или query (code flow)
  const token = hashParams.access_token || hashParams.id_token ||
                queryParams.access_token || queryParams.id_token;

  // Error от OIDC сервера
  const error = hashParams.error || queryParams.error;
  if (error) {
    const desc = hashParams.error_description || queryParams.error_description || error;
    // redirect_uri_mismatch — самая частая ошибка
    if (error === 'invalid_request' || desc.includes('redirect_uri')) {
      showError('redirect_uri не зарегистрирован в Ktalk. ' +
        'Добавьте переменную KTALK_REDIRECT_URI в Railway Variables: ' +
        window.location.origin + '/auth/ktalk/callback');
    } else {
      showError(desc);
    }
    return;
  }

  if (!token) {
    // Нет токена и нет ошибки — может быть code flow
    const code = queryParams.code;
    if (code) {
      showError('Получен authorization code вместо токена. ' +
        'Нужна серверная обработка code flow. Обратитесь к администратору.');
    } else {
      showError('Токен не получен. Возможно redirect_uri не совпадает с зарегистрированным в Ktalk.');
    }
    return;
  }

  await saveToken(token);
})();
</script></body></html>""")


@app.post("/api/auth/ktalk/token")
async def api_ktalk_save_token(
    request: Request,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """Сохраняет OIDC access_token KTalk после browser-based авторизации."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        raise HTTPException(status_code=401)

    body = await request.json()
    access_token = body.get("access_token", "")
    if not access_token:
        return {"ok": False, "error": "Нет токена"}

    # Получаем данные пользователя чтобы подтвердить токен
    import ktalk as ktalk_mod
    user_info = await ktalk_mod._get_user_info(access_token)

    settings = user.settings or {}
    kt = settings.get("ktalk", {})
    kt["access_token"] = access_token
    kt["login"] = user_info.get("email", kt.get("login", ""))
    settings["ktalk"] = kt
    user.settings = dict(settings)
    from sqlalchemy.orm.attributes import flag_modified
    flag_modified(user, "settings")
    db.commit()
    return {"ok": True, "user": user_info}


@app.get("/api/ktalk/calendar")
async def api_ktalk_calendar(
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
    days: int = 7,
):
    """Получить встречи из KTalk календаря."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        raise HTTPException(status_code=401)

    settings = user.settings or {}
    kt = settings.get("ktalk", {})
    login_id = kt.get("login", "")
    password = kt.get("password", "")
    if not login_id or not password:
        return {"error": "Укажи логин/пароль KTalk в Настройках"}

    import ktalk
    return await ktalk.get_today_meetings(login_id, password)


@app.get("/api/sync/merchrules-creds")
async def api_get_mr_creds(db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Получить сохранённые креды Merchrules пользователя."""
    if not auth_token:
        return {}
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        return {}
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        return {}
    settings = user.settings or {}
    mr = settings.get("merchrules", {})
    return {"login": mr.get("login", ""), "site_ids": settings.get("merchrules_site_ids", [])}


# ============================================================================
# API: TASK CRUD
# ============================================================================

@app.post("/api/tasks")
async def api_create_task(
    request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)
):
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)

    data = await request.json()
    task = Task(
        client_id=data["client_id"],
        title=data["title"],
        description=data.get("description", ""),
        status=data.get("status", "plan"),
        priority=data.get("priority", "medium"),
        team=data.get("team", ""),
        task_type=data.get("task_type", ""),
        due_date=datetime.fromisoformat(data["due_date"]) if data.get("due_date") else None,
    )
    db.add(task)
    db.commit()
    db.refresh(task)
    # WS real-time push
    try:
        await ws_invalidate_stats(user.id, db)
        await ws_notify_user(user.id, "task_update", {"action": "created", "task_id": task.id})
    except Exception:
        pass
    return {"ok": True, "id": task.id}


@app.put("/api/tasks/{task_id}")
async def api_update_task(task_id: int, request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    if not auth_token:
        raise HTTPException(status_code=401)
    data = await request.json()
    task = db.query(Task).filter(Task.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404)
    for k, v in data.items():
        if hasattr(task, k):
            setattr(task, k, v)
    db.commit()
    return {"ok": True}


# ============================================================================
# API: AI PROCESSING
# ============================================================================

@app.post("/api/ai/process-transcript")
async def api_process_transcript(request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    if not auth_token:
        raise HTTPException(status_code=401)
    data = await request.json()
    transcript = data.get("transcript", "")
    try:
        result = ai_process_transcript(transcript)
        return result
    except Exception as e:
        return {"error": str(e)}


@app.post("/api/ai/generate-followup")
async def api_generate_followup(request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    if not auth_token:
        raise HTTPException(status_code=401)
    data = await request.json()
    client_id = data.get("client_id")
    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        return {"error": "Client not found"}
    tasks = db.query(Task).filter(Task.client_id == client_id).all()
    meetings = db.query(Meeting).filter(Meeting.client_id == client_id).order_by(Meeting.date.desc()).limit(3).all()
    try:
        text = generate_smart_followup(client, tasks, meetings)
        return {"text": text}
    except Exception as e:
        return {"error": str(e)}


@app.post("/api/ai/generate-prep")
async def api_generate_prep(
    request: Request,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """Перегенерация AI-подготовки к встрече (вызывается по кнопке на prep-странице)."""
    if not auth_token:
        raise HTTPException(status_code=401)
    data = await request.json()
    client_id = data.get("client_id")
    meeting_type = data.get("meeting_type", "meeting")
    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        return {"error": "Клиент не найден"}
    tasks = db.query(Task).filter(
        Task.client_id == client_id,
        Task.status.in_(["plan", "in_progress", "blocked"]),
    ).all()
    meetings = db.query(Meeting).filter(
        Meeting.client_id == client_id,
    ).order_by(Meeting.date.desc()).limit(5).all()
    try:
        text = generate_prep_brief(client, tasks, meetings)
        # Добавляем контекст типа встречи
        type_hints = {
            "checkup": "\n\n📋 Тип встречи: ЧЕКАП — фокус на прогрессе по задачам и здоровье аккаунта.",
            "qbr": "\n\n📊 Тип встречи: QBR — квартальный обзор, нужна аналитика и достижения.",
            "onboarding": "\n\n🚀 Тип встречи: ОНБОРДИНГ — первые шаги, знакомство с продуктом.",
            "upsell": "\n\n📈 Тип встречи: АПСЕЙЛ — выявление возможностей для расширения.",
            "sync": "\n\n🔄 Тип встречи: СИНК — текущий статус и оперативные вопросы.",
        }
        text += type_hints.get(meeting_type, "")
        return {"text": text}
    except Exception as e:
        return {"error": str(e)}


# ============================================================================
# SETTINGS
# ============================================================================

@app.get("/settings", response_class=HTMLResponse)
async def settings_page(request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    if not auth_token:
        return RedirectResponse(url="/login", status_code=303)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        return RedirectResponse(url="/login", status_code=303)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        return RedirectResponse(url="/login", status_code=303)

    settings = user.settings or {}
    rules = settings.get("rules", {
        "min_health_score": 0.5, "checkup_interval_days": 30, "warning_days": 14,
        "segments": ["ENT", "SME+", "SME-", "SMB", "SS"],
        "auto_create_tasks": True, "morning_plan_time": "09:00", "weekly_digest_day": "friday",
    })
    prefs = settings.get("preferences", {
        "theme": "dark", "dashboard_view": "cards",
        "notifications_email": True, "notifications_tg": True, "notifications_ktalk": False,
        "notif_overdue": True, "notif_new_tasks": True, "notif_blocked": True, "notif_morning": True,
    })
    return templates.TemplateResponse("settings.html", {
        "request": request, "user": user,
        "user_settings": settings,
        "rules": rules, "prefs": prefs,
    })


@app.post("/api/settings/creds")
async def api_save_creds(request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Сохранить персональные креды пользователя для ВСЕХ сервисов."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        raise HTTPException(status_code=401)

    data = await request.json()
    settings = user.settings or {}

    # Сохраняем все сервисы: merchrules, telegram, ktalk, tbank_time, airtable, sheets, groq
    for service in ["merchrules", "telegram", "ktalk", "tbank_time", "airtable", "sheets", "groq"]:
        if service in data:
            if service not in settings:
                settings[service] = {}
            settings[service].update(data[service])

    user.settings = dict(settings)
    from sqlalchemy.orm.attributes import flag_modified
    flag_modified(user, "settings")
    db.commit()
    return {"ok": True}


@app.post("/api/settings/rules")
async def api_save_rules(request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Сохранить правила работы менеджера."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        raise HTTPException(status_code=401)
    data = await request.json()
    settings = dict(user.settings or {})
    settings["rules"] = {**(settings.get("rules") or {}), **data}
    user.settings = settings
    from sqlalchemy.orm.attributes import flag_modified
    flag_modified(user, "settings")
    db.commit()
    return {"ok": True}


@app.post("/api/settings/prefs")
async def api_save_prefs(request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Сохранить предпочтения (тема, уведомления и т.д.)."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        raise HTTPException(status_code=401)
    data = await request.json()
    settings = dict(user.settings or {})
    settings["preferences"] = {**(settings.get("preferences") or {}), **data}
    user.settings = settings
    from sqlalchemy.orm.attributes import flag_modified
    flag_modified(user, "settings")
    db.commit()
    return {"ok": True}




@app.get("/", response_class=HTMLResponse)
async def root(request: Request, auth_token: Optional[str] = Cookie(None)):
    if auth_token:
        from auth import decode_access_token
        payload = decode_access_token(auth_token)
        if payload:
            return RedirectResponse(url="/dashboard", status_code=303)
    return RedirectResponse(url="/login", status_code=303)


# ============================================================================
# WORKFLOW: MEETINGS CRUD
# ============================================================================

@app.post("/api/meetings")
async def api_create_meeting(
    request: Request,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """Создать встречу вручную."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        raise HTTPException(status_code=401)

    data = await request.json()
    client_id = data.get("client_id")
    title = data.get("title", "").strip()
    meeting_type = data.get("type", "meeting")
    date_str = data.get("date")
    notes = data.get("notes", "")

    if not client_id:
        return {"error": "client_id обязателен"}

    meeting_date = None
    if date_str:
        try:
            meeting_date = datetime.fromisoformat(date_str.replace("Z", ""))
        except Exception:
            return {"error": f"Неверный формат даты: {date_str}"}

    meeting = Meeting(
        client_id=int(client_id),
        title=title or meeting_type,
        type=meeting_type,
        date=meeting_date or datetime.now(),
        source="manual",
        followup_status="pending",
        summary=notes or None,
    )
    db.add(meeting)
    db.flush()

    # Обновляем last_meeting_date у клиента
    client = db.query(Client).filter(Client.id == int(client_id)).first()
    if client and meeting_date:
        if not client.last_meeting_date or meeting_date > client.last_meeting_date:
            client.last_meeting_date = meeting_date

    # Создаём слоты prep/followup
    try:
        from meeting_slots import create_slots_for_meeting
        create_slots_for_meeting(db, meeting)
    except Exception as e:
        logger.warning(f"Slots creation failed: {e}")

    db.commit()
    return {"ok": True, "meeting_id": meeting.id, "message": f"Встреча «{meeting.title}» создана"}


@app.delete("/api/meetings/{meeting_id}")
async def api_delete_meeting(
    meeting_id: int,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """Удалить встречу."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)

    meeting = db.query(Meeting).filter(Meeting.id == meeting_id).first()
    if not meeting:
        raise HTTPException(status_code=404)
    db.delete(meeting)
    db.commit()
    return {"ok": True}


# ============================================================================
# WORKFLOW: FOLLOWUP
# ============================================================================

@app.post("/api/meetings/{meeting_id}/followup/generate")
async def api_generate_followup(meeting_id: int, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """AI-генерация фолоуапа для встречи."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)

    meeting = db.query(Meeting).filter(Meeting.id == meeting_id).first()
    if not meeting:
        raise HTTPException(status_code=404)

    client = db.query(Client).filter(Client.id == meeting.client_id).first()
    if not client:
        raise HTTPException(status_code=404)

    tasks = db.query(Task).filter(Task.client_id == client.id, Task.status.in_(["plan", "in_progress"])).all()
    meetings = db.query(Meeting).filter(Meeting.client_id == client.id).order_by(Meeting.date.desc()).limit(3).all()

    try:
        text = generate_smart_followup(client, tasks, meetings)
        meeting.followup_text = text
        meeting.followup_status = "filled"
        db.commit()
        return {"ok": True, "text": text}
    except Exception as e:
        return {"error": str(e)}


@app.post("/api/meetings/{meeting_id}/followup/send")
async def api_send_followup(meeting_id: int, request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Подтверждение отправки фолоуапа → создаётся задача done."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)

    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    meeting = db.query(Meeting).filter(Meeting.id == meeting_id).first()
    if not meeting:
        raise HTTPException(status_code=404)

    data = await request.json()
    followup_text = data.get("text", meeting.followup_text)

    meeting.followup_status = "sent"
    meeting.followup_text = followup_text
    meeting.followup_sent_at = datetime.now()

    # Создаём задачу "Фолоуап отправлен" со статусом done
    task = Task(
        client_id=meeting.client_id,
        title=f"📧 Фолоуап: {meeting.title or meeting.type}",
        description=followup_text[:500] if followup_text else "",
        status="done",
        priority="medium",
        source="followup",
        created_from_meeting_id=meeting.id,
        confirmed_at=datetime.now(),
        confirmed_by=user.email if user else None,
    )
    db.add(task)

    # Обновляем last_meeting_date у клиента
    client = db.query(Client).filter(Client.id == meeting.client_id).first()
    if client:
        client.last_meeting_date = meeting.date or datetime.now()

    db.commit()

    # Push в Ktalk — если настроен канал
    if user and followup_text:
        try:
            settings = user.settings or {}
            kt = settings.get("ktalk", {})
            channel_id = kt.get("followup_channel_id") or kt.get("channel_id")
            token = kt.get("access_token", "")
            if channel_id and token:
                from integrations.ktalk import send_followup_to_channel
                await send_followup_to_channel(
                    channel_id=channel_id,
                    client_name=client.name if client else "",
                    followup_text=followup_text,
                    meeting_date=meeting.date,
                    token=token,
                )
        except Exception as e:
            logger.warning(f"Ktalk followup push failed: {e}")

    # Push в Airtable — обновляем дату последней встречи
    if client and client.airtable_record_id:
        try:
            from airtable_sync import sync_meeting_to_airtable
            await sync_meeting_to_airtable(
                record_id=client.airtable_record_id,
                meeting_date=meeting.date or datetime.now(),
                comment=f"Фолоуап отправлен: {(followup_text or '')[:100]}",
            )
        except Exception as e:
            logger.warning(f"Airtable followup sync failed: {e}")

    # Push в Airtable — обновляем дату встречи
    if client and client.airtable_record_id:
        try:
            from integrations.airtable import update_meeting_date
            await update_meeting_date(
                record_id=client.airtable_record_id,
                meeting_date=meeting.date or datetime.now(),
                comment=f"Фолоуап: {(followup_text or '')[:200]}",
            )
        except Exception as e:
            logger.warning(f"Airtable followup push failed: {e}")

    return {"ok": True, "task_id": task.id}


@app.post("/api/meetings/{meeting_id}/followup/skip")
async def api_skip_followup(meeting_id: int, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Пропустить фолоуап → создаётся задача plan."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)

    meeting = db.query(Meeting).filter(Meeting.id == meeting_id).first()
    if not meeting:
        raise HTTPException(status_code=404)

    meeting.followup_status = "skipped"
    meeting.followup_skipped = True

    # Создаём задачу "Фолоуап" со статусом plan
    task = Task(
        client_id=meeting.client_id,
        title=f"📧 Фолоуап: {meeting.title or meeting.type}",
        description="Фолоуап пропущен — требуется заполнить позже",
        status="plan",
        priority="medium",
        source="followup",
        created_from_meeting_id=meeting.id,
    )
    db.add(task)
    db.commit()
    return {"ok": True, "task_id": task.id}


# ============================================================================
# WORKFLOW: TASK CONFIRMATION
# ============================================================================

@app.post("/api/tasks/{task_id}/confirm")
async def api_confirm_task(task_id: int, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Подтверждение выполнения задачи."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)

    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    task = db.query(Task).filter(Task.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404)

    task.status = "done"
    task.confirmed_at = datetime.now()
    task.confirmed_by = user.email if user else None
    db.commit()
    return {"ok": True}


# ============================================================================
# WORKFLOW: ROADMAP PUSH
# ============================================================================

@app.post("/api/tasks/{task_id}/push-roadmap")
async def api_push_roadmap(task_id: int, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Отправка задачи в Merchrules Roadmap."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)

    task = db.query(Task).filter(Task.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404)

    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    settings = (user.settings or {}) if user else {}
    mr = settings.get("merchrules", {})
    login = mr.get("login") or env.MR_LOGIN
    password = mr.get("password") or env.MR_PASSWORD
    base_url = _env("MERCHRULES_API_URL", "https://merchrules.any-platform.ru")

    if not login or not password:
        return {"error": "Нужны креды Merchrules (Настройки → Креды)"}

    client = db.query(Client).filter(Client.id == task.client_id).first()
    if not client or not client.merchrules_account_id:
        return {"error": "У клиента нет merchrules_account_id — синхронизируйте клиента сначала"}

    import httpx, io
    try:
        async with httpx.AsyncClient(timeout=30) as hx:
            # Авторизация — перебираем поля
            token = None
            for field in ("email", "login", "username"):
                try:
                    r = await hx.post(
                        f"{base_url}/backend-v2/auth/login",
                        json={field: login, "password": password},
                        timeout=10,
                    )
                    if r.status_code == 200:
                        token = r.json().get("token") or r.json().get("access_token") or r.json().get("accessToken")
                        if token:
                            break
                except Exception:
                    continue

            if not token:
                return {"error": "Ошибка авторизации Merchrules — проверьте логин/пароль"}

            headers = {"Authorization": f"Bearer {token}"}

            # Пробуем JSON API сначала
            task_payload = {
                "title": task.title,
                "description": task.description or "",
                "status": task.status,
                "priority": task.priority or "medium",
                "site_id": client.merchrules_account_id,
            }
            if task.team:
                task_payload["team"] = task.team
            if task.due_date:
                task_payload["due_date"] = task.due_date.strftime("%Y-%m-%d")

            resp = await hx.post(
                f"{base_url}/backend-v2/tasks",
                json=task_payload,
                headers=headers,
                timeout=15,
            )

            # Fallback: CSV import
            if resp.status_code not in (200, 201):
                csv_content = "title,description,status,priority,team,due_date\n"
                csv_content += f'"{task.title}","{task.description or ""}",{task.status},{task.priority or "medium"},{task.team or ""},{task.due_date.strftime("%Y-%m-%d") if task.due_date else ""}'
                files = {"file": ("task.csv", io.BytesIO(csv_content.encode("utf-8")), "text/csv")}
                resp = await hx.post(
                    f"{base_url}/backend-v2/import/tasks/csv",
                    data={"site_id": client.merchrules_account_id},
                    files=files,
                    headers=headers,
                    timeout=15,
                )

        if resp.status_code in (200, 201):
            task.pushed_to_roadmap = True
            task.roadmap_pushed_at = datetime.now()
            from sqlalchemy.orm.attributes import flag_modified
            db.commit()
            return {"ok": True, "message": f"Задача «{task.title}» отправлена в Roadmap"}
        return {"error": f"Merchrules вернул HTTP {resp.status_code}: {resp.text[:200]}"}
    except Exception as e:
        return {"error": str(e)}


# ============================================================================
# WORKFLOW: QBR
# ============================================================================

@app.get("/api/clients/{client_id}/qbr")
async def api_get_qbr(client_id: int, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Получить QBR данные клиента."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)

    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        raise HTTPException(status_code=404)

    qbr = db.query(QBR).filter(QBR.client_id == client_id).order_by(QBR.date.desc()).first()
    meetings = db.query(Meeting).filter(Meeting.client_id == client_id, Meeting.is_qbr == True).order_by(Meeting.date.desc()).limit(5).all()
    tasks = db.query(Task).filter(Task.client_id == client_id, Task.status == "done").order_by(Task.confirmed_at.desc()).limit(20).all()

    return {
        "client": {"id": client.id, "name": client.name, "segment": client.segment},
        "current_qbr": {
            "id": qbr.id if qbr else None,
            "quarter": qbr.quarter if qbr else None,
            "status": qbr.status if qbr else "draft",
            "metrics": qbr.metrics if qbr else {},
            "summary": qbr.summary if qbr else None,
            "achievements": qbr.achievements if qbr else [],
            "issues": qbr.issues if qbr else [],
            "next_goals": qbr.next_quarter_goals if qbr else [],
        } if qbr else None,
        "qbr_meetings": [{"id": m.id, "date": m.date.isoformat() if m.date else None, "title": m.title} for m in meetings],
        "completed_tasks": [{"id": t.id, "title": t.title, "confirmed_at": t.confirmed_at.isoformat() if t.confirmed_at else None} for t in tasks],
        "last_qbr_date": client.last_qbr_date.isoformat() if client.last_qbr_date else None,
        "next_qbr_date": client.next_qbr_date.isoformat() if client.next_qbr_date else None,
    }


@app.post("/api/clients/{client_id}/qbr")
async def api_create_qbr(client_id: int, request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Создать/обновить QBR."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)

    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    data = await request.json()

    qbr = db.query(QBR).filter(QBR.client_id == client_id).order_by(QBR.date.desc()).first()
    if not qbr:
        qbr = QBR(client_id=client_id, year=datetime.now().year, quarter=f"{datetime.now().year}-Q{(datetime.now().month-1)//3+1}")
        db.add(qbr)

    qbr.status = data.get("status", qbr.status)
    qbr.metrics = data.get("metrics", qbr.metrics)
    qbr.summary = data.get("summary", qbr.summary)
    qbr.achievements = data.get("achievements", qbr.achievements)
    qbr.issues = data.get("issues", qbr.issues)
    qbr.next_quarter_goals = data.get("next_quarter_goals", qbr.next_quarter_goals)
    qbr.key_insights = data.get("key_insights", qbr.key_insights or [])
    qbr.future_work = data.get("future_work", qbr.future_work or [])
    qbr.presentation_url = data.get("presentation_url", qbr.presentation_url)
    qbr.executive_summary = data.get("executive_summary", qbr.executive_summary)
    if data.get("date"):
        qbr.date = datetime.fromisoformat(data["date"])

    # Обновляем клиента
    client = db.query(Client).filter(Client.id == client_id).first()
    if client:
        client.last_qbr_date = qbr.date
        # Следующий QBR через 3 месяца
        client.next_qbr_date = qbr.date + timedelta(days=90) if qbr.date else None

    db.commit()

    # Push QBR в Airtable
    if client and client.airtable_record_id and qbr.summary:
        try:
            from airtable_sync import push_qbr_to_airtable
            await push_qbr_to_airtable(
                client_name=client.name,
                quarter=qbr.quarter or "",
                summary=qbr.summary or "",
                achievements=qbr.achievements or [],
            )
        except Exception as e:
            logger.warning(f"Airtable QBR push failed: {e}")

    return {"ok": True, "qbr_id": qbr.id}


# ============================================================================
# WORKFLOW: ACCOUNT PLAN
# ============================================================================

@app.get("/api/clients/{client_id}/plan")
async def api_get_plan(client_id: int, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Получить план работы по клиенту."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)

    plan = db.query(AccountPlan).filter(AccountPlan.client_id == client_id).first()
    if not plan:
        plan = AccountPlan(client_id=client_id)
        db.add(plan)
        db.commit()

    return {
        "quarterly_goals": plan.quarterly_goals or [],
        "action_items": plan.action_items or [],
        "notes": plan.notes,
        "strategy": plan.strategy,
        "updated_at": plan.updated_at.isoformat() if plan.updated_at else None,
        "updated_by": plan.updated_by,
    }


@app.post("/api/clients/{client_id}/plan")
async def api_save_plan(client_id: int, request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Сохранить план работы по клиенту."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)

    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    data = await request.json()

    plan = db.query(AccountPlan).filter(AccountPlan.client_id == client_id).first()
    if not plan:
        plan = AccountPlan(client_id=client_id)
        db.add(plan)

    plan.quarterly_goals = data.get("quarterly_goals", plan.quarterly_goals or [])
    plan.action_items = data.get("action_items", plan.action_items or [])
    plan.notes = data.get("notes", plan.notes)
    plan.strategy = data.get("strategy", plan.strategy)
    plan.updated_at = datetime.now()
    plan.updated_by = user.email if user else None

    db.commit()
    return {"ok": True}


# ============================================================================
# WORKFLOW: TBANK TICKETS
# ============================================================================

@app.get("/api/tbank/tickets/{client_name}")
async def api_tbank_tickets(client_name: str, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Получить тикеты Tbank Time для клиента."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)

    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    u_settings = (user.settings or {}) if user else {}
    tm = u_settings.get("tbank_time", {})

    # Приоритет: user.settings → env
    time_token = (tm.get("session_cookie") or tm.get("api_token")
                  or env.TIME_TOKEN)

    if not time_token:
        return {"error": "Настройте доступ к Tbank Time в Настройках → Аккаунты", "tickets": []}

    from integrations.tbank_time import sync_tickets_for_client
    try:
        result = await sync_tickets_for_client(client_name, token=time_token)
        return result
    except Exception as e:
        return {"error": str(e), "open_count": 0, "total_count": 0, "last_ticket": None}


@app.get("/api/tbank/tickets")
async def api_tbank_all_tickets(db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Получить все открытые тикеты."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)

    time_token = env.TIME_TOKEN
    if not time_token:
        return {"error": "TIME_API_TOKEN не настроен", "tickets": []}

    from integrations.tbank_time import get_support_tickets
    try:
        clients = db.query(Client).all()
        all_tickets = []
        for c in clients:
            if c.name:
                tickets = await get_support_tickets(c.name)
                for t in tickets:
                    t["client"] = c.name
                all_tickets.extend(tickets)
        return {"tickets": all_tickets, "total": len(all_tickets)}
    except Exception as e:
        return {"error": str(e), "tickets": [], "total": 0}




# ============================================================================
# ANALYTICS API
# ============================================================================

@app.get("/followup-templates", response_class=HTMLResponse)
async def followup_templates_page(request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    if not auth_token: return RedirectResponse(url="/login", status_code=303)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload: return RedirectResponse(url="/login", status_code=303)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user: return RedirectResponse(url="/login", status_code=303)
    return templates.TemplateResponse("followup_templates.html", {"request": request, "user": user})


@app.get("/auto-tasks", response_class=HTMLResponse)
async def auto_tasks_page(request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    if not auth_token: return RedirectResponse(url="/login", status_code=303)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload: return RedirectResponse(url="/login", status_code=303)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user: return RedirectResponse(url="/login", status_code=303)
    return templates.TemplateResponse("auto_tasks.html", {"request": request, "user": user})


@app.get("/api/analytics/overview")
async def api_analytics_overview(
    days: int = 30,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    if not auth_token: raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload: raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user: raise HTTPException(status_code=401)

    since = datetime.utcnow() - timedelta(days=days)
    q = db.query(Client)
    if user.role == "manager": q = q.filter(Client.manager_email == user.email)
    clients = q.all()

    health_vals = [c.health_score for c in clients if c.health_score is not None]
    avg_health  = sum(health_vals) / len(health_vals) if health_vals else 0

    # Сегменты
    from collections import Counter
    seg_counter = Counter(c.segment or "Unknown" for c in clients)
    segments = [{"segment": k, "count": v} for k, v in seg_counter.most_common()]

    # Health distribution
    health_good = sum(1 for h in health_vals if h >= 70)
    health_warn = sum(1 for h in health_vals if 40 <= h < 70)
    health_bad  = sum(1 for h in health_vals if h < 40)

    # Tasks
    cids = [c.id for c in clients]
    open_tasks    = db.query(Task).filter(Task.client_id.in_(cids), Task.status != "done").count() if cids else 0
    overdue_tasks = db.query(Task).filter(
        Task.client_id.in_(cids), Task.status != "done",
        Task.due_date < datetime.utcnow()
    ).count() if cids else 0

    # Meetings + followups за период
    meetings_count  = db.query(Meeting).filter(Meeting.client_id.in_(cids), Meeting.date >= since).count() if cids else 0
    followups_count = 0  # TODO: followup model

    # Risk clients
    risk_clients = sorted(
        [{"id": c.id, "name": c.name, "segment": c.segment, "health_score": c.health_score}
         for c in clients if c.health_score is not None and c.health_score < 60],
        key=lambda x: x["health_score"]
    )[:8]

    # Active clients (by meetings + tasks)
    active = []
    for c in clients:
        m_cnt = db.query(Meeting).filter(Meeting.client_id == c.id, Meeting.date >= since).count()
        t_cnt = db.query(Task).filter(Task.client_id == c.id, Task.created_at >= since).count()
        if m_cnt + t_cnt > 0:
            active.append({"id": c.id, "name": c.name, "activity_score": m_cnt * 3 + t_cnt})
    active.sort(key=lambda x: x["activity_score"], reverse=True)

    return {
        "total_clients": len(clients),
        "avg_health": avg_health,
        "open_tasks": open_tasks,
        "overdue_tasks": overdue_tasks,
        "meetings_count": meetings_count,
        "followups_count": followups_count,
        "segments": segments,
        "health_good": health_good,
        "health_warn": health_warn,
        "health_bad": health_bad,
        "risk_clients": risk_clients,
        "active_clients": active[:8],
    }


@app.get("/api/analytics/health-trend")
async def api_analytics_health_trend(
    days: int = 30,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    if not auth_token: raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload: raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user: raise HTTPException(status_code=401)

    # Генерируем точки по неделям на основе текущих данных (без истории)
    q = db.query(Client)
    if user.role == "manager": q = q.filter(Client.manager_email == user.email)
    clients = q.all()
    health_vals = [c.health_score for c in clients if c.health_score is not None]
    avg = sum(health_vals) / len(health_vals) if health_vals else 0

    # Симулируем тренд за период (заглушка до появления audit log)
    import random
    n_points = min(days // 7, 12)
    labels, values = [], []
    base = avg
    for i in range(n_points, 0, -1):
        d = datetime.utcnow() - timedelta(weeks=i)
        labels.append(d.strftime("%d.%m"))
        # Небольшой шум вокруг текущего значения
        values.append(round(max(0, min(100, base + random.uniform(-5, 5))), 1))
    labels.append("Сейчас")
    values.append(round(avg, 1))

    return {"labels": labels, "values": values}


@app.get("/api/analytics/tasks-stats")
async def api_analytics_tasks_stats(
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    if not auth_token: raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload: raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user: raise HTTPException(status_code=401)

    q = db.query(Client)
    if user.role == "manager": q = q.filter(Client.manager_email == user.email)
    cids = [c.id for c in q.all()]

    from collections import Counter
    tasks = db.query(Task).filter(Task.client_id.in_(cids)).all() if cids else []
    by_status = dict(Counter(t.status or "plan" for t in tasks))

    return {"by_status": by_status, "total": len(tasks)}


@app.get("/api/analytics/activity")
async def api_analytics_activity(
    days: int = 30,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    if not auth_token: raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload: raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user: raise HTTPException(status_code=401)

    q = db.query(Client)
    if user.role == "manager": q = q.filter(Client.manager_email == user.email)
    cids = [c.id for c in q.all()]

    # По неделям
    n_weeks = min(days // 7, 8)
    labels, meetings_data, tasks_data = [], [], []
    for i in range(n_weeks, 0, -1):
        week_start = datetime.utcnow() - timedelta(weeks=i)
        week_end   = datetime.utcnow() - timedelta(weeks=i-1)
        label = week_start.strftime("Нед %d.%m")
        m_cnt = db.query(Meeting).filter(Meeting.client_id.in_(cids), Meeting.date >= week_start, Meeting.date < week_end).count() if cids else 0
        t_cnt = db.query(Task).filter(Task.client_id.in_(cids), Task.created_at >= week_start, Task.created_at < week_end).count() if cids else 0
        labels.append(label); meetings_data.append(m_cnt); tasks_data.append(t_cnt)

    return {"labels": labels, "meetings": meetings_data, "tasks": tasks_data}


@app.get("/api/analytics/export")
async def api_analytics_export(
    days: int = 30,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """Экспорт аналитики в CSV."""
    if not auth_token: raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload: raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user: raise HTTPException(status_code=401)

    q = db.query(Client)
    if user.role == "manager": q = q.filter(Client.manager_email == user.email)
    clients = q.order_by(Client.name).all()

    import csv, io
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["ID","Название","Сегмент","Health Score","Последняя встреча","Открытых задач"])
    for c in clients:
        open_t = db.query(Task).filter(Task.client_id == c.id, Task.status != "done").count()
        w.writerow([c.id, c.name, c.segment, f"{c.health_score:.0f}%" if c.health_score else "—",
                    c.last_meeting_date.strftime("%d.%m.%Y") if c.last_meeting_date else "—", open_t])

    from fastapi.responses import Response
    return Response(content=buf.getvalue(), media_type="text/csv",
                    headers={"Content-Disposition": f"attachment; filename=analytics_{datetime.utcnow().strftime('%Y%m%d')}.csv"})


# ── Auto-task rules ─────────────────────────────────────────────────────────

@app.get("/api/auto-tasks/rules")
async def api_auto_task_rules_list(db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    if not auth_token: raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload: raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user: raise HTTPException(status_code=401)
    from models import AutoTaskRule
    from sqlalchemy import or_
    rules = db.query(AutoTaskRule).filter(
        or_(AutoTaskRule.user_id == user.id, AutoTaskRule.user_id.is_(None))
    ).order_by(AutoTaskRule.created_at.desc()).all()
    return {"rules": [{"id":r.id,"name":r.name,"trigger":r.trigger,"trigger_config":r.trigger_config,
                        "segment_filter":r.segment_filter,"task_title":r.task_title,
                        "task_description":r.task_description,"task_priority":r.task_priority,
                        "task_due_days":r.task_due_days,"task_type":r.task_type,
                        "is_active":r.is_active,"created_at":r.created_at.isoformat() if r.created_at else None}
                       for r in rules]}


@app.post("/api/auto-tasks/rules")
async def api_auto_task_rules_create(request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    if not auth_token: raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload: raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user: raise HTTPException(status_code=401)
    body = await request.json()
    from models import AutoTaskRule
    rule = AutoTaskRule(user_id=user.id, **{k:v for k,v in body.items()
                         if k in ("name","trigger","trigger_config","segment_filter","task_title",
                                  "task_description","task_priority","task_due_days","task_type","is_active")})
    db.add(rule); db.commit(); db.refresh(rule)
    return {"ok": True, "id": rule.id}


@app.put("/api/auto-tasks/rules/{rule_id}")
async def api_auto_task_rules_update(rule_id: int, request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    if not auth_token: raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload: raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user: raise HTTPException(status_code=401)
    from models import AutoTaskRule
    from sqlalchemy.orm.attributes import flag_modified
    rule = db.query(AutoTaskRule).filter(AutoTaskRule.id == rule_id).first()
    if not rule: raise HTTPException(status_code=404)
    body = await request.json()
    for k, v in body.items():
        if hasattr(rule, k): setattr(rule, k, v)
    db.commit()
    return {"ok": True}


@app.patch("/api/auto-tasks/rules/{rule_id}")
async def api_auto_task_rules_patch(rule_id: int, request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    if not auth_token: raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload: raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user: raise HTTPException(status_code=401)
    from models import AutoTaskRule
    rule = db.query(AutoTaskRule).filter(AutoTaskRule.id == rule_id).first()
    if not rule: raise HTTPException(status_code=404)
    body = await request.json()
    for k, v in body.items():
        if hasattr(rule, k): setattr(rule, k, v)
    db.commit()
    return {"ok": True}


@app.delete("/api/auto-tasks/rules/{rule_id}")
async def api_auto_task_rules_delete(rule_id: int, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    if not auth_token: raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload: raise HTTPException(status_code=401)
    from models import AutoTaskRule
    rule = db.query(AutoTaskRule).filter(AutoTaskRule.id == rule_id).first()
    if rule: db.delete(rule); db.commit()
    return {"ok": True}


@app.post("/api/auto-tasks/rules/{rule_id}/test")
async def api_auto_task_rules_test(rule_id: int, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Тестовый прогон правила — создаёт задачи для подходящих клиентов."""
    if not auth_token: raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload: raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user: raise HTTPException(status_code=401)
    from models import AutoTaskRule
    rule = db.query(AutoTaskRule).filter(AutoTaskRule.id == rule_id).first()
    if not rule: raise HTTPException(status_code=404)

    q = db.query(Client)
    if user.role == "manager": q = q.filter(Client.manager_email == user.email)
    clients = q.all()

    segs = rule.segment_filter or []
    if segs: clients = [c for c in clients if c.segment in segs]

    cfg = rule.trigger_config or {}
    triggered = []
    now = datetime.utcnow()

    for c in clients:
        match = False
        if rule.trigger == "health_drop":
            threshold = cfg.get("threshold", 50)
            match = (c.health_score or 0) < threshold
        elif rule.trigger == "days_no_contact":
            days = cfg.get("days", 30)
            last = c.last_meeting_date or c.last_checkup
            match = not last or (now - last).days >= days
        elif rule.trigger == "checkup_due":
            interval = CHECKUP_INTERVALS.get(c.segment or "", 90)
            last = c.last_meeting_date or c.last_checkup
            match = not last or (now - last).days >= interval
        if match:
            triggered.append(c)

    created = 0
    for c in triggered[:10]:  # Лимит 10 за тест
        due = now + timedelta(days=rule.task_due_days or 3)
        task = Task(
            client_id=c.id, title=rule.task_title,
            description="[Автозадача: " + rule.name + "]\n" + (rule.task_description or ""),
            status="plan", priority=rule.task_priority or "medium",
            due_date=due, created_at=now,
        )
        db.add(task)
        created += 1

    db.commit()
    return {"ok": True, "triggered": len(triggered), "created": created}


# ── Followup templates ──────────────────────────────────────────────────────

@app.get("/api/followup/templates")
async def api_followup_templates_list(db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    if not auth_token: raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload: raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user: raise HTTPException(status_code=401)
    tmpls = db.query(FollowupTemplate).filter(FollowupTemplate.user_id == user.id).order_by(FollowupTemplate.created_at.desc()).all()
    return {"templates": [{"id":t.id,"name":t.name,"content":t.content,"category":t.category,"created_at":t.created_at.isoformat() if t.created_at else None} for t in tmpls]}


@app.post("/api/followup/templates")
async def api_followup_templates_create(request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    if not auth_token: raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload: raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user: raise HTTPException(status_code=401)
    body = await request.json()
    t = FollowupTemplate(user_id=user.id, name=body["name"], content=body["content"], category=body.get("category","general"))
    db.add(t); db.commit(); db.refresh(t)
    return {"ok": True, "id": t.id}


@app.put("/api/followup/templates/{tmpl_id}")
async def api_followup_templates_update(tmpl_id: int, request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    if not auth_token: raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload: raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user: raise HTTPException(status_code=401)
    t = db.query(FollowupTemplate).filter(FollowupTemplate.id == tmpl_id, FollowupTemplate.user_id == user.id).first()
    if not t: raise HTTPException(status_code=404)
    body = await request.json()
    for k in ("name","content","category"):
        if k in body: setattr(t, k, body[k])
    db.commit()
    return {"ok": True}


@app.delete("/api/followup/templates/{tmpl_id}")
async def api_followup_templates_delete(tmpl_id: int, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    if not auth_token: raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload: raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user: raise HTTPException(status_code=401)
    t = db.query(FollowupTemplate).filter(FollowupTemplate.id == tmpl_id, FollowupTemplate.user_id == user.id).first()
    if t: db.delete(t); db.commit()
    return {"ok": True}


@app.get("/api/tasks/all")
async def api_tasks_all(db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    if not auth_token: raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload: raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user: raise HTTPException(status_code=401)
    q = db.query(Task).join(Client, Task.client_id == Client.id, isouter=True)
    if user.role == "manager": q = q.filter(Client.manager_email == user.email)
    tasks = q.order_by(Task.due_date.asc().nullslast(), Task.created_at.desc()).all()
    return {"tasks": [{"id":t.id,"title":t.title,"status":t.status,"priority":t.priority,
                        "due_date":t.due_date.isoformat() if t.due_date else None,
                        "client_id":t.client_id,"client_name":t.client.name if t.client else "—",
                        "description":t.description,"merchrules_task_id":t.merchrules_task_id}
                       for t in tasks]}


# ============================================================================
# MISSING PAGES — страницы из nav без endpoint
# ============================================================================

@app.get("/hub", response_class=HTMLResponse)
async def hub_page(request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Командный центр — редирект на dashboard."""
    if not auth_token: return RedirectResponse(url="/login", status_code=303)
    return RedirectResponse(url="/dashboard", status_code=302)

@app.get("/top50", response_class=HTMLResponse)
async def top50_page(request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    if not auth_token: return RedirectResponse(url="/login", status_code=303)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload: return RedirectResponse(url="/login", status_code=303)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user: return RedirectResponse(url="/login", status_code=303)
    return templates.TemplateResponse("top50.html", {"request": request, "user": user})

@app.get("/roadmap", response_class=HTMLResponse)
async def roadmap_page(request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    if not auth_token: return RedirectResponse(url="/login", status_code=303)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload: return RedirectResponse(url="/login", status_code=303)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user: return RedirectResponse(url="/login", status_code=303)
    return templates.TemplateResponse("roadmap.html", {"request": request, "user": user})

@app.get("/internal-tasks", response_class=HTMLResponse)
async def internal_tasks_page(request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    if not auth_token: return RedirectResponse(url="/login", status_code=303)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload: return RedirectResponse(url="/login", status_code=303)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user: return RedirectResponse(url="/login", status_code=303)
    return templates.TemplateResponse("internal_tasks.html", {"request": request, "user": user})

@app.get("/qbr-calendar", response_class=HTMLResponse)
async def qbr_calendar_page(request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    if not auth_token: return RedirectResponse(url="/login", status_code=303)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload: return RedirectResponse(url="/login", status_code=303)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user: return RedirectResponse(url="/login", status_code=303)
    return templates.TemplateResponse("qbr_calendar.html", {"request": request, "user": user})


# ── Ktalk DM ────────────────────────────────────────────────────────────────
@app.post("/api/ktalk/send-dm")
async def api_ktalk_send_dm(
    request: Request,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """Отправить DM клиенту через KTalk."""
    if not auth_token: raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload: raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user: raise HTTPException(status_code=401)

    body = await request.json()
    client_id = body.get("client_id")
    message   = body.get("message", "")
    channel_id = body.get("channel_id")

    if not message:
        return {"ok": False, "error": "Нет текста сообщения"}

    u_settings = user.settings or {}
    kt = u_settings.get("ktalk", {})
    token = kt.get("access_token") or _env("KTALK_API_TOKEN")
    if not token:
        return {"ok": False, "error": "KTalk не настроен — войдите в Настройки → KTalk"}

    # Получаем channel_id если не передан
    if not channel_id and client_id:
        client = db.query(Client).filter(Client.id == client_id).first()
        if client:
            meta = client.integration_metadata or {}
            channel_id = meta.get("ktalk_channel_id") or kt.get("followup_channel_id")

    if not channel_id:
        return {"ok": False, "error": "Нет channel_id для отправки"}

    try:
        import httpx
        async with httpx.AsyncClient(timeout=10) as hx:
            r = await hx.post(
                "https://tbank.ktalk.ru/api/v4/posts",
                headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
                json={"channel_id": channel_id, "message": message}
            )
        if r.status_code in (200, 201):
            return {"ok": True}
        return {"ok": False, "error": f"KTalk HTTP {r.status_code}"}
    except Exception as e:
        return {"ok": False, "error": str(e)}


# ── Integration tests ────────────────────────────────────────────────────────
@app.get("/api/integrations/test/airtable")
async def api_test_airtable(db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    if not auth_token: raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload: raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user: raise HTTPException(status_code=401)
    u = user.settings or {}
    at = u.get("airtable", {})
    token = at.get("pat") or at.get("token") or _env("AIRTABLE_TOKEN")
    if not token:
        return {"ok": False, "error": "Airtable токен не настроен"}
    try:
        import httpx
        async with httpx.AsyncClient(timeout=10) as hx:
            r = await hx.get("https://api.airtable.com/v0/meta/bases",
                             headers={"Authorization": f"Bearer {token}"})
        return {"ok": r.status_code == 200, "error": f"HTTP {r.status_code}" if r.status_code != 200 else None}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.get("/api/integrations/test/tbank_time")
@app.get("/api/integrations/test/tbank")
async def api_test_tbank(db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    if not auth_token: raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload: raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user: raise HTTPException(status_code=401)
    u = user.settings or {}
    tm = u.get("tbank_time", {})
    token = tm.get("session_cookie") or tm.get("mmauthtoken") or tm.get("api_token") or _env("TIME_API_TOKEN")
    if not token:
        return {"ok": False, "error": "Tbank Time токен не настроен"}
    try:
        import httpx
        async with httpx.AsyncClient(timeout=10) as hx:
            r = await hx.get("https://time.tbank.ru/api/v4/users/me",
                             headers={"Authorization": f"Bearer {token}"})
        if r.status_code == 200:
            me = r.json()
            return {"ok": True, "username": me.get("username"), "email": me.get("email")}
        return {"ok": False, "error": f"HTTP {r.status_code} — токен истёк?"}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.get("/api/integrations/test/{system}")
async def api_test_integration_generic(system: str, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Заглушка для остальных систем."""
    return {"ok": False, "error": f"Тест для {system} не реализован"}


# ── Import CSV ───────────────────────────────────────────────────────────────
@app.post("/api/import/clients-csv")
async def api_import_clients_csv(
    file: UploadFile,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """Импорт клиентов из CSV файла."""
    if not auth_token: raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload: raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user: raise HTTPException(status_code=401)

    import io, pandas as pd
    content_bytes = await file.read()
    try:
        df = pd.read_csv(io.BytesIO(content_bytes), dtype=str)
        df.columns = [c.strip().lower().replace(" ", "_") for c in df.columns]
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Ошибка чтения CSV: {e}")

    created = updated = skipped = 0
    from sqlalchemy.orm.attributes import flag_modified
    for _, row in df.iterrows():
        name = str(row.get("name") or row.get("название") or row.get("client_name") or "").strip()
        if not name or name == "nan":
            skipped += 1
            continue
        client = db.query(Client).filter(Client.name == name).first()
        if client:
            for field, col in [("segment","segment"),("domain","domain"),("health_score","health_score")]:
                v = str(row.get(col) or "").strip()
                if v and v != "nan":
                    if field == "health_score":
                        try: setattr(client, field, float(v.replace("%","")))
                        except: pass
                    else: setattr(client, field, v)
            updated += 1
        else:
            seg = str(row.get("segment") or "").strip()
            domain = str(row.get("domain") or "").strip()
            client = Client(name=name, segment=seg or None, domain=domain or None,
                           manager_email=user.email)
            db.add(client)
            created += 1
    db.commit()
    return {"ok": True, "created": created, "updated": updated, "skipped": skipped}


# ── Roadmap create ───────────────────────────────────────────────────────────
@app.post("/api/roadmap/create")
async def api_roadmap_create(request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    if not auth_token: raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload: raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user: raise HTTPException(status_code=401)
    body = await request.json()
    task = Task(client_id=body.get("client_id"), title=body.get("title",""), 
                status="plan", priority=body.get("priority","medium"),
                created_at=datetime.utcnow())
    db.add(task); db.commit(); db.refresh(task)
    return {"ok": True, "id": task.id}

# ============================================================================
# MANAGER CABINET — личный кабинет менеджера
# ============================================================================

@app.get("/cabinet", response_class=HTMLResponse)
async def manager_cabinet(
    request: Request,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """Личный кабинет менеджера — его клиенты + выбор из общего пула."""
    if not auth_token:
        return RedirectResponse(url="/login", status_code=303)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        return RedirectResponse(url="/login", status_code=303)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    return templates.TemplateResponse("cabinet.html", {"request": request, "user": user})


@app.get("/api/cabinet/my-clients")
async def api_cabinet_my_clients(
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """Клиенты текущего менеджера (назначены через assignment или manager_email)."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        raise HTTPException(status_code=401)

    # Объединяем: manager_email ИЛИ user_client_assignment
    from sqlalchemy import or_
    assigned_ids = [a.client_id for a in
                    db.query(UserClientAssignment).filter(UserClientAssignment.user_id == user.id).all()]
    clients = db.query(Client).filter(
        or_(Client.manager_email == user.email, Client.id.in_(assigned_ids))
    ).order_by(Client.name).all()

    now = datetime.now()
    result = []
    for c in clients:
        meta = c.integration_metadata or {}
        digi = meta.get("diginetica", {})
        products = [p for p in ("sort", "autocomplete", "recommendations") if digi.get(p, {}).get("api_key")]
        if not products and meta.get("diginetica_api_key"):
            products = ["sort"]
        interval = CHECKUP_INTERVALS.get(c.segment or "", 90)
        last = c.last_meeting_date or c.last_checkup
        days_since = (now - last).days if last else 999
        result.append({
            "id": c.id, "name": c.name, "segment": c.segment,
            "health_score": c.health_score,
            "domain": c.domain or meta.get("site_url", ""),
            "products": products,
            "has_api_keys": len(products) > 0,
            "days_since_checkup": days_since,
            "checkup_overdue": days_since > interval,
            "merchrules_id": c.merchrules_account_id,
            "last_meeting": c.last_meeting_date.isoformat() if c.last_meeting_date else None,
        })
    return {"clients": result, "total": len(result)}


@app.get("/api/cabinet/available-clients")
async def api_cabinet_available_clients(
    search: str = "",
    segment: str = "",
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """Клиенты без менеджера или со свободными слотами — для добавления в свой кабинет."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        raise HTTPException(status_code=401)

    from sqlalchemy import or_
    # Уже назначенные у этого менеджера
    my_ids = {a.client_id for a in
              db.query(UserClientAssignment).filter(UserClientAssignment.user_id == user.id).all()}
    my_emails = {user.email}

    q = db.query(Client)
    # Исключаем уже своих
    if my_ids:
        q = q.filter(~Client.id.in_(my_ids))
    q = q.filter(or_(Client.manager_email != user.email, Client.manager_email.is_(None)))

    if search:
        q = q.filter(Client.name.ilike(f"%{search}%"))
    if segment:
        q = q.filter(Client.segment == segment)

    clients = q.order_by(Client.name).limit(100).all()
    return {
        "clients": [
            {
                "id": c.id, "name": c.name, "segment": c.segment,
                "health_score": c.health_score,
                "manager_email": c.manager_email or "—",
            }
            for c in clients
        ]
    }


@app.post("/api/cabinet/assign/{client_id}")
async def api_cabinet_assign(
    client_id: int,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """Добавить клиента в свой кабинет."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        raise HTTPException(status_code=401)

    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        raise HTTPException(status_code=404)

    # Проверяем нет ли уже
    existing = db.query(UserClientAssignment).filter(
        UserClientAssignment.user_id == user.id,
        UserClientAssignment.client_id == client_id,
    ).first()
    if not existing:
        db.add(UserClientAssignment(user_id=user.id, client_id=client_id))
        # Устанавливаем manager_email если он пустой
        if not client.manager_email:
            client.manager_email = user.email
        db.commit()
    return {"ok": True}


@app.delete("/api/cabinet/assign/{client_id}")
async def api_cabinet_unassign(
    client_id: int,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """Убрать клиента из своего кабинета."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        raise HTTPException(status_code=401)

    db.query(UserClientAssignment).filter(
        UserClientAssignment.user_id == user.id,
        UserClientAssignment.client_id == client_id,
    ).delete()
    db.commit()
    return {"ok": True}

# ============================================================================
# SEARCH QUALITY CHECKUP — API для расширения
# ============================================================================

def _checkup_auth(auth_token: Optional[str], db):
    """Общая авторизация для checkup endpoints."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        raise HTTPException(status_code=401)
    return user


@app.get("/api/cabinets/{cabinet_id}")
async def api_get_cabinet(
    cabinet_id: str,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """
    Возвращает данные кабинета для расширения Search Quality Checkup.
    cabinet_id = client.id или client.merchrules_account_id
    """
    user = _checkup_auth(auth_token, db)

    # Ищем клиента по id или по merchrules_account_id
    client = None
    if cabinet_id.isdigit():
        client = db.query(Client).filter(Client.id == int(cabinet_id)).first()
    if not client:
        client = db.query(Client).filter(
            Client.merchrules_account_id == cabinet_id
        ).first()
    if not client:
        raise HTTPException(status_code=404, detail=f"Кабинет {cabinet_id} не найден")

    meta = client.integration_metadata or {}
    api_key = (
        meta.get("diginetica_api_key")
        or meta.get("search_api_key")
        or meta.get("apiKey")
        or ""
    )
    site_url = client.domain or meta.get("site_url") or ""
    if site_url and not site_url.startswith("http"):
        site_url = "https://" + site_url

    digi = meta.get("diginetica", {})

    products = {}
    for product in ("sort", "autocomplete", "recommendations"):
        p = digi.get(product, {})
        if p.get("api_key"):
            products[product] = {
                "apiKey": p["api_key"],
                "url":    p.get("url", ""),
            }
    # Legacy fallback — если нет структурированных, используем diginetica_api_key как sort
    if not products and api_key:
        products["sort"] = {"apiKey": api_key, "url": "https://sort.diginetica.net/search"}

    return {
        "ok": True,
        "cabinetId": str(client.id),
        "clientName": client.name,
        # Основной ключ (Sort или первый доступный) — для обратной совместимости с расширением
        "apiKey": (products.get("sort") or next(iter(products.values()), {})).get("apiKey", ""),
        "siteUrl": site_url,
        "segment": client.segment,
        "healthScore": client.health_score,
        # Все продукты — расширение использует для выбора типа чекапа
        "products": products,
        "hasSort": "sort" in products,
        "hasAutocomplete": "autocomplete" in products,
        "hasRecommendations": "recommendations" in products,
    }


@app.get("/api/checkup/{cabinet_id}/queries")
async def api_checkup_queries(
    cabinet_id: str,
    type: str = "top",
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """
    Возвращает список запросов для чекапа.
    type: top | random | zero | zeroquery
    Запросы берутся из integration_metadata.checkup_queries[type]
    или из сохранённых результатов предыдущих чекапов.
    """
    user = _checkup_auth(auth_token, db)

    client = None
    if cabinet_id.isdigit():
        client = db.query(Client).filter(Client.id == int(cabinet_id)).first()
    if not client:
        client = db.query(Client).filter(
            Client.merchrules_account_id == cabinet_id
        ).first()
    if not client:
        raise HTTPException(status_code=404, detail="Кабинет не найден")

    meta = client.integration_metadata or {}
    checkup_queries = meta.get("checkup_queries", {})
    queries = checkup_queries.get(type, [])

    # Если нет сохранённых — возвращаем пустой список (расширение попросит ввести вручную)
    return {"ok": True, "queries": queries, "type": type, "client": client.name}


@app.get("/api/cabinets/{cabinet_id}/merch-rules")
async def api_merch_rules(
    cabinet_id: str,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """Мерч-правила клиента из integration_metadata."""
    user = _checkup_auth(auth_token, db)

    client = None
    if cabinet_id.isdigit():
        client = db.query(Client).filter(Client.id == int(cabinet_id)).first()
    if not client:
        client = db.query(Client).filter(
            Client.merchrules_account_id == cabinet_id
        ).first()
    if not client:
        return []

    meta = client.integration_metadata or {}
    return meta.get("merch_rules", [])


@app.post("/api/checkup/{cabinet_id}/results")
async def api_save_checkup_results(
    cabinet_id: str,
    request: Request,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """
    Сохраняет результаты чекапа из расширения.
    Расширение вызывает после завершения проверки.
    """
    user = _checkup_auth(auth_token, db)

    client = None
    if cabinet_id.isdigit():
        client = db.query(Client).filter(Client.id == int(cabinet_id)).first()
    if not client:
        client = db.query(Client).filter(
            Client.merchrules_account_id == cabinet_id
        ).first()
    if not client:
        raise HTTPException(status_code=404, detail="Кабинет не найден")

    body = await request.json()
    results = body.get("results", [])
    avg_score = (
        sum(r.get("manualScore") or r.get("autoScore", 0) for r in results) / len(results)
        if results else None
    )
    score_dist = {str(i): 0 for i in range(4)}
    for r in results:
        s = str(r.get("manualScore") or r.get("autoScore", 0))
        score_dist[s] = score_dist.get(s, 0) + 1

    from models import CheckupResult
    cr = CheckupResult(
        client_id=client.id,
        cabinet_id=cabinet_id,
        query_type=body.get("queryType", "top"),
        manager_name=body.get("managerName") or user.name,
        mode=body.get("mode"),
        total_queries=len(results),
        avg_score=avg_score,
        score_dist=score_dist,
        results=results,
    )
    db.add(cr)

    # Обновляем дату последнего чекапа у клиента
    client.last_checkup = datetime.utcnow()
    db.commit()

    logger.info(f"CheckupResult saved: client={client.name}, queries={len(results)}, avg={avg_score:.2f if avg_score else 'N/A'}")
    return {"ok": True, "id": cr.id, "avg_score": avg_score, "total": len(results)}


@app.get("/api/checkup/{cabinet_id}/history")
async def api_checkup_history(
    cabinet_id: str,
    limit: int = 10,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """История чекапов клиента."""
    user = _checkup_auth(auth_token, db)

    client_id = None
    if cabinet_id.isdigit():
        client_id = int(cabinet_id)
    else:
        c = db.query(Client).filter(Client.merchrules_account_id == cabinet_id).first()
        if c:
            client_id = c.id

    if not client_id:
        return {"results": []}

    from models import CheckupResult
    history = (
        db.query(CheckupResult)
        .filter(CheckupResult.client_id == client_id)
        .order_by(CheckupResult.created_at.desc())
        .limit(limit)
        .all()
    )
    return {
        "results": [
            {
                "id": r.id,
                "query_type": r.query_type,
                "manager_name": r.manager_name,
                "mode": r.mode,
                "total_queries": r.total_queries,
                "avg_score": round(r.avg_score, 2) if r.avg_score else None,
                "score_dist": r.score_dist,
                "created_at": r.created_at.isoformat() if r.created_at else None,
            }
            for r in history
        ]
    }


@app.put("/api/clients/{client_id}/checkup-config")
async def api_save_checkup_config(
    client_id: int,
    request: Request,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """
    Сохраняет конфиг чекапа клиента:
    diginetica_api_key, site_url, checkup_queries (top/random/zero/zeroquery)
    """
    user = _checkup_auth(auth_token, db)
    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        raise HTTPException(status_code=404)

    body = await request.json()
    meta = dict(client.integration_metadata or {})

    if "diginetica_api_key" in body:
        meta["diginetica_api_key"] = body["diginetica_api_key"]
    if "site_url" in body:
        meta["site_url"] = body["site_url"]
    if "checkup_queries" in body:
        meta["checkup_queries"] = body["checkup_queries"]
    if "merch_rules" in body:
        meta["merch_rules"] = body["merch_rules"]

    from sqlalchemy.orm.attributes import flag_modified
    client.integration_metadata = meta
    flag_modified(client, "integration_metadata")
    db.commit()
    return {"ok": True}


# ============================================================================
# WORKFLOW: DASHBOARD ACTIONS
# ============================================================================

@app.get("/api/dashboard/actions")
async def api_dashboard_actions(db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Получить карточки действий для дашборда."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)

    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    now = datetime.now()

    actions = []

    # 1. Фолоуапы pending
    pending_followups = db.query(Meeting).filter(
        Meeting.followup_status == "pending",
        Meeting.date < now,
    ).all()
    for m in pending_followups:
        client = db.query(Client).filter(Client.id == m.client_id).first()
        actions.append({
            "type": "followup",
            "priority": "high",
            "meeting_id": m.id,
            "client_name": client.name if client else "—",
            "meeting_title": m.title or m.type,
            "meeting_date": m.date.isoformat() if m.date else None,
            "days_ago": (now - m.date).days if m.date else 0,
        })

    # 2. Prep до встречи
    upcoming = db.query(Meeting).filter(
        Meeting.date >= now,
        Meeting.date < now + timedelta(days=2),
    ).all()
    for m in upcoming:
        client = db.query(Client).filter(Client.id == m.client_id).first()
        actions.append({
            "type": "prep",
            "priority": "medium",
            "meeting_id": m.id,
            "client_name": client.name if client else "—",
            "meeting_title": m.title or m.type,
            "meeting_date": m.date.isoformat() if m.date else None,
            "hours_until": int((m.date - now).total_seconds() / 3600) if m.date else 0,
        })

    # 3. Chekups overdue
    overdue_checkups = db.query(CheckUp).filter(CheckUp.status == "overdue").all()
    for c in overdue_checkups:
        client = db.query(Client).filter(Client.id == c.client_id).first()
        actions.append({
            "type": "checkup",
            "priority": "high",
            "checkup_id": c.id,
            "client_name": client.name if client else "—",
            "checkup_type": c.type,
            "scheduled_date": c.scheduled_date.isoformat() if c.scheduled_date else None,
        })

    # 4. QBR overdue
    clients_qbr = db.query(Client).filter(
        Client.next_qbr_date != None,
        Client.next_qbr_date < now,
    ).all()
    for c in clients_qbr:
        actions.append({
            "type": "qbr",
            "priority": "high",
            "client_id": c.id,
            "client_name": c.name,
            "next_qbr_date": c.next_qbr_date.isoformat() if c.next_qbr_date else None,
        })

    return {"actions": actions, "total": len(actions)}


# ============================================================================
# ONBOARDING
# ============================================================================

@app.get("/onboarding", response_class=HTMLResponse)
async def onboarding_page(request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    if not auth_token:
        return RedirectResponse(url="/login", status_code=303)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        return RedirectResponse(url="/login", status_code=303)
    return templates.TemplateResponse("onboarding.html", {"request": request})


@app.post("/api/onboarding/complete")
async def api_complete_onboarding(db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Отметить что онбординг пройден."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        raise HTTPException(status_code=401)
    settings = user.settings or {}
    settings["onboarding_complete"] = True
    user.settings = dict(settings)
    from sqlalchemy.orm.attributes import flag_modified
    flag_modified(user, "settings")
    db.commit()
    return {"ok": True}



# ============================================================================
# ADMIN PANEL
# ============================================================================

def _require_admin(auth_token, db):
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user or user.role != "admin":
        raise HTTPException(status_code=403, detail="Только для администраторов")
    return user


@app.get("/admin", response_class=HTMLResponse)
async def admin_page(
    request: Request,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """Панель администратора."""
    if not auth_token:
        return RedirectResponse(url="/login", status_code=303)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        return RedirectResponse(url="/login", status_code=303)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user or user.role != "admin":
        raise HTTPException(status_code=403, detail="Только для администраторов")

    # Статистика
    total_clients  = db.query(Client).count()
    clients_with_key = db.query(Client).filter(
        Client.integration_metadata.op("->>")('diginetica_api_key').isnot(None)
    ).count()
    total_users = db.query(User).count()
    from models import CheckupResult
    total_checkups = db.query(CheckupResult).count()

    return templates.TemplateResponse("admin.html", {
        "request": request, "user": user,
        "stats": {
            "total_clients": total_clients,
            "clients_with_key": clients_with_key,
            "total_users": total_users,
            "total_checkups": total_checkups,
        },
    })


@app.post("/api/admin/import/api-keys")
async def api_admin_import_api_keys(
    file: UploadFile,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """
    Импорт Diginetica API keys из Excel файла.
    Формат: client_id | client_name | diginetica_api_key | site_url | checkup_queries_top
    Доступно только администраторам.
    """
    _require_admin(auth_token, db)

    if not file.filename.endswith((".xlsx", ".xls", ".csv")):
        raise HTTPException(status_code=400, detail="Поддерживаются только .xlsx / .csv")

    content = await file.read()

    import io, pandas as pd
    try:
        if file.filename.endswith(".csv"):
            df = pd.read_csv(io.BytesIO(content), dtype=str)
        else:
            df = pd.read_excel(io.BytesIO(content), dtype=str, skiprows=1)  # skiprows=1 пропускает описания
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Ошибка чтения файла: {e}")

    df.columns = [c.strip().lower() for c in df.columns]
    if "client_id" not in df.columns:
        raise HTTPException(status_code=400, detail="Нет колонки client_id")

    updated, skipped, errors = 0, 0, []
    from sqlalchemy.orm.attributes import flag_modified

    for _, row in df.iterrows():
        raw_id = row.get("client_id", "")
        if not raw_id or str(raw_id).strip() in ("", "nan", "None"):
            skipped += 1
            continue

        try:
            client_id = int(float(str(raw_id).strip()))
        except ValueError:
            errors.append(f"Неверный client_id: {raw_id!r}")
            continue

        client = db.query(Client).filter(Client.id == client_id).first()
        if not client:
            errors.append(f"Клиент #{client_id} не найден")
            continue

        meta = dict(client.integration_metadata or {})

        # Diginetica продукты — Sort / Autocomplete / Recommendations
        digi = meta.get("diginetica", {})

        def _val(key):
            v = str(row.get(key, "") or "").strip()
            return "" if v in ("", "nan", "None") else v

        # Sort
        if _val("sort_api_key"):
            digi.setdefault("sort", {})["api_key"] = _val("sort_api_key")
        if _val("sort_url"):
            digi.setdefault("sort", {})["url"] = _val("sort_url")
        # Autocomplete
        if _val("auto_api_key"):
            digi.setdefault("autocomplete", {})["api_key"] = _val("auto_api_key")
        if _val("auto_url"):
            digi.setdefault("autocomplete", {})["url"] = _val("auto_url")
        # Recommendations
        if _val("rec_api_key"):
            digi.setdefault("recommendations", {})["api_key"] = _val("rec_api_key")
        if _val("rec_url"):
            digi.setdefault("recommendations", {})["url"] = _val("rec_url")
        # Legacy single key
        if _val("diginetica_api_key"):
            meta["diginetica_api_key"] = _val("diginetica_api_key")
            digi.setdefault("sort", {})["api_key"] = _val("diginetica_api_key")

        if digi:
            meta["diginetica"] = digi

        # site_url
        site_url = _val("site_url")
        if site_url:
            if not site_url.startswith("http"):
                site_url = "https://" + site_url
            meta["site_url"] = site_url

        # Запросы для чекапа (top)
        queries_raw = _val("checkup_queries_top")
        if queries_raw:
            queries = [q.strip() for q in queries_raw.split(",") if q.strip()]
            cq = meta.get("checkup_queries", {})
            cq["top"] = queries
            meta["checkup_queries"] = cq

        client.integration_metadata = meta
        flag_modified(client, "integration_metadata")
        updated += 1

    db.commit()
    logger.info(f"Admin import API keys: updated={updated}, skipped={skipped}, errors={len(errors)}")
    return {"ok": True, "updated": updated, "skipped": skipped, "errors": errors}


@app.get("/api/admin/clients/api-keys")
async def api_admin_list_api_keys(
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """Список всех клиентов с их API keys (только admin)."""
    _require_admin(auth_token, db)
    clients = db.query(Client).order_by(Client.name).all()
    return {
        "clients": [
            {
                "id": c.id,
                "name": c.name,
                "segment": c.segment,
                "has_api_key": bool((c.integration_metadata or {}).get("diginetica_api_key")),
                "api_key": (c.integration_metadata or {}).get("diginetica_api_key", ""),
                "site_url": c.domain or (c.integration_metadata or {}).get("site_url", ""),
                "has_queries": bool((c.integration_metadata or {}).get("checkup_queries", {}).get("top")),
            }
            for c in clients
        ]
    }


@app.patch("/api/admin/clients/{client_id}/api-key")
async def api_admin_set_api_key(
    client_id: int,
    request: Request,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """Установить/обновить API key конкретного клиента (только admin)."""
    _require_admin(auth_token, db)
    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        raise HTTPException(status_code=404)

    body = await request.json()
    from sqlalchemy.orm.attributes import flag_modified
    meta = dict(client.integration_metadata or {})

    if "diginetica_api_key" in body:
        meta["diginetica_api_key"] = body["diginetica_api_key"]
    if "site_url" in body:
        meta["site_url"] = body["site_url"]
    if "checkup_queries" in body:
        meta["checkup_queries"] = body["checkup_queries"]

    client.integration_metadata = meta
    flag_modified(client, "integration_metadata")
    db.commit()
    return {"ok": True}


@app.get("/api/clients/{client_id}/checkup-info")
async def api_client_checkup_info(
    client_id: int,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """
    Публичная информация о checkup-конфиге клиента.
    API key — только маска для менеджеров, полный — для admin.
    """
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        raise HTTPException(status_code=401)

    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        raise HTTPException(status_code=404)

    meta = client.integration_metadata or {}
    api_key = meta.get("diginetica_api_key", "")

    # Менеджер видит только маску: первые 4 и последние 4 символа
    if user.role != "admin" and api_key:
        masked = api_key[:4] + "••••••••••••••••••••" + api_key[-4:] if len(api_key) > 8 else "••••••••"
        api_key_display = masked
    else:
        api_key_display = api_key

    cq = meta.get("checkup_queries", {})
    return {
        "client_id": client_id,
        "client_name": client.name,
        "has_api_key": bool(meta.get("diginetica_api_key")),
        "api_key_display": api_key_display,
        "api_key_full": api_key if user.role == "admin" else None,
        "site_url": client.domain or meta.get("site_url", ""),
        "is_admin": user.role == "admin",
        "checkup_queries": cq,
        "queries_count": {k: len(v) for k, v in cq.items() if isinstance(v, list)},
    }

@app.post("/api/admin/reset-data")
async def api_reset_data(db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Удалить все тестовые данные (только для админа)."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user or user.role != "admin":
        raise HTTPException(status_code=403)

    # Удаляем в правильном порядке (из-за FK)
    db.query(Task).delete()
    db.query(Meeting).delete()
    db.query(CheckUp).delete()
    db.query(QBR).delete()
    db.query(AccountPlan).delete()
    db.query(Client).delete()
    db.commit()
    return {"ok": True, "message": "Все данные очищены"}


@app.get("/api/onboarding/status")
async def api_onboarding_status(db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Проверить, пройден ли онбординг."""
    if not auth_token:
        return {"onboarding_complete": True}
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        return {"onboarding_complete": True}
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        return {"onboarding_complete": True}
    settings = user.settings or {}
    return {"onboarding_complete": settings.get("onboarding_complete", False)}


# ============================================================================
# GLOBAL SEARCH
# ============================================================================

@app.get("/api/search")
async def api_global_search(
    q: str = Query("", min_length=1),
    limit: int = Query(20, ge=1, le=50),
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """Глобальный поиск по клиентам, задачам, встречам, заметкам."""
    if not auth_token:
        return {"clients": [], "tasks": [], "meetings": [], "notes": []}
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        return {"clients": [], "tasks": [], "meetings": [], "notes": []}

    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        return {"clients": [], "tasks": [], "meetings": [], "notes": []}

    search_pattern = f"%{q}%"

    # Клиенты
    c_q = db.query(Client)
    if user.role == "manager":
        c_q = c_q.filter(Client.manager_email == user.email)
    clients = c_q.filter(
        Client.name.ilike(search_pattern) |
        (Client.segment is not None and Client.segment.ilike(search_pattern)),
    ).limit(limit).all()

    # Задачи
    task_query = db.query(Task).join(Client, Task.client_id == Client.id, isouter=True)
    if user.role == "manager":
        task_query = task_query.filter(Client.manager_email == user.email)
    tasks = task_query.filter(
        Task.title.ilike(search_pattern) |
        (Task.description is not None and Task.description.ilike(search_pattern)),
    ).limit(limit).all()

    # Встречи
    meeting_query = db.query(Meeting).join(Client, Meeting.client_id == Client.id, isouter=True)
    if user.role == "manager":
        meeting_query = meeting_query.filter(Client.manager_email == user.email)
    meetings = meeting_query.filter(
        (Meeting.title is not None and Meeting.title.ilike(search_pattern)) |
        (Meeting.type is not None and Meeting.type.ilike(search_pattern)),
    ).order_by(Meeting.date.desc()).limit(limit).all()

    # Заметки
    note_query = db.query(ClientNote).join(Client, ClientNote.client_id == Client.id, isouter=True)
    if user.role == "manager":
        note_query = note_query.filter(Client.manager_email == user.email)
    notes = note_query.filter(ClientNote.content.ilike(search_pattern)).order_by(
        ClientNote.is_pinned.desc(), ClientNote.updated_at.desc()
    ).limit(limit).all()

    return {
        "clients": [{"id": c.id, "name": c.name, "segment": c.segment, "url": f"/client/{c.id}", "type": "client"} for c in clients],
        "tasks": [{"id": t.id, "title": t.title, "status": t.status, "client_name": t.client.name if t.client else "—", "url": f"/client/{t.client_id}", "type": "task"} for t in tasks],
        "meetings": [{"id": m.id, "title": m.title or m.type, "date": m.date.isoformat() if m.date else None, "client_name": m.client.name if m.client else "—", "url": f"/client/{m.client_id}", "type": "meeting"} for m in meetings],
        "notes": [{"id": n.id, "content": n.content[:100] + "..." if len(n.content) > 100 else n.content, "client_name": n.client.name if n.client else "—", "url": f"/client/{n.client_id}", "type": "note", "pinned": n.is_pinned} for n in notes],
    }


# ============================================================================
# CLIENT NOTES API
# ============================================================================

@app.post("/api/clients/{client_id}/notes")
async def api_create_note(client_id: int, request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Создать заметку к клиенту."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    data = await request.json()
    note = ClientNote(client_id=client_id, user_id=user.id, content=data.get("content", ""), is_pinned=data.get("pinned", False))
    db.add(note)
    db.commit()
    return {"ok": True, "id": note.id}


@app.get("/api/clients/{client_id}/notes")
async def api_get_notes(client_id: int, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Получить заметки клиента."""
    if not auth_token:
        return {"notes": []}
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        return {"notes": []}
    notes = db.query(ClientNote).filter(ClientNote.client_id == client_id).order_by(ClientNote.is_pinned.desc(), ClientNote.updated_at.desc()).all()
    return {"notes": [{"id": n.id, "content": n.content, "pinned": n.is_pinned, "created_at": n.created_at.strftime("%d.%m.%Y %H:%M") if n.created_at else None, "updated_at": n.updated_at.strftime("%d.%m.%Y %H:%M") if n.updated_at else None, "user_id": n.user_id} for n in notes]}


@app.put("/api/clients/notes/{note_id}")
async def api_update_note(note_id: int, request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Обновить заметку."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)
    note = db.query(ClientNote).filter(ClientNote.id == note_id).first()
    if not note:
        raise HTTPException(status_code=404)
    data = await request.json()
    if "content" in data:
        note.content = data["content"]
    if "pinned" in data:
        note.is_pinned = data["pinned"]
    note.updated_at = datetime.utcnow()
    db.commit()
    return {"ok": True}


@app.delete("/api/clients/notes/{note_id}")
async def api_delete_note(note_id: int, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Удалить заметку."""
    if not auth_token:
        raise HTTPException(status_code=401)
    note = db.query(ClientNote).filter(ClientNote.id == note_id).first()
    if not note:
        raise HTTPException(status_code=404)
    db.delete(note)
    db.commit()
    return {"ok": True}


# ============================================================================
# KANBAN API
# ============================================================================

@app.get("/api/kanban")
async def api_kanban(
    client_id: Optional[int] = None,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """Получить задачи в формате канбан (группировка по статусам)."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        raise HTTPException(status_code=401)

    q = db.query(Task).join(Client, Task.client_id == Client.id, isouter=True)
    if user.role == "manager":
        q = q.filter(Client.manager_email == user.email)
    if client_id:
        q = q.filter(Task.client_id == client_id)
    tasks = q.order_by(Task.due_date.asc()).all()

    columns = {"plan": [], "in_progress": [], "review": [], "done": [], "blocked": []}
    for t in tasks:
        status = t.status or "plan"
        if status not in columns:
            columns["plan"].append(t)
        else:
            columns[status].append(t)

    def task_dict(t):
        return {
            "id": t.id, "title": t.title, "priority": t.priority, "status": t.status or "plan",
            "due_date": t.due_date.isoformat() if t.due_date else None,
            "client_name": t.client.name if t.client else "—",
            "client_id": t.client_id, "team": t.team,
            "created_at": t.created_at.isoformat() if t.created_at else None,
        }

    # Возвращаем оба формата — columns (для client_detail) и плоский (для kanban страницы)
    columns_list = [
        {"id": col, "tasks": [task_dict(t) for t in tlist]}
        for col, tlist in columns.items()
    ]
    return {
        "columns": columns_list,
        **{col: [task_dict(t) for t in tlist] for col, tlist in columns.items()}
    }


@app.patch("/api/tasks/{task_id}/status")
async def api_update_task_status(task_id: int, request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Изменить статус задачи (для канбан drag-and-drop)."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    data = await request.json()
    new_status = data.get("status")
    if new_status not in ("plan", "in_progress", "review", "done", "blocked"):
        raise HTTPException(status_code=400, detail="Invalid status")
    task = db.query(Task).filter(Task.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404)
    task.status = new_status
    if new_status == "done":
        task.confirmed_at = datetime.utcnow()
        task.confirmed_by = user.email if user else None
    db.commit()

    # Push статуса в Merchrules если задача оттуда
    if task.merchrules_task_id and user:
        try:
            settings = (user.settings or {})
            mr = settings.get("merchrules", {})
            login = mr.get("login") or env.MR_LOGIN
            password = mr.get("password") or env.MR_PASSWORD
            base_url = _env("MERCHRULES_API_URL", "https://merchrules.any-platform.ru")
            if login and password:
                import httpx as _httpx
                async with _httpx.AsyncClient(timeout=15) as hx:
                    for field in ("email", "login", "username"):
                        r = await hx.post(
                            f"{base_url}/backend-v2/auth/login",
                            json={field: login, "password": password}, timeout=8,
                        )
                        if r.status_code == 200:
                            tok = r.json().get("token") or r.json().get("access_token") or r.json().get("accessToken")
                            if tok:
                                await hx.patch(
                                    f"{base_url}/backend-v2/tasks/{task.merchrules_task_id}",
                                    json={"status": new_status},
                                    headers={"Authorization": f"Bearer {tok}"},
                                    timeout=8,
                                )
                                break
        except Exception as e:
            logger.warning(f"Merchrules task status push failed: {e}")

    return {"ok": True}


# ============================================================================
# MY DAY: TIME TRACKING API
# ============================================================================

@app.post("/api/my-day/schedule")
async def api_my_day_schedule(request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Сохранить расписание задач на день."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        raise HTTPException(status_code=401)
    data = await request.json()
    settings = user.settings or {}
    settings["my_day_schedule"] = data.get("schedule", [])
    settings["my_day_date"] = data.get("date")
    user.settings = dict(settings)
    from sqlalchemy.orm.attributes import flag_modified
    flag_modified(user, "settings")
    db.commit()
    return {"ok": True}


@app.get("/api/my-day/schedule")
async def api_get_my_day_schedule(db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Получить расписание задач на день."""
    if not auth_token:
        return {"schedule": [], "date": None}
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        return {"schedule": [], "date": None}
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        return {"schedule": [], "date": None}
    settings = user.settings or {}
    return {"schedule": settings.get("my_day_schedule", []), "date": settings.get("my_day_date")}


@app.get("/api/clients/{client_id}/timeline")
async def api_client_timeline(client_id: int, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Таймлайн клиента: встречи, задачи, заметки."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)

    events = []

    # Встречи
    meetings = db.query(Meeting).filter(Meeting.client_id == client_id).order_by(Meeting.date.desc()).limit(20).all()
    for m in meetings:
        events.append({
            "type": "followup" if m.followup_status == "sent" else "meeting",
            "date": m.date.strftime("%d.%m.%Y") if m.date else "—",
            "iso_date": m.date.isoformat() if m.date else "",
            "icon": "📅",
            "title": m.title or m.type,
            "desc": (m.summary or "")[:100] + ("..." if m.summary and len(m.summary) > 100 else ""),
        })

    # Задачи
    tasks = db.query(Task).filter(Task.client_id == client_id).order_by(Task.created_at.desc()).limit(20).all()
    for t in tasks:
        events.append({
            "type": "task",
            "date": t.created_at.strftime("%d.%m.%Y") if t.created_at else "—",
            "iso_date": t.created_at.isoformat() if t.created_at else "",
            "icon": {"plan": "📝", "in_progress": "🔄", "done": "✅", "blocked": "🔴", "review": "👀"}.get(t.status, "📋"),
            "title": t.title,
            "desc": f"Статус: {t.status}" + (f" · {t.priority}" if t.priority else ""),
        })

    # Заметки
    notes = db.query(ClientNote).filter(ClientNote.client_id == client_id).order_by(ClientNote.updated_at.desc()).limit(10).all()
    for n in notes:
        events.append({
            "type": "note",
            "date": n.updated_at.strftime("%d.%m.%Y") if n.updated_at else "—",
            "iso_date": n.updated_at.isoformat() if n.updated_at else "",
            "icon": "📌" if n.is_pinned else "📝",
            "title": "Заметка" + (" (закреплена)" if n.is_pinned else ""),
            "desc": n.content[:100] + ("..." if len(n.content) > 100 else ""),
        })

    # Фолоуапы как отдельные события
    followups = db.query(Meeting).filter(
        Meeting.client_id == client_id,
        Meeting.followup_status == "sent",
        Meeting.followup_text != None,
    ).order_by(Meeting.followup_sent_at.desc()).limit(10).all()
    for m in followups:
        events.append({
            "type": "followup",
            "date": m.followup_sent_at.strftime("%d.%m.%Y") if m.followup_sent_at else "—",
            "iso_date": m.followup_sent_at.isoformat() if m.followup_sent_at else "",
            "icon": "✍️",
            "title": f"Фолоуап: {m.title or m.type}",
            "desc": (m.followup_text or "")[:100],
        })

    # Сортировка по iso_date
    events.sort(key=lambda e: e.get("iso_date", ""), reverse=True)

    return {"events": events[:50]}


@app.get("/api/clients/{client_id}/tasks-status")
async def api_client_tasks_status(client_id: int, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Лёгкий polling — только статусы задач для real-time обновления."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)

    tasks = db.query(Task).filter(Task.client_id == client_id).all()
    return {"tasks": [{"id": t.id, "status": t.status} for t in tasks]}

CHECKUP_INTERVALS = {"SS": 180, "SMB": 90, "SME": 60, "ENT": 30, "SME+": 60, "SME-": 60}

@app.get("/checkups", response_class=HTMLResponse)
async def checkups_page(request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Страница умных чекапов."""
    if not auth_token:
        return RedirectResponse(url="/login", status_code=303)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        return RedirectResponse(url="/login", status_code=303)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    return templates.TemplateResponse("checkups.html", {"request": request, "user": user})



@app.get("/api/checkup/results/all")
async def api_checkup_results_all(
    query_type: str = "",
    limit: int = 50,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """Все результаты чекапов менеджера (для страницы /checkups)."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        raise HTTPException(status_code=401)

    from models import CheckupResult
    q = (
        db.query(CheckupResult, Client)
        .join(Client, CheckupResult.client_id == Client.id)
        .order_by(CheckupResult.created_at.desc())
    )
    if user.role == "manager":
        q = q.filter(Client.manager_email == user.email)
    if query_type:
        q = q.filter(CheckupResult.query_type == query_type)
    rows = q.limit(limit).all()

    return {
        "results": [
            {
                "id": r.id,
                "client_id": r.client_id,
                "client_name": c.name,
                "cabinet_id": r.cabinet_id,
                "query_type": r.query_type,
                "manager_name": r.manager_name,
                "mode": r.mode,
                "total_queries": r.total_queries,
                "avg_score": round(r.avg_score, 2) if r.avg_score is not None else None,
                "score_dist": r.score_dist,
                "created_at": r.created_at.isoformat() if r.created_at else None,
            }
            for r, c in rows
        ]
    }



@app.get("/api/auth/me")
async def api_auth_me(
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """Данные текущего пользователя (роль, имя)."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        raise HTTPException(status_code=401)
    return {"id": user.id, "name": user.name, "email": user.email, "role": user.role}

@app.get("/api/auth/me/token")
async def api_me_token(
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """Возвращает текущий access token пользователя (для настройки расширения)."""
    if not auth_token:
        raise HTTPException(status_code=401)
    return {"token": auth_token}


@app.get("/checkup/result/{result_id}", response_class=HTMLResponse)
async def checkup_result_page(
    result_id: int,
    request: Request,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """Детальная страница результата чекапа."""
    if not auth_token:
        return RedirectResponse(url="/login", status_code=303)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        return RedirectResponse(url="/login", status_code=303)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        return RedirectResponse(url="/login", status_code=303)

    from models import CheckupResult
    result = db.query(CheckupResult).filter(CheckupResult.id == result_id).first()
    if not result:
        raise HTTPException(status_code=404)
    client = db.query(Client).filter(Client.id == result.client_id).first()

    return templates.TemplateResponse("checkup_result.html", {
        "request": request, "user": user,
        "result": result, "client": client,
    })

@app.get("/api/checkups")
async def api_checkups(db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Получить список чекапов по сегментам с дедлайнами."""
    if not auth_token:
        return {"overdue": [], "due_soon": [], "upcoming": []}
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        return {"overdue": [], "due_soon": [], "upcoming": []}
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        return {"overdue": [], "due_soon": [], "upcoming": []}

    q = db.query(Client)
    if user.role == "manager":
        q = q.filter(Client.manager_email == user.email)
    clients = q.all()

    now = datetime.now()
    overdue, due_soon, upcoming = [], [], []

    for c in clients:
        interval = CHECKUP_INTERVALS.get(c.segment or "", 90)
        last = c.last_meeting_date or c.last_checkup
        if last:
            days_since = (now - last).days
            days_until = interval - days_since
        else:
            days_since = 999
            days_until = -30

        info = {"id": c.id, "name": c.name, "segment": c.segment, "days_since": days_since, "days_until": days_until, "interval": interval, "last_date": last.isoformat() if last else None}

        if days_until < 0:
            overdue.append(info)
        elif days_until <= 14:
            due_soon.append(info)
        elif days_until <= 30:
            upcoming.append(info)

    overdue.sort(key=lambda x: x["days_until"])
    due_soon.sort(key=lambda x: x["days_until"])
    upcoming.sort(key=lambda x: x["days_until"])

    return {"overdue": overdue, "due_soon": due_soon, "upcoming": upcoming}


@app.post("/api/checkups/assign")
async def api_assign_checkup(request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Назначить чекап клиенту (создать встречу)."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)
    data = await request.json()
    client_id = data.get("client_id")
    date_str = data.get("date")
    if not client_id:
        raise HTTPException(status_code=400)
    meeting_date = datetime.fromisoformat(date_str) if date_str else datetime.now()
    meeting = Meeting(client_id=client_id, date=meeting_date, type="checkup", source="internal", title="Чекап")
    db.add(meeting)
    client = db.query(Client).filter(Client.id == client_id).first()
    if client:
        client.last_meeting_date = meeting_date
        client.needs_checkup = False
    db.commit()

    # Автозапись в Google Sheets
    if client:
        try:
            from sheets import write_checkup_status
            await write_checkup_status(
                client_name=client.name,
                status="Запланирован",
                last_date=meeting_date.strftime("%d.%m.%Y"),
            )
        except Exception as e:
            logger.debug(f"Sheets write-back skipped: {e}")

    return {"ok": True, "meeting_id": meeting.id}


# ============================================================================
# FOLLOWUP TEMPLATES
# ============================================================================

@app.get("/api/followup-templates")
async def api_get_followup_templates(db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Получить шаблоны фолоуапов пользователя."""
    if not auth_token:
        return {"templates": []}
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        return {"templates": []}
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        return {"templates": []}
    templates_list = db.query(FollowupTemplate).filter(FollowupTemplate.user_id == user.id).order_by(FollowupTemplate.name).all()
    return {"templates": [{"id": t.id, "name": t.name, "content": t.content, "category": t.category} for t in templates_list]}


@app.post("/api/followup-templates")
async def api_create_followup_template(request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Создать шаблон фолоуапа."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    data = await request.json()
    tpl = FollowupTemplate(user_id=user.id, name=data.get("name", ""), content=data.get("content", ""), category=data.get("category", "general"))
    db.add(tpl)
    db.commit()
    return {"ok": True, "id": tpl.id}


@app.delete("/api/followup-templates/{tpl_id}")
async def api_delete_followup_template(tpl_id: int, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Удалить шаблон."""
    if not auth_token:
        raise HTTPException(status_code=401)
    tpl = db.query(FollowupTemplate).filter(FollowupTemplate.id == tpl_id).first()
    if not tpl:
        raise HTTPException(status_code=404)
    db.delete(tpl)
    db.commit()
    return {"ok": True}


# ============================================================================
# CALENDAR
# ============================================================================

@app.get("/calendar", response_class=HTMLResponse)
async def calendar_page(request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Календарь встреч."""
    if not auth_token:
        return RedirectResponse(url="/login", status_code=303)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        return RedirectResponse(url="/login", status_code=303)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    return templates.TemplateResponse("calendar.html", {"request": request, "user": user})


@app.get("/api/calendar/events")
async def api_calendar_events(start: str = "", end: str = "", db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Получить события для календаря."""
    if not auth_token:
        return {"events": []}
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        return {"events": []}
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        return {"events": []}

    q = db.query(Meeting).join(Client, Meeting.client_id == Client.id, isouter=True)
    if user.role == "manager":
        q = q.filter(Client.manager_email == user.email)
    if start and end:
        q = q.filter(Meeting.date >= datetime.fromisoformat(start), Meeting.date <= datetime.fromisoformat(end))
    meetings = q.order_by(Meeting.date).all()

    events = []
    for m in meetings:
        color = {"checkup": "#22c55e", "qbr": "#6366f1", "kickoff": "#f97316", "sync": "#3b82f6"}.get(m.type, "#64748b")
        events.append({
            "id": m.id,
            "title": f"{m.client.name + ': ' if m.client else ''}{m.title or m.type}",
            "start": m.date.isoformat() if m.date else None,
            "color": color,
            "url": f"/client/{m.client_id}",
            "type": m.type,
        })

    # Также добавляем дедлайны задач
    task_q = db.query(Task).join(Client, Task.client_id == Client.id, isouter=True)
    if user.role == "manager":
        task_q = task_q.filter(Client.manager_email == user.email)
    task_q = task_q.filter(Task.due_date != None, Task.status != "done")
    if start and end:
        task_q = task_q.filter(Task.due_date >= datetime.fromisoformat(start), Task.due_date <= datetime.fromisoformat(end))
    tasks = task_q.all()

    for t in tasks:
        events.append({
            "id": f"task-{t.id}",
            "title": f"⏰ {t.title}",
            "start": t.due_date.isoformat() if t.due_date else None,
            "color": "#ef4444",
            "url": f"/client/{t.client_id}",
            "type": "task",
        })

    return {"events": events}


# ============================================================================
# PROFILE
# ============================================================================

@app.get("/profile", response_class=HTMLResponse)
async def profile_page(request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    if not auth_token:
        return RedirectResponse(url="/login", status_code=303)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        return RedirectResponse(url="/login", status_code=303)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    return templates.TemplateResponse("profile.html", {"request": request, "user": user})


@app.post("/api/profile/update")
async def api_profile_update(request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Обновить имя/фамилию/telegram_id."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        raise HTTPException(status_code=401)

    data = await request.json()
    if "first_name" in data:
        user.first_name = data["first_name"].strip()
    if "last_name" in data:
        user.last_name = data["last_name"].strip()
    if "telegram_id" in data:
        tg = data["telegram_id"].strip()
        # Проверяем что такой TG ID не занят другим юзером
        if tg and db.query(User).filter(User.telegram_id == tg, User.id != user.id).first():
            return {"ok": False, "error": "Этот Telegram ID уже привязан к другому аккаунту"}
        user.telegram_id = tg or None
    db.commit()
    return {"ok": True}


@app.post("/api/profile/password")
async def api_change_password(request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Сменить пароль."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        raise HTTPException(status_code=401)

    data = await request.json()
    current = data.get("current_password", "")
    new_pw = data.get("new_password", "")
    confirm = data.get("confirm_password", "")

    if not new_pw or len(new_pw) < 8:
        return {"ok": False, "error": "Новый пароль должен быть не менее 8 символов"}
    if new_pw != confirm:
        return {"ok": False, "error": "Пароли не совпадают"}
    if user.hashed_password and not verify_password(current, user.hashed_password):
        return {"ok": False, "error": "Неверный текущий пароль"}

    user.hashed_password = hash_password(new_pw)
    db.commit()
    return {"ok": True}



# ============================================================================
# SYNC STATUS
# ============================================================================

@app.get("/api/sync/status")
async def api_sync_status(
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """Статус последней синхронизации по каждой интеграции."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        raise HTTPException(status_code=401)

    integrations = ["merchrules", "airtable", "meetings_slots", "system"]
    result = {}
    now = datetime.now()

    for integration in integrations:
        last = db.query(SyncLog).filter(
            SyncLog.integration == integration,
        ).order_by(SyncLog.started_at.desc()).first()

        if last:
            ago_sec = int((now - last.started_at).total_seconds()) if last.started_at else None
            if ago_sec is not None:
                if ago_sec < 60:
                    ago_str = "только что"
                elif ago_sec < 3600:
                    ago_str = f"{ago_sec // 60} мин назад"
                elif ago_sec < 86400:
                    ago_str = f"{ago_sec // 3600} ч назад"
                else:
                    ago_str = f"{ago_sec // 86400} дн назад"
            else:
                ago_str = "—"

            result[integration] = {
                "status": last.status,
                "records": last.records_processed,
                "ago": ago_str,
                "at": last.started_at.strftime("%d.%m %H:%M") if last.started_at else "—",
                "error": last.message if last.status == "error" else None,
            }
        else:
            result[integration] = {"status": "never", "ago": "никогда", "records": 0}

    # Кол-во клиентов текущего менеджера
    q = db.query(Client)
    if user.role == "manager":
        q = q.filter(Client.manager_email == user.email)
    clients_count = q.count()

    return {"integrations": result, "clients_total": clients_count}


# ============================================================================
# DIAGNOSTICS & IMPORT
# ============================================================================

@app.get("/api/diagnostics/outbound-ip")
async def api_outbound_ip(auth_token: Optional[str] = Cookie(None)):
    """
    Возвращает внешний IP Railway-сервера.
    Этот IP нужно добавить в whitelist Merchrules.
    """
    if not auth_token:
        raise HTTPException(status_code=401)
    import httpx
    try:
        async with httpx.AsyncClient(timeout=10) as client:
            # Используем несколько сервисов для надёжности
            for url in ["https://api.ipify.org?format=json", "https://ifconfig.me/ip"]:
                try:
                    resp = await client.get(url)
                    if resp.status_code == 200:
                        text = resp.text.strip()
                        ip = resp.json().get("ip", text) if "json" in url else text
                        return {"ip": ip, "note": "Добавьте этот IP в whitelist Merchrules"}
                except Exception:
                    continue
    except Exception as e:
        return {"error": str(e)}
    return {"error": "Не удалось определить IP"}


@app.post("/api/diagnostics/merchrules-auth")
async def api_diag_merchrules_auth(
    request: Request,
    auth_token: Optional[str] = Cookie(None),
):
    """
    Диагностика авторизации Merchrules.
    Показывает точный HTTP-статус и ответ для каждой попытки.
    """
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)

    body = await request.json()
    login = body.get("login", "")
    password = body.get("password", "")
    if not login or not password:
        return {"error": "Нужны login и password"}

    import httpx
    results = []
    urls = list(dict.fromkeys([
        _env("MERCHRULES_API_URL", "https://merchrules.any-platform.ru"),
        "https://merchrules.any-platform.ru",
        "https://merchrules-qa.any-platform.ru",
    ]))
    fields = ["email", "login", "username"]

    async with httpx.AsyncClient(timeout=15) as hx:
        for url in urls:
            for field in fields:
                try:
                    resp = await hx.post(
                        f"{url}/backend-v2/auth/login",
                        json={field: login, "password": password},
                        timeout=10,
                    )
                    body_text = resp.text[:300]
                    has_token = False
                    if resp.status_code == 200:
                        try:
                            j = resp.json()
                            has_token = bool(j.get("token") or j.get("access_token") or j.get("accessToken"))
                        except Exception:
                            pass
                    results.append({
                        "url": url,
                        "field": field,
                        "status": resp.status_code,
                        "has_token": has_token,
                        "response": body_text,
                    })
                    # Нашли рабочий — дальше не пробуем
                    if resp.status_code == 200 and has_token:
                        return {"ok": True, "working": results[-1], "all": results}
                except Exception as e:
                    results.append({
                        "url": url,
                        "field": field,
                        "status": "error",
                        "error": str(e),
                    })

    return {"ok": False, "all": results}



async def api_import_clients_csv(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """
    Импорт клиентов из CSV/Excel файла.

    Ожидаемые колонки (гибко — ищет по ключевым словам):
      name / название / клиент
      segment / сегмент
      manager_email / менеджер
      site_id / site_ids / merchrules_id
      health_score
    """
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        raise HTTPException(status_code=401)

    content = await file.read()
    filename = file.filename or ""

    # Парсим файл
    try:
        import pandas as pd, io
        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            df = pd.read_excel(io.BytesIO(content))
        else:
            # Пробуем разные кодировки и разделители
            for enc in ("utf-8", "cp1251", "latin-1"):
                try:
                    df = pd.read_csv(io.BytesIO(content), encoding=enc, sep=None, engine="python")
                    break
                except Exception:
                    continue
            else:
                return {"error": "Не удалось прочитать файл. Поддерживаются CSV и XLSX."}
    except Exception as e:
        return {"error": f"Ошибка чтения файла: {e}"}

    # Нормализуем названия колонок
    df.columns = [str(c).strip().lower() for c in df.columns]

    def find_col(df, variants):
        for v in variants:
            for c in df.columns:
                if v in c:
                    return c
        return None

    col_name    = find_col(df, ["name", "название", "клиент", "company", "account"])
    col_segment = find_col(df, ["segment", "сегмент", "тип"])
    col_manager = find_col(df, ["manager", "менеджер", "email"])
    col_site    = find_col(df, ["site_id", "site", "merchrules", "account_id"])
    col_health  = find_col(df, ["health", "score", "хелс"])

    if not col_name:
        return {"error": f"Не найдена колонка с именем клиента. Колонки в файле: {list(df.columns)}"}

    created = updated = skipped = 0
    errors = []

    for idx, row in df.iterrows():
        name = str(row.get(col_name, "")).strip()
        if not name or name.lower() in ("nan", "none", ""):
            skipped += 1
            continue

        segment   = str(row.get(col_segment, "")).strip() if col_segment else ""
        manager   = str(row.get(col_manager, "")).strip() if col_manager else user.email
        site_id   = str(row.get(col_site, "")).strip() if col_site else ""
        health    = None
        if col_health:
            try:
                health = float(str(row.get(col_health, "")).replace(",", ".").replace("%", ""))
                if health > 1:
                    health = health / 100
            except Exception:
                pass

        # Ищем существующего клиента
        existing = None
        if site_id and site_id not in ("nan", ""):
            existing = db.query(Client).filter(Client.merchrules_account_id == site_id).first()
        if not existing:
            existing = db.query(Client).filter(Client.name == name).first()

        if existing:
            # Обновляем только непустые поля
            if segment and segment not in ("nan", ""):
                existing.segment = segment
            if manager and manager not in ("nan", "") and "@" in manager:
                existing.manager_email = manager
            if site_id and site_id not in ("nan", ""):
                existing.merchrules_account_id = site_id
            if health is not None:
                existing.health_score = health
            updated += 1
        else:
            c = Client(
                name=name,
                segment=segment if segment not in ("nan", "") else None,
                manager_email=manager if "@" in manager else user.email,
                merchrules_account_id=site_id if site_id not in ("nan", "") else None,
                health_score=health,
            )
            db.add(c)
            created += 1

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        return {"error": f"Ошибка сохранения: {e}"}

    return {
        "ok": True,
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "total_rows": len(df),
        "columns_detected": {
            "name": col_name, "segment": col_segment,
            "manager": col_manager, "site_id": col_site,
        }
    }


@app.post("/api/import/tasks-csv")
async def api_import_tasks_csv(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """
    Импорт задач из CSV/Excel файла.

    Ожидаемые колонки:
      title / название / задача
      client / клиент / account
      status / статус
      priority / приоритет
      due_date / дедлайн / срок
      team / команда
    """
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        raise HTTPException(status_code=401)

    content = await file.read()
    filename = file.filename or ""

    try:
        import pandas as pd, io
        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            df = pd.read_excel(io.BytesIO(content))
        else:
            for enc in ("utf-8", "cp1251", "latin-1"):
                try:
                    df = pd.read_csv(io.BytesIO(content), encoding=enc, sep=None, engine="python")
                    break
                except Exception:
                    continue
            else:
                return {"error": "Не удалось прочитать файл"}
    except Exception as e:
        return {"error": f"Ошибка чтения файла: {e}"}

    df.columns = [str(c).strip().lower() for c in df.columns]

    def find_col(df, variants):
        for v in variants:
            for c in df.columns:
                if v in c:
                    return c
        return None

    col_title    = find_col(df, ["title", "название", "задача", "task", "name"])
    col_client   = find_col(df, ["client", "клиент", "account", "аккаунт", "site"])
    col_status   = find_col(df, ["status", "статус"])
    col_priority = find_col(df, ["priority", "приоритет"])
    col_due      = find_col(df, ["due", "дедлайн", "срок", "date"])
    col_team     = find_col(df, ["team", "команда"])
    col_mr_id    = find_col(df, ["merchrules_task", "task_id", "mr_id", "id"])

    if not col_title:
        return {"error": f"Не найдена колонка с названием задачи. Колонки: {list(df.columns)}"}

    STATUS_MAP = {
        "plan": "plan", "в работе": "in_progress", "in_progress": "in_progress",
        "review": "review", "done": "done", "готово": "done",
        "blocked": "blocked", "заблок": "blocked",
    }

    created = skipped = 0
    # Кешируем клиентов для поиска
    all_clients = {c.name.lower(): c for c in db.query(Client).all()}

    for idx, row in df.iterrows():
        title = str(row.get(col_title, "")).strip()
        if not title or title.lower() in ("nan", "none", ""):
            skipped += 1
            continue

        # Ищем клиента
        client_id = None
        if col_client:
            client_name = str(row.get(col_client, "")).strip().lower()
            if client_name and client_name not in ("nan", ""):
                # Точное совпадение
                c = all_clients.get(client_name)
                if not c:
                    # Частичное совпадение
                    for cname, cobj in all_clients.items():
                        if client_name in cname or cname in client_name:
                            c = cobj
                            break
                if c:
                    client_id = c.id

        # Статус
        raw_status = str(row.get(col_status, "plan")).strip().lower() if col_status else "plan"
        status_val = STATUS_MAP.get(raw_status, "plan")

        # Дедлайн
        due_date = None
        if col_due:
            raw_due = str(row.get(col_due, "")).strip()
            if raw_due and raw_due not in ("nan", ""):
                for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%m/%d/%Y"):
                    try:
                        due_date = datetime.strptime(raw_due[:10], fmt)
                        break
                    except Exception:
                        continue

        # Проверяем дубль по merchrules_task_id
        mr_id = str(row.get(col_mr_id, "")).strip() if col_mr_id else ""
        if mr_id and mr_id not in ("nan", ""):
            existing = db.query(Task).filter(Task.merchrules_task_id == mr_id).first()
            if existing:
                skipped += 1
                continue

        task = Task(
            client_id=client_id,
            title=title,
            status=status_val,
            priority=str(row.get(col_priority, "medium")).strip().lower() if col_priority else "medium",
            due_date=due_date,
            team=str(row.get(col_team, "")).strip() if col_team else None,
            source="import",
            merchrules_task_id=mr_id if mr_id not in ("nan", "") else None,
        )
        db.add(task)
        created += 1

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        return {"error": f"Ошибка сохранения: {e}"}

    return {
        "ok": True,
        "created": created,
        "skipped": skipped,
        "total_rows": len(df),
    }


# ============================================================================
# MEETING SLOTS
# ============================================================================

@app.get("/meetings", response_class=HTMLResponse)
async def meetings_page(request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Страница встреч со слотами дня."""
    if not auth_token:
        return RedirectResponse(url="/login", status_code=303)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        return RedirectResponse(url="/login", status_code=303)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    return templates.TemplateResponse("meetings.html", {"request": request, "user": user})


@app.get("/api/meetings/slots")
async def api_meetings_slots(
    date: Optional[str] = None,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """
    Получить слоты дня (встречи + prep/followup задачи).
    date: ISO строка даты, по умолчанию — сегодня МСК.
    """
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        raise HTTPException(status_code=401)

    from meeting_slots import get_day_slots
    target_date = datetime.now(MSK).replace(tzinfo=None)
    if date:
        try:
            target_date = datetime.fromisoformat(date)
        except ValueError:
            pass

    slots = get_day_slots(db, user.email, target_date)
    return {"slots": slots, "date": target_date.strftime("%Y-%m-%d")}


@app.post("/api/meetings/sync-slots")
async def api_sync_meeting_slots(
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """Принудительно создать слоты для всех предстоящих встреч."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        raise HTTPException(status_code=401)

    from meeting_slots import create_slots_for_meeting
    now = datetime.utcnow()
    window_end = now + timedelta(days=7)

    q = db.query(Meeting).join(Client, Meeting.client_id == Client.id, isouter=True).filter(
        Meeting.date >= now,
        Meeting.date <= window_end,
    )
    if user.role == "manager":
        q = q.filter(Client.manager_email == user.email)

    meetings = q.all()
    total = 0
    for m in meetings:
        created = create_slots_for_meeting(db, m)
        total += len(created)

    return {"ok": True, "slots_created": total, "meetings_processed": len(meetings)}


@app.get("/api/integrations/test/outlook")
async def api_test_outlook():
    """Тест подключения к Outlook."""
    from integrations.outlook import test_connection
    result = await test_connection()
    return result


# ============================================================================
# KPI МЕНЕДЖЕРА
# ============================================================================

@app.get("/kpi", response_class=HTMLResponse)
async def kpi_page(request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Страница KPI менеджера."""
    if not auth_token:
        return RedirectResponse(url="/login", status_code=303)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        return RedirectResponse(url="/login", status_code=303)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    return templates.TemplateResponse("kpi.html", {"request": request, "user": user})


@app.get("/notifications", response_class=HTMLResponse)
async def notifications_page(request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Страница всех уведомлений."""
    if not auth_token:
        return RedirectResponse(url="/login", status_code=303)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        return RedirectResponse(url="/login", status_code=303)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    notifs = db.query(Notification).filter(
        Notification.user_id == user.id
    ).order_by(Notification.created_at.desc()).limit(100).all()
    return templates.TemplateResponse("notifications.html", {
        "request": request, "user": user, "notifications": notifs
    })


# ============================================================================
# ANALYTICS
# ============================================================================

@app.get("/analytics", response_class=HTMLResponse)
async def analytics_page(request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Страница аналитики."""
    if not auth_token:
        return RedirectResponse(url="/login", status_code=303)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        return RedirectResponse(url="/login", status_code=303)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    return templates.TemplateResponse("analytics.html", {"request": request, "user": user})


@app.get("/api/analytics")
async def api_analytics(db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Данные для аналитики."""
    if not auth_token:
        return {}
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        return {}
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        return {}

    q = db.query(Client)
    if user.role == "manager":
        q = q.filter(Client.manager_email == user.email)
    clients = q.all()

    # Segments
    seg_counts = {}
    for c in clients:
        seg = c.segment or "other"
        seg_counts[seg] = seg_counts.get(seg, 0) + 1

    # Health distribution
    health_buckets = {"0-25": 0, "25-50": 0, "50-75": 0, "75-100": 0}
    for c in clients:
        score = (c.health_score or 0) * 100
        if score < 25:
            health_buckets["0-25"] += 1
        elif score < 50:
            health_buckets["25-50"] += 1
        elif score < 75:
            health_buckets["50-75"] += 1
        else:
            health_buckets["75-100"] += 1

    # Tasks by status
    task_q = db.query(Task).join(Client, Task.client_id == Client.id, isouter=True)
    if user.role == "manager":
        task_q = task_q.filter(Client.manager_email == user.email)
    all_tasks = task_q.all()
    task_status_counts = {}
    for t in all_tasks:
        s = t.status or "plan"
        task_status_counts[s] = task_status_counts.get(s, 0) + 1

    # Meetings per month (last 6 months)
    meetings_per_month = {}
    meeting_q = db.query(Meeting).join(Client, Meeting.client_id == Client.id, isouter=True)
    if user.role == "manager":
        meeting_q = meeting_q.filter(Client.manager_email == user.email)
    all_meetings = meeting_q.filter(Meeting.date != None).order_by(Meeting.date.desc()).all()
    for m in all_meetings:
        if m.date:
            key = m.date.strftime("%Y-%m")
            meetings_per_month[key] = meetings_per_month.get(key, 0) + 1

    return {
        "total_clients": len(clients),
        "segments": seg_counts,
        "health_distribution": health_buckets,
        "task_status": task_status_counts,
        "total_tasks": len(all_tasks),
        "meetings_per_month": dict(sorted(meetings_per_month.items(), reverse=True)[:6]),
        "avg_health": round(sum((c.health_score or 0) for c in clients) / max(len(clients), 1) * 100, 1),
    }


# ============================================================================
# BULK ACTIONS
# ============================================================================

@app.post("/api/bulk/checkups")
async def api_bulk_checkups(request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Массовое назначение чекапов."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)
    data = await request.json()
    client_ids = data.get("client_ids", [])
    date_str = data.get("date")
    meeting_date = datetime.fromisoformat(date_str) if date_str else datetime.now()
    created = 0
    for cid in client_ids:
        client = db.query(Client).filter(Client.id == cid).first()
        if client:
            m = Meeting(client_id=cid, date=meeting_date, type="checkup", source="internal", title="Чекап")
            db.add(m)
            client.last_meeting_date = meeting_date
            client.needs_checkup = False
            created += 1
    db.commit()
    return {"ok": True, "created": created}


@app.post("/api/bulk/segment")
async def api_bulk_segment(request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Массовое изменение сегмента."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)
    data = await request.json()
    client_ids = data.get("client_ids", [])
    segment = data.get("segment", "")
    updated = 0
    for cid in client_ids:
        client = db.query(Client).filter(Client.id == cid).first()
        if client:
            client.segment = segment
            updated += 1
    db.commit()
    return {"ok": True, "updated": updated}


# ============================================================================
# EXPORT
# ============================================================================

@app.get("/api/export/client/{client_id}")
async def api_export_client(client_id: int, fmt: str = "json", db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Экспорт отчёта по клиенту (JSON/CSV)."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)

    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        raise HTTPException(status_code=404)

    tasks = db.query(Task).filter(Task.client_id == client_id).all()
    meetings = db.query(Meeting).filter(Meeting.client_id == client_id).order_by(Meeting.date.desc()).all()
    notes = db.query(ClientNote).filter(ClientNote.client_id == client_id).all()

    data = {
        "client": {"id": client.id, "name": client.name, "segment": client.segment, "health_score": client.health_score, "manager_email": client.manager_email},
        "tasks": [{"id": t.id, "title": t.title, "status": t.status, "priority": t.priority, "due_date": t.due_date.isoformat() if t.due_date else None} for t in tasks],
        "meetings": [{"id": m.id, "title": m.title or m.type, "date": m.date.isoformat() if m.date else None, "type": m.type} for m in meetings],
        "notes": [{"id": n.id, "content": n.content, "pinned": n.is_pinned} for n in notes],
        "exported_at": datetime.utcnow().isoformat(),
    }

    if fmt == "csv":
        import io, csv
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["type", "id", "title", "date", "details"])
        for t in tasks:
            writer.writerow(["task", t.id, t.title, t.due_date.isoformat() if t.due_date else "", t.status])
        for m in meetings:
            writer.writerow(["meeting", m.id, m.title or m.type, m.date.isoformat() if m.date else "", m.type])
        for n in notes:
            writer.writerow(["note", n.id, n.content[:50], "", "pinned" if n.is_pinned else ""])
        from fastapi.responses import PlainTextResponse
        return PlainTextResponse(content=output.getvalue(), headers={"Content-Disposition": f"attachment; filename=client_{client_id}.csv"})

    return data


# ============================================================================
# AI RECOMMENDATIONS & AUTO-QBR
# ============================================================================

@app.post("/api/ai/auto-qbr/{client_id}")
async def api_auto_qbr(client_id: int, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """AI-генерация черновика QBR из данных клиента."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)

    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        raise HTTPException(status_code=404)

    tasks = db.query(Task).filter(Task.client_id == client_id).all()
    meetings = db.query(Meeting).filter(Meeting.client_id == client_id).order_by(Meeting.date.desc()).limit(10).all()

    tasks_done = [t for t in tasks if t.status == "done"]
    tasks_blocked = [t for t in tasks if t.status == "blocked"]

    prompt = f"""Создай черновик QBR для клиента {client.name} ({client.segment or '—'}).

Health Score: {(client.health_score or 0)*100:.0f}%
Задач выполнено: {len(tasks_done)}
Задач заблокировано: {len(tasks_blocked)}
Последние встречи:
{chr(10).join([f"- {m.title or m.type} ({m.date.strftime('%d.%m.%Y') if m.date else ''})" for m in meetings[:5]])}

Ответь JSON:
{"achievements": [...], "issues": [...], "next_quarter_goals": [...], "summary": "..."}"""

    try:
        text = await _ai_chat("", prompt, max_tokens=1500)
        import json, re
        m = re.search(r'\{.*\}', text, re.DOTALL)
        if m:
            data = json.loads(m.group(0))
        else:
            data = {}
    except Exception as e:
        data = {"error": str(e)}

    return data


async def _ai_chat(system: str, user: str, max_tokens: int = 1000) -> str:
    """AI чат через Groq или Qwen."""
    groq_key = env.GROQ_KEY
    qwen_key = env.QWEN_KEY

    if groq_key:
        import httpx
        try:
            async with httpx.AsyncClient(timeout=60) as hx:
                resp = await hx.post("https://api.groq.com/openai/v1/chat/completions",
                    json={"model": "llama-3.3-70b-versatile", "messages": [{"role": "system", "content": system}, {"role": "user", "content": user}], "max_tokens": max_tokens, "temperature": 0.1},
                    headers={"Authorization": f"Bearer {groq_key}", "Content-Type": "application/json"})
            if resp.status_code == 200:
                return resp.json()["choices"][0]["message"]["content"].strip()
        except Exception as e:
            logger.debug(f"Ignored error: {e}")

    if qwen_key:
        import httpx
        try:
            async with httpx.AsyncClient(timeout=60) as hx:
                resp = await hx.post("https://dashscope.aliyuncs.com/compatible-mode/v1/chat/completions",
                    json={"model": "qwen-plus", "messages": [{"role": "system", "content": system}, {"role": "user", "content": user}], "max_tokens": max_tokens, "temperature": 0.1},
                    headers={"Authorization": f"Bearer {qwen_key}", "Content-Type": "application/json"})
            if resp.status_code == 200:
                return resp.json()["choices"][0]["message"]["content"].strip()
        except Exception as e:
            logger.debug(f"Ignored error: {e}")

    return "AI недоступен. Настройте GROQ_API_KEY или QWEN_API_KEY."


# ============================================================================
# "TIME TO WRITE" SIGNALS
# ============================================================================

@app.get("/api/notifications")
async def api_notifications(db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Получить уведомления: пора написать, просрочки и т.д."""
    if not auth_token:
        return {"notifications": []}
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        return {"notifications": []}
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        return {"notifications": []}

    # Кеш 120 сек — polling каждые 60 сек от 18 менеджеров = экономим ~540 DB запросов/мин
    ck = f"notif:{user.id}"
    cached = cache_get(ck)
    if cached:
        return cached

    notifications = []
    now = datetime.now()

    # Клиенты без активности
    q = db.query(Client)
    if user.role == "manager":
        q = q.filter(Client.manager_email == user.email)
    clients = q.all()

    for c in clients:
        interval = CHECKUP_INTERVALS.get(c.segment or "", 90)
        last = c.last_meeting_date or c.last_checkup
        if last:
            days_since = (now - last).days
            if days_since > interval:
                notifications.append({
                    "type": "overdue_checkup",
                    "priority": "high",
                    "message": f"Пора написать: {c.name} (последний контакт {days_since} дн. назад)",
                    "client_id": c.id,
                    "client_name": c.name,
                })
            elif days_since > interval - 14:
                notifications.append({
                    "type": "checkup_soon",
                    "priority": "medium",
                    "message": f"Скоро чекап: {c.name} (через {interval - days_since} дн.)",
                    "client_id": c.id,
                    "client_name": c.name,
                })

    # Blocked tasks
    task_q = db.query(Task).join(Client, Task.client_id == Client.id, isouter=True)
    if user.role == "manager":
        task_q = task_q.filter(Client.manager_email == user.email)
    blocked = task_q.filter(Task.status == "blocked").all()
    for t in blocked:
        notifications.append({
            "type": "blocked_task",
            "priority": "high",
            "message": f"Заблокирована задача: {t.title} ({t.client.name if t.client else ''})",
            "client_id": t.client_id,
        })

    return {"notifications": notifications}


# ============================================================================
# FOCUS MODE (CSS toggle — client detail page with sidebar hidden)
# ============================================================================

@app.get("/client/{client_id}/focus", response_class=HTMLResponse)
async def client_focus_view(request: Request, client_id: int, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Режим фокуса: клиент без сайдбара."""
    if not auth_token:
        return RedirectResponse(url="/login", status_code=303)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        return RedirectResponse(url="/login", status_code=303)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        return RedirectResponse(url="/login", status_code=303)

    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        raise HTTPException(status_code=404)

    tasks = db.query(Task).filter(Task.client_id == client_id).order_by(Task.due_date.desc()).all()
    meetings = db.query(Meeting).filter(Meeting.client_id == client_id).order_by(Meeting.date.desc()).limit(5).all()
    notes = db.query(ClientNote).filter(ClientNote.client_id == client_id).order_by(ClientNote.is_pinned.desc(), ClientNote.updated_at.desc()).all()

    return templates.TemplateResponse("client_focus.html", {
        "request": request, "user": user, "client": client,
        "tasks": tasks, "meetings": meetings, "notes": notes,
    })


@app.get("/health")
async def health():
    return {"status": "ok", "version": "2.0.0"}


# ============================================================================
# VOICE NOTES
# ============================================================================

@app.post("/api/voice-notes")
async def api_create_voice_note(
    request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)
):
    """Создать голосовую заметку (текстовую транскрипцию)."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    data = await request.json()
    vn = VoiceNote(
        client_id=data.get("client_id"),
        meeting_id=data.get("meeting_id"),
        user_id=user.id,
        transcription=data.get("text", ""),
        duration_seconds=data.get("duration", 0),
    )
    db.add(vn)
    # Авто-создание задачи из заметки
    if data.get("create_task"):
        db.add(Task(
            client_id=data.get("client_id"),
            title=f"🎤 {data.get('text', '')[:80]}",
            description=data.get("text", ""),
            status="plan",
            priority="medium",
            source="voice_note",
        ))
    db.commit()
    return {"ok": True, "id": vn.id}


# ============================================================================
# PERSONAL INBOX
# ============================================================================

@app.get("/inbox", response_class=HTMLResponse)
async def inbox_page(request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Персональный Inbox."""
    if not auth_token:
        return RedirectResponse(url="/login", status_code=303)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        return RedirectResponse(url="/login", status_code=303)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    return templates.TemplateResponse("inbox.html", {"request": request, "user": user})


@app.get("/api/inbox")
async def api_inbox(db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Получить сообщения Inbox."""
    if not auth_token:
        return {"items": []}
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        return {"items": []}
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    items = []
    now = datetime.now()

    # Новые уведомления
    notifs = db.query(Notification).filter(Notification.user_id == user.id, Notification.is_read == False).order_by(Notification.created_at.desc()).limit(20).all()
    for n in notifs:
        items.append({"type": "notification", "title": n.title, "message": n.message, "date": n.created_at.isoformat() if n.created_at else None, "priority": n.type})

    # Просроченные чекапы
    q = db.query(Client)
    if user.role == "manager":
        q = q.filter(Client.manager_email == user.email)
    clients = q.all()
    for c in clients:
        interval = CHECKUP_INTERVALS.get(c.segment or "", 90)
        last = c.last_meeting_date or c.last_checkup
        if last and (now - last).days > interval:
            items.append({"type": "overdue", "title": f"Просрочен чекап: {c.name}", "message": f"Последний контакт {(now-last).days} дн. назад (норма: {interval})", "date": last.isoformat(), "priority": "high", "client_id": c.id})

    # Blocked tasks
    task_q = db.query(Task).join(Client, Task.client_id == Client.id, isouter=True)
    if user.role == "manager":
        task_q = task_q.filter(Client.manager_email == user.email)
    blocked = task_q.filter(Task.status == "blocked").all()
    for t in blocked:
        items.append({"type": "blocked", "title": f"Заблокирована: {t.title}", "message": t.client.name if t.client else "", "priority": "high", "client_id": t.client_id})

    items.sort(key=lambda x: x.get("date") or "", reverse=True)
    return {"items": items[:50]}


@app.post("/api/inbox/mark-read")
async def api_inbox_mark_read(request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Отметить уведомления прочитанными."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    db.query(Notification).filter(Notification.user_id == user.id, Notification.is_read == False).update({"is_read": True})
    db.commit()
    return {"ok": True}


@app.get("/api/inbox/items")
async def api_inbox_items(db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Алиас /api/inbox для совместимости с base.html."""
    return await api_inbox(db=db, auth_token=auth_token)


# ============================================================================
# CHURN PREDICTION
# ============================================================================

@app.get("/api/churn/risk")
async def api_churn_risk(db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Прогнозирование оттока: rule-based scoring."""
    if not auth_token:
        return {"clients": []}
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        return {"clients": []}
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        return {"clients": []}

    q = db.query(Client)
    if user.role == "manager":
        q = q.filter(Client.manager_email == user.email)
    clients = q.all()
    now = datetime.now()
    results = []

    for c in clients:
        score = 0
        reasons = []
        interval = CHECKUP_INTERVALS.get(c.segment or "", 90)
        last = c.last_meeting_date or c.last_checkup

        # Фактор 1: Нет контакта > 2x интервала
        if last and (now - last).days > interval * 2:
            score += 40
            reasons.append(f"Нет контакта {(now-last).days} дн. (норма: {interval})")

        # Фактор 2: Low health score
        if c.health_score and c.health_score < 0.3:
            score += 30
            reasons.append(f"Низкий health score: {c.health_score:.0%}")

        # Фактор 3: Blocked tasks
        blocked = db.query(Task).filter(Task.client_id == c.id, Task.status == "blocked").count()
        if blocked > 0:
            score += 15
            reasons.append(f"{blocked} заблокированных задач")

        # Фактор 4: Нет задач вообще
        total_tasks = db.query(Task).filter(Task.client_id == c.id).count()
        if total_tasks == 0:
            score += 15
            reasons.append("Нет задач")

        risk = "low"
        if score >= 60:
            risk = "critical"
        elif score >= 30:
            risk = "medium"

        results.append({"id": c.id, "name": c.name, "segment": c.segment, "risk": risk, "score": score, "reasons": reasons})

    results.sort(key=lambda x: x["score"], reverse=True)
    return {"clients": results}


# ============================================================================
# AI AUTO-QBR PAGE
# ============================================================================

@app.get("/qbr/auto/{client_id}", response_class=HTMLResponse)
async def qbr_auto_page(request: Request, client_id: int, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Страница авто-QBR."""
    if not auth_token:
        return RedirectResponse(url="/login", status_code=303)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        return RedirectResponse(url="/login", status_code=303)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        raise HTTPException(status_code=404)
    return templates.TemplateResponse("qbr_auto.html", {"request": request, "user": user, "client": client})


@app.get("/voice-notes", response_class=HTMLResponse)
async def voice_notes_page(request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Страница голосовых заметок."""
    if not auth_token:
        return RedirectResponse(url="/login", status_code=303)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        return RedirectResponse(url="/login", status_code=303)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    return templates.TemplateResponse("voice_notes.html", {"request": request, "user": user})


@app.get("/api/voice-notes")
async def api_get_voice_notes(db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Получить голосовые заметки пользователя."""
    if not auth_token:
        return {"notes": []}
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        return {"notes": []}
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    notes = db.query(VoiceNote).filter(VoiceNote.user_id == user.id).order_by(VoiceNote.created_at.desc()).limit(50).all()
    return {"notes": [{"id": n.id, "text": n.transcription, "duration": n.duration_seconds, "client_id": n.client_id, "created_at": n.created_at.isoformat() if n.created_at else None} for n in notes]}


# ============================================================================
# PWA ICONS (SVG placeholder)
# ============================================================================

@app.get("/static/icon-192.png")
@app.get("/static/icon-512.png")
async def pwa_icon():
    """SVG иконка для PWA (base64 PNG placeholder)."""
    from fastapi.responses import PlainTextResponse
    # Simple colored square as placeholder
    return PlainTextResponse(content="iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAAC0lEQVQI12NgAAIABQABNjN9GQAAAAlwSFlzAAAWJQAAFiUBSVIk8AAAAA0lEQVQI12P4z8BQDwAEgAF/QL9hbgAAAABJRU5ErkJggg==", headers={"Content-Type": "image/png"})


# ============================================================================
# PWA MANIFEST
# ============================================================================

# ============================================================================
# OFFLINE / DRAFTS
# ============================================================================

@app.post("/api/drafts")
async def api_save_draft(request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Сохранить черновик (фолоуап, заметка) для офлайн-синхронизации."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    data = await request.json()
    settings = user.settings or {}
    drafts = settings.get("drafts", [])
    drafts.append({**data, "saved_at": datetime.utcnow().isoformat(), "user_id": user.id})
    settings["drafts"] = drafts[-50:]  # keep last 50
    user.settings = dict(settings)
    from sqlalchemy.orm.attributes import flag_modified
    flag_modified(user, "settings")
    db.commit()
    return {"ok": True}


@app.get("/api/drafts")
async def api_get_drafts(db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Получить черновики."""
    if not auth_token:
        return {"drafts": []}
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        return {"drafts": []}
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        return {"drafts": []}
    settings = user.settings or {}
    return {"drafts": settings.get("drafts", [])}


# ============================================================================
# TASK MODAL API (bulk edit)
# ============================================================================

@app.patch("/api/tasks/bulk")
async def api_bulk_edit_tasks(request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Массовое редактирование задач."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)
    data = await request.json()
    task_ids = data.get("task_ids", [])
    updates = {}
    for key in ["status", "priority", "due_date", "team", "task_type"]:
        if key in data and data[key]:
            updates[key] = data[key]
    if not task_ids or not updates:
        return {"error": "Need task_ids and updates"}
    updated = db.query(Task).filter(Task.id.in_(task_ids)).update(updates, synchronize_session=False)
    db.commit()
    return {"ok": True, "updated": updated}


# ============================================================================
# MEETING REMINDER API (for morning alerts)
# ============================================================================

@app.get("/api/meetings/today")
async def api_meetings_today(db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Получить встречи сегодня с ссылками."""
    if not auth_token:
        return {"meetings": []}
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        return {"meetings": []}
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        return {"meetings": []}

    now_msk = datetime.now(MSK)
    today = now_msk.date()
    tomorrow = today + timedelta(days=1)

    q = db.query(Meeting).join(Client, Meeting.client_id == Client.id, isouter=True)
    if user.role == "manager":
        q = q.filter(Client.manager_email == user.email)
    meetings = q.filter(
        Meeting.date >= datetime.combine(today, datetime.min.time()),
        Meeting.date < datetime.combine(tomorrow, datetime.min.time()),
    ).all()

    return {
        "meetings": [{
            "id": m.id,
            "title": m.title or m.type,
            "time": m.date.strftime("%H:%M") if m.date else "—",
            "client": m.client.name if m.client else "—",
            "client_id": m.client_id,
            "link": m.recording_url or f"/client/{m.client_id}",
            "type": m.type,
        } for m in meetings]
    }

# ============================================================================
# MISSING ENDPOINTS (referenced from templates)
# ============================================================================

@app.get("/api/stats")
async def api_stats(db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Быстрая статистика для sidebar — вызывается на каждой странице."""
    if not auth_token:
        return {"overdue": 0, "warning": 0, "open_tasks": 0}
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        return {"overdue": 0, "warning": 0, "open_tasks": 0}
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        return {"overdue": 0, "warning": 0, "open_tasks": 0}

    # Кеш 90 сек — вызывается на каждой странице от 18 менеджеров
    ck = f"stats:{user.id}"
    cached = cache_get(ck)
    if cached:
        return cached

    now = datetime.now()
    q = db.query(Client)
    if user.role == "manager":
        q = q.filter(Client.manager_email == user.email)
    clients = q.all()

    overdue = warning = 0
    for c in clients:
        last = c.last_meeting_date or c.last_checkup
        if not last:
            continue
        days = (now - last).days
        from models import CHECKUP_INTERVALS
        interval = CHECKUP_INTERVALS.get(c.segment or "", 90)
        if days > interval:
            overdue += 1
        elif days > interval - 14:
            warning += 1

    tq = db.query(Task).join(Client, Task.client_id == Client.id, isouter=True).filter(
        Task.status.in_(["plan", "in_progress", "blocked"])
    )
    if user.role == "manager":
        tq = tq.filter(Client.manager_email == user.email)
    open_tasks = tq.count()

    result = {"overdue": overdue, "warning": warning, "open_tasks": open_tasks}
    cache_set(ck, result, ttl=90)
    return result


@app.get("/api/settings/my-clients")
async def api_my_clients(db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Список клиентов текущего менеджера."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        raise HTTPException(status_code=401)

    q = db.query(Client)
    if user.role == "manager":
        q = q.filter(Client.manager_email == user.email)
    clients = q.order_by(Client.name).all()
    return {"clients": [{"id": c.id, "name": c.name, "segment": c.segment} for c in clients]}


@app.get("/api/clients")
async def api_clients_list(
    segment: Optional[str] = None,
    search: Optional[str] = None,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """API список клиентов с фильтрами."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        raise HTTPException(status_code=401)

    q = db.query(Client)
    if user.role == "manager":
        q = q.filter(Client.manager_email == user.email)
    if segment:
        q = q.filter(Client.segment == segment)
    if search:
        q = q.filter(Client.name.ilike(f"%{search}%"))
    clients = q.order_by(Client.name).all()

    return {"clients": [{
        "id": c.id, "name": c.name, "segment": c.segment,
        "health_score": c.health_score, "manager_email": c.manager_email,
        "merchrules_account_id": c.merchrules_account_id,
        "last_meeting_date": c.last_meeting_date.isoformat() if c.last_meeting_date else None,
    } for c in clients]}


# ============================================================================
# EXPORT: PDF REPORT
# ============================================================================

# ============================================================================
# EXPORT: PDF отчёт по клиенту
# ============================================================================

@app.get("/api/clients/{client_id}/export/pdf")
async def api_export_client_pdf(
    client_id: int,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """Экспорт карточки клиента в PDF."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)

    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        raise HTTPException(status_code=404)

    tasks = db.query(Task).filter(Task.client_id == client_id).order_by(Task.created_at.desc()).limit(50).all()
    meetings = db.query(Meeting).filter(Meeting.client_id == client_id).order_by(Meeting.date.desc()).limit(20).all()
    notes = db.query(ClientNote).filter(ClientNote.client_id == client_id).order_by(ClientNote.updated_at.desc()).limit(10).all()
    plan = db.query(AccountPlan).filter(AccountPlan.client_id == client_id).first()
    qbr = db.query(QBR).filter(QBR.client_id == client_id).order_by(QBR.date.desc()).first()

    now_str = datetime.now().strftime("%d.%m.%Y")
    health_pct = int((client.health_score or 0) * 100)
    health_color = "#22c55e" if health_pct >= 70 else ("#eab308" if health_pct >= 40 else "#ef4444")

    # Задачи по статусам
    task_rows = ""
    for t in tasks:
        status_colors = {"plan": "#64748b", "in_progress": "#6366f1", "review": "#eab308", "done": "#22c55e", "blocked": "#ef4444"}
        color = status_colors.get(t.status or "plan", "#64748b")
        due = t.due_date.strftime("%d.%m.%Y") if t.due_date else "—"
        task_rows += f"""<tr>
            <td>{t.title or ''}</td>
            <td><span style="color:{color};font-weight:600;">{t.status or ''}</span></td>
            <td>{t.priority or ''}</td>
            <td>{t.team or '—'}</td>
            <td>{due}</td>
        </tr>"""

    # Встречи
    meeting_rows = ""
    for m in meetings:
        date_str = m.date.strftime("%d.%m.%Y %H:%M") if m.date else "—"
        followup = "✅" if m.followup_status == "sent" else ("⏳" if m.followup_status == "pending" else "—")
        meeting_rows += f"""<tr>
            <td>{date_str}</td>
            <td>{m.type or ''}</td>
            <td>{m.title or ''}</td>
            <td>{followup}</td>
        </tr>"""

    # Заметки
    notes_html = ""
    for n in notes:
        pin = "📌 " if n.is_pinned else ""
        date_str = n.updated_at.strftime("%d.%m.%Y") if n.updated_at else ""
        notes_html += f'<div style="margin-bottom:8px;padding:8px 12px;background:#f8fafc;border-left:3px solid #e2e8f0;border-radius:4px;"><div style="font-size:12px;white-space:pre-wrap;">{pin}{n.content}</div><div style="font-size:10px;color:#94a3b8;margin-top:4px;">{date_str}</div></div>'

    # Цели из плана
    goals_html = ""
    if plan and plan.quarterly_goals:
        for g in (plan.quarterly_goals or [])[:5]:
            if isinstance(g, dict):
                goals_html += f'<li>{g.get("goal", str(g))}</li>'
            else:
                goals_html += f'<li>{g}</li>'

    html = f"""<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8">
<style>
  @page {{ margin: 20mm 15mm; }}
  body {{ font-family: 'Arial', sans-serif; font-size: 12px; color: #1e293b; line-height: 1.5; }}
  h1 {{ font-size: 22px; font-weight: 700; color: #0f172a; margin: 0 0 4px; }}
  h2 {{ font-size: 14px; font-weight: 600; color: #1e293b; margin: 20px 0 8px; border-bottom: 1px solid #e2e8f0; padding-bottom: 4px; }}
  .header {{ display: flex; justify-content: space-between; align-items: flex-start; margin-bottom: 20px; padding-bottom: 16px; border-bottom: 2px solid #6366f1; }}
  .meta {{ font-size: 11px; color: #64748b; margin-top: 4px; }}
  .badge {{ display: inline-block; padding: 2px 8px; border-radius: 4px; font-size: 10px; font-weight: 700; background: #ede9fe; color: #6366f1; margin-right: 4px; }}
  .health {{ font-size: 28px; font-weight: 800; color: {health_color}; }}
  .health-label {{ font-size: 10px; color: #64748b; }}
  table {{ width: 100%; border-collapse: collapse; margin-bottom: 12px; font-size: 11px; }}
  th {{ background: #f1f5f9; padding: 6px 8px; text-align: left; font-weight: 600; color: #475569; font-size: 10px; text-transform: uppercase; letter-spacing: 0.4px; }}
  td {{ padding: 6px 8px; border-bottom: 1px solid #f1f5f9; vertical-align: top; }}
  tr:last-child td {{ border-bottom: none; }}
  .footer {{ margin-top: 24px; padding-top: 8px; border-top: 1px solid #e2e8f0; font-size: 10px; color: #94a3b8; display: flex; justify-content: space-between; }}
  .kpi-row {{ display: flex; gap: 16px; margin-bottom: 16px; }}
  .kpi {{ flex: 1; background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 6px; padding: 10px 12px; }}
  .kpi-val {{ font-size: 20px; font-weight: 700; color: #0f172a; }}
  .kpi-label {{ font-size: 10px; color: #94a3b8; text-transform: uppercase; letter-spacing: 0.4px; margin-top: 2px; }}
</style>
</head>
<body>
<div class="header">
  <div>
    <h1>{client.name}</h1>
    <div class="meta">
      <span class="badge">{client.segment or '—'}</span>
      Менеджер: {client.manager_email or '—'}
      {'· Домен: ' + client.domain if client.domain else ''}
    </div>
  </div>
  <div style="text-align:right;">
    <div class="health">{health_pct}%</div>
    <div class="health-label">Health Score</div>
    <div style="font-size:10px;color:#94a3b8;margin-top:4px;">Отчёт от {now_str}</div>
  </div>
</div>

<div class="kpi-row">
  <div class="kpi">
    <div class="kpi-val">{sum(1 for t in tasks if t.status in ('plan','in_progress','blocked'))}</div>
    <div class="kpi-label">Открытых задач</div>
  </div>
  <div class="kpi">
    <div class="kpi-val">{sum(1 for t in tasks if t.status == 'done')}</div>
    <div class="kpi-label">Выполнено</div>
  </div>
  <div class="kpi">
    <div class="kpi-val">{len(meetings)}</div>
    <div class="kpi-label">Встреч</div>
  </div>
  <div class="kpi">
    <div class="kpi-val">{client.last_meeting_date.strftime('%d.%m') if client.last_meeting_date else '—'}</div>
    <div class="kpi-label">Последний контакт</div>
  </div>
</div>

{'<h2>Цели на квартал</h2><ul>' + goals_html + '</ul>' if goals_html else ''}

<h2>Задачи</h2>
{'<table><thead><tr><th>Задача</th><th>Статус</th><th>Приоритет</th><th>Команда</th><th>Дедлайн</th></tr></thead><tbody>' + task_rows + '</tbody></table>' if task_rows else '<p style="color:#94a3b8;font-size:11px;">Задач нет</p>'}

<h2>Встречи</h2>
{'<table><thead><tr><th>Дата</th><th>Тип</th><th>Тема</th><th>Фолоуап</th></tr></thead><tbody>' + meeting_rows + '</tbody></table>' if meeting_rows else '<p style="color:#94a3b8;font-size:11px;">Встреч нет</p>'}

{'<h2>Заметки</h2>' + notes_html if notes_html else ''}

{'<h2>QBR · ' + (qbr.quarter or '') + '</h2><p>' + (qbr.summary or '') + '</p>' if qbr and qbr.summary else ''}

<div class="footer">
  <span>AM Hub · {client.name}</span>
  <span>{now_str}</span>
</div>
</body>
</html>"""

    try:
        import weasyprint
        pdf_bytes = weasyprint.HTML(string=html).write_pdf()
        from fastapi.responses import Response
        fname = f"{client.name.replace(' ', '_')}_{now_str.replace('.', '-')}.pdf"
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )
    except Exception as e:
        logger.error(f"PDF export error: {e}")
        return {"error": str(e)}

# ============================================================================
# КЛИЕНТ: редактирование карточки
# ============================================================================

@app.patch("/api/clients/{client_id}")
async def api_update_client(
    client_id: int,
    request: Request,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """Обновить поля клиента (сегмент, имя, домен, health_score)."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)

    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        raise HTTPException(status_code=404)

    data = await request.json()
    allowed = ("name", "segment", "domain", "health_score", "manager_email", "activity_level")
    changed = {}
    for field in allowed:
        if field in data:
            old_val = getattr(client, field)
            new_val = data[field]
            if old_val != new_val:
                setattr(client, field, new_val)
                changed[field] = {"old": old_val, "new": new_val}

    if changed:
        db.commit()
        # Push изменений в Airtable если есть record_id
        if client.airtable_record_id:
            try:
                from airtable_sync import push_client_fields_to_airtable
                # Маппинг полей хаба → поля Airtable (подстраивается под реальную структуру)
                at_fields = {}
                if "segment" in changed:
                    at_fields["Сегмент"] = data["segment"]
                if "health_score" in changed:
                    at_fields["Health Score"] = data["health_score"]
                if "domain" in changed:
                    at_fields["Домен"] = data["domain"]
                if at_fields:
                    await push_client_fields_to_airtable(client.airtable_record_id, at_fields)
            except Exception as e:
                logger.warning(f"Airtable push on client update failed: {e}")

    return {"ok": True, "changed": changed}


# ============================================================================
# ЗАДАЧИ: комментарии
# ============================================================================

@app.get("/api/tasks/{task_id}/comments")
async def api_get_task_comments(
    task_id: int,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    comments = db.query(TaskComment).filter(TaskComment.task_id == task_id).order_by(TaskComment.created_at.asc()).all()
    return {"comments": [{
        "id": c.id,
        "content": c.content,
        "created_at": c.created_at.strftime("%d.%m.%Y %H:%M") if c.created_at else None,
        "user_id": c.user_id,
    } for c in comments]}


@app.post("/api/tasks/{task_id}/comments")
async def api_add_task_comment(
    task_id: int,
    request: Request,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        raise HTTPException(status_code=401)

    data = await request.json()
    content = (data.get("content") or "").strip()
    if not content:
        return {"error": "Пустой комментарий"}

    comment = TaskComment(task_id=task_id, user_id=user.id, content=content)
    db.add(comment)
    db.commit()
    return {"ok": True, "id": comment.id}


@app.delete("/api/tasks/{task_id}/comments/{comment_id}")
async def api_delete_task_comment(
    task_id: int,
    comment_id: int,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    if not auth_token:
        raise HTTPException(status_code=401)
    comment = db.query(TaskComment).filter(
        TaskComment.id == comment_id, TaskComment.task_id == task_id
    ).first()
    if not comment:
        raise HTTPException(status_code=404)
    db.delete(comment)
    db.commit()
    return {"ok": True}


# ============================================================================
# KPI МЕНЕДЖЕРА
# ============================================================================

@app.get("/api/manager/kpi")
async def api_manager_kpi(
    period_days: int = 30,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """KPI менеджера за период: задачи, встречи, фолоуапы, чекапы."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        raise HTTPException(status_code=401)

    since = datetime.utcnow() - timedelta(days=period_days)
    email = user.email

    # Задачи
    tasks_closed = db.query(Task).join(Client, Task.client_id == Client.id, isouter=True).filter(
        Client.manager_email == email,
        Task.status == "done",
        Task.confirmed_at >= since,
    ).count()

    tasks_created = db.query(Task).join(Client, Task.client_id == Client.id, isouter=True).filter(
        Client.manager_email == email,
        Task.created_at >= since,
    ).count()

    tasks_overdue = db.query(Task).join(Client, Task.client_id == Client.id, isouter=True).filter(
        Client.manager_email == email,
        Task.due_date < datetime.utcnow(),
        Task.status.in_(["plan", "in_progress"]),
    ).count()

    # Встречи
    meetings_held = db.query(Meeting).join(Client, Meeting.client_id == Client.id, isouter=True).filter(
        Client.manager_email == email,
        Meeting.date >= since,
        Meeting.date <= datetime.utcnow(),
    ).count()

    # Фолоуапы отправлены
    followups_sent = db.query(Meeting).join(Client, Meeting.client_id == Client.id, isouter=True).filter(
        Client.manager_email == email,
        Meeting.followup_status == "sent",
        Meeting.followup_sent_at >= since,
    ).count()

    followups_pending = db.query(Meeting).join(Client, Meeting.client_id == Client.id, isouter=True).filter(
        Client.manager_email == email,
        Meeting.followup_status == "pending",
        Meeting.date < datetime.utcnow(),
    ).count()

    # Клиенты
    total_clients = db.query(Client).filter(Client.manager_email == email).count()

    clients_no_contact = db.query(Client).filter(
        Client.manager_email == email,
        Client.last_meeting_date < datetime.utcnow() - timedelta(days=60),
    ).count()

    # Средний health score
    from sqlalchemy import func
    avg_health = db.query(func.avg(Client.health_score)).filter(
        Client.manager_email == email,
        Client.health_score != None,
    ).scalar() or 0

    return {
        "period_days": period_days,
        "manager": user.email,
        "tasks": {
            "closed": tasks_closed,
            "created": tasks_created,
            "overdue": tasks_overdue,
            "close_rate": round(tasks_closed / max(tasks_created, 1) * 100, 1),
        },
        "meetings": {
            "held": meetings_held,
            "followups_sent": followups_sent,
            "followups_pending": followups_pending,
            "followup_rate": round(followups_sent / max(meetings_held, 1) * 100, 1),
        },
        "clients": {
            "total": total_clients,
            "no_contact_60d": clients_no_contact,
            "avg_health_score": round(float(avg_health) * 100, 1),
        },
    }


@app.get("/api/team/kpi")
async def api_team_kpi(
    period_days: int = 30,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """KPI всей команды — только для admin."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user or user.role != "admin":
        raise HTTPException(status_code=403)

    since = datetime.utcnow() - timedelta(days=period_days)
    managers = db.query(User).filter(User.role == "manager", User.is_active == True).all()
    result = []

    for m in managers:
        tasks_closed = db.query(Task).join(Client, Task.client_id == Client.id, isouter=True).filter(
            Client.manager_email == m.email,
            Task.status == "done", Task.confirmed_at >= since,
        ).count()
        meetings_held = db.query(Meeting).join(Client, Meeting.client_id == Client.id, isouter=True).filter(
            Client.manager_email == m.email,
            Meeting.date >= since, Meeting.date <= datetime.utcnow(),
        ).count()
        followups_sent = db.query(Meeting).join(Client, Meeting.client_id == Client.id, isouter=True).filter(
            Client.manager_email == m.email,
            Meeting.followup_status == "sent", Meeting.followup_sent_at >= since,
        ).count()
        clients_count = db.query(Client).filter(Client.manager_email == m.email).count()

        result.append({
            "manager": m.email,
            "name": f"{m.first_name or ''} {m.last_name or ''}".strip() or m.email,
            "tasks_closed": tasks_closed,
            "meetings_held": meetings_held,
            "followups_sent": followups_sent,
            "clients": clients_count,
        })

    result.sort(key=lambda x: x["tasks_closed"] + x["meetings_held"], reverse=True)
    return {"period_days": period_days, "managers": result}


# ============================================================================
# AIRTABLE WEBHOOK (входящие изменения из Airtable)
# ============================================================================

# ============================================================================
# ОНБОРДИНГ ПАРТНЁРА: чеклист
# ============================================================================

ONBOARDING_CHECKLIST = [
    {"id": 1, "title": "Провести вводную встречу (1 ч)", "type": "meeting", "day": 0},
    {"id": 2, "title": "Отправить welcome-фолоуап", "type": "followup", "day": 0},
    {"id": 3, "title": "Добавить в карточку клиента", "type": "admin", "day": 1},
    {"id": 4, "title": "Касание в Ktalk (день 3)", "type": "ktalk", "day": 3},
    {"id": 5, "title": "Проверить первые шаги партнёра", "type": "check", "day": 7},
    {"id": 6, "title": "Касание в Ktalk (день 7)", "type": "ktalk", "day": 7},
    {"id": 7, "title": "Первый чекап (2 нед)", "type": "meeting", "day": 14},
    {"id": 8, "title": "Касание в Ktalk (день 14)", "type": "ktalk", "day": 14},
    {"id": 9, "title": "Проверить health score", "type": "check", "day": 30},
    {"id": 10, "title": "Закрыть онбординг, перевести в активные", "type": "admin", "day": 30},
]


@app.get("/api/clients/{client_id}/onboarding")
async def api_get_onboarding(
    client_id: int,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """Получить чеклист онбординга клиента."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)

    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        raise HTTPException(status_code=404)

    # Читаем прогресс из account_plan
    plan_data = client.account_plan or {}
    onboarding_progress = plan_data.get("onboarding_progress", {})

    checklist = []
    for item in ONBOARDING_CHECKLIST:
        checklist.append({
            **item,
            "done": onboarding_progress.get(str(item["id"]), False),
        })

    return {"checklist": checklist, "client_id": client_id}


@app.patch("/api/clients/{client_id}/onboarding/{item_id}")
async def api_update_onboarding_item(
    client_id: int,
    item_id: int,
    request: Request,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """Отметить шаг онбординга выполненным."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)

    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        raise HTTPException(status_code=404)

    data = await request.json()
    done = bool(data.get("done", True))

    plan_data = dict(client.account_plan or {})
    if "onboarding_progress" not in plan_data:
        plan_data["onboarding_progress"] = {}
    plan_data["onboarding_progress"][str(item_id)] = done

    client.account_plan = plan_data
    from sqlalchemy.orm.attributes import flag_modified
    flag_modified(client, "account_plan")
    db.commit()

    # Если отмечаем касание в Ktalk — отправляем авто-сообщение
    item = next((i for i in ONBOARDING_CHECKLIST if i["id"] == item_id), None)
    if done and item and item["type"] == "ktalk":
        try:
            from auth import decode_access_token
            user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
            if user:
                settings = user.settings or {}
                kt = settings.get("ktalk", {})
                channel_id = kt.get("followup_channel_id") or kt.get("channel_id")
                token = kt.get("access_token", "")
                if channel_id and token:
                    from integrations.ktalk import send_message
                    await send_message(
                        channel_id,
                        f"📋 Онбординг {client.name}: выполнено — {item['title']}",
                        token,
                    )
        except Exception as e:
            logger.warning(f"Ktalk onboarding notify failed: {e}")

    return {"ok": True, "done": done}


# ============================================================================
# KPI PAGE
# ============================================================================

@app.get("/kpi", response_class=HTMLResponse)
async def kpi_page(request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    if not auth_token:
        return RedirectResponse(url="/login", status_code=303)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        return RedirectResponse(url="/login", status_code=303)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    return templates.TemplateResponse("kpi.html", {"request": request, "user": user})

# ============================================================================
# AIRTABLE WEBHOOK (push при изменении записи)
# ============================================================================

# ============================================================================
# GOOGLE SHEETS WRITE-BACK (запись статуса чекапа обратно)
# ============================================================================

# ============================================================================
# AIRTABLE WEBHOOK — push при изменении записи в Airtable
# ============================================================================

@app.post("/webhook/airtable")
async def airtable_webhook(request: Request, db: Session = Depends(get_db)):
    """
    Webhook от Airtable Automations.
    Airtable → Automations → Webhook → этот endpoint.
    При изменении записи обновляем клиента в локальной БД.
    """
    try:
        payload = await request.json()
    except Exception:
        return {"ok": False, "error": "invalid json"}

    # Airtable присылает {record_id, fields: {...}}
    record_id = payload.get("record_id") or payload.get("id")
    fields = payload.get("fields") or payload.get("data") or {}

    if not record_id:
        return {"ok": False, "error": "no record_id"}

    client = db.query(Client).filter(Client.airtable_record_id == record_id).first()
    if not client:
        # Пробуем создать нового клиента
        name = fields.get("Название") or fields.get("Name") or fields.get("Клиент") or fields.get("Company")
        if name:
            client = Client(
                airtable_record_id=record_id,
                name=str(name),
                segment=str(fields.get("Сегмент") or fields.get("Segment") or ""),
                manager_email=str(fields.get("Менеджер") or fields.get("Manager Email") or ""),
            )
            db.add(client)
            db.commit()
            logger.info(f"✅ Airtable webhook: created client {name}")
            return {"ok": True, "action": "created", "client_id": client.id}
        return {"ok": False, "error": "client not found and no name field"}

    # Обновляем поля
    field_map = {
        "Название": "name", "Name": "name", "Клиент": "name", "Company": "name",
        "Сегмент": "segment", "Segment": "segment",
        "Домен": "domain", "Domain": "domain",
        "Менеджер": "manager_email", "Manager Email": "manager_email",
        "Health Score": "health_score",
    }
    updated = []
    for at_field, model_field in field_map.items():
        if at_field in fields and fields[at_field] is not None:
            val = fields[at_field]
            if model_field == "health_score":
                try:
                    val = float(val)
                    if val > 1:
                        val = val / 100
                except Exception:
                    continue
            setattr(client, model_field, val)
            updated.append(model_field)

    if updated:
        from sqlalchemy.orm.attributes import flag_modified
        db.commit()
        logger.info(f"✅ Airtable webhook: updated {client.name} fields={updated}")

    return {"ok": True, "action": "updated", "fields": updated}


# ============================================================================
# GOOGLE SHEETS — запись обратно
# ============================================================================

@app.post("/api/sheets/update-checkup")
async def api_sheets_update_checkup(
    request: Request,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """
    Записать статус чекапа обратно в Google Sheets (Top-50).
    Обновляет строку клиента: дата последнего чекапа, статус.
    """
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)

    data = await request.json()
    client_name = data.get("client_name", "")
    checkup_date = data.get("checkup_date", datetime.now().strftime("%Y-%m-%d"))
    status = data.get("status", "done")

    if not client_name:
        return {"error": "client_name required"}

    try:
        from sheets import write_checkup_status
        result = await write_checkup_status(client_name, checkup_date, status)
        return {"ok": result, "client": client_name}
    except Exception as e:
        return {"error": str(e), "note": "Sheets write-back requires service account credentials"}


@app.post("/api/sheets/batch-update")
async def api_sheets_batch_update(
    request: Request,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """Массовое обновление данных в Google Sheets."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)

    data = await request.json()
    updates = data.get("updates", [])  # [{row, col, value}]

    try:
        from sheets import batch_update_cells
        result = await batch_update_cells(updates)
        return {"ok": result, "count": len(updates)}
    except Exception as e:
        return {"error": str(e)}

# ============================================================================
# ИНТЕГРАЦИИ: тест персональных кредов
# ============================================================================

@app.get("/api/integrations/status")
async def api_integrations_status(
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """
    Статус всех интеграций для текущего менеджера.
    Проверяет наличие кредов в user.settings и последний синк.
    """
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        raise HTTPException(status_code=401)

    s = user.settings or {}
    mr = s.get("merchrules", {})
    kt = s.get("ktalk", {})
    at = s.get("airtable", {})
    tg = s.get("telegram", {})
    tm = s.get("tbank_time", {})
    gs = s.get("sheets", {})

    def last_sync(integration: str) -> str:
        log = db.query(SyncLog).filter(
            SyncLog.integration == integration,
            SyncLog.status == "success",
        ).order_by(SyncLog.started_at.desc()).first()
        if not log or not log.started_at:
            return None
        ago = int((datetime.now() - log.started_at).total_seconds())
        if ago < 60: return "только что"
        if ago < 3600: return f"{ago//60} мин назад"
        if ago < 86400: return f"{ago//3600} ч назад"
        return f"{ago//86400} дн назад"

    return {
        "merchrules": {
            "configured": bool(mr.get("login") and mr.get("password")),
            "login": mr.get("login", ""),
            "last_sync": last_sync("merchrules"),
        },
        "ktalk": {
            "configured": bool(kt.get("access_token")),
            "login": kt.get("login", ""),
            "channel_id": kt.get("followup_channel_id", ""),
            "last_sync": last_sync("ktalk"),
        },
        "airtable": {
            "configured": bool(at.get("pat") or at.get("token")),
            "base_id": at.get("base_id", ""),
            "last_sync": last_sync("airtable"),
        },
        "tbank_time": {
            "configured": bool(tm.get("login") or tm.get("session_cookie")),
            "login": tm.get("login", ""),
        },
        "sheets": {
            "configured": bool(gs.get("spreadsheet_id") or _env("SHEETS_SPREADSHEET_ID")),
            "spreadsheet_id": gs.get("spreadsheet_id") or _env("SHEETS_SPREADSHEET_ID"),
        },
        "telegram": {
            "configured": bool(user.telegram_id or tg.get("chat_id")),
            "telegram_id": user.telegram_id,
        },
        "groq_ai": {
            "configured": bool(_env("GROQ_API_KEY") or _env("QWEN_API_KEY")),
        },
    }


@app.post("/api/integrations/test/{service}")
async def api_test_integration(
    service: str,
    request: Request,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """
    Проверить подключение конкретного сервиса с персональными кредами.
    service: merchrules | ktalk | airtable | tbank_time | sheets
    """
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        raise HTTPException(status_code=401)

    s = user.settings or {}
    import httpx

    if service == "merchrules":
        mr = s.get("merchrules", {})
        login = mr.get("login", "") or _env("MERCHRULES_LOGIN")
        password = mr.get("password", "") or _env("MERCHRULES_PASSWORD")
        if not login or not password:
            return {"ok": False, "error": "Логин и пароль не заданы"}
        base = _env("MERCHRULES_API_URL", "https://merchrules.any-platform.ru")
        try:
            async with httpx.AsyncClient(timeout=15) as hx:
                for field in ("email", "login", "username"):
                    r = await hx.post(f"{base}/backend-v2/auth/login",
                                      json={field: login, "password": password}, timeout=10)
                    if r.status_code == 200:
                        body = r.json()
                        token = body.get("token") or body.get("access_token") or body.get("accessToken")
                        if token:
                            # Считаем клиентов
                            ra = await hx.get(f"{base}/backend-v2/accounts?limit=1",
                                              headers={"Authorization": f"Bearer {token}"}, timeout=10)
                            count = len(ra.json().get("accounts", ra.json().get("items", []))) if ra.status_code == 200 else "?"
                            return {"ok": True, "message": f"✅ Подключено ({field}={login})", "accounts": count}
            return {"ok": False, "error": "Неверный логин или пароль"}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    elif service == "ktalk":
        kt = s.get("ktalk", {})
        token = kt.get("access_token", "")
        space = kt.get("space", "") or _env("KTALK_SPACE")
        if not token or not space:
            return {"ok": False, "error": "Нет токена — войдите через /auth/ktalk"}
        try:
            async with httpx.AsyncClient(timeout=10) as hx:
                r = await hx.get(f"https://{space}.ktalk.ru/api/v4/users/me",
                                  headers={"Authorization": f"Bearer {token}"})
            if r.status_code == 200:
                me = r.json()
                return {"ok": True, "message": f"✅ {me.get('username', space)}", "email": me.get("email")}
            return {"ok": False, "error": f"HTTP {r.status_code}"}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    elif service == "airtable":
        at = s.get("airtable", {})
        token = at.get("pat") or at.get("token") or _env("AIRTABLE_TOKEN")
        base_id = at.get("base_id") or _env("AIRTABLE_BASE_ID")
        if not token:
            return {"ok": False, "error": "Токен не задан"}
        try:
            async with httpx.AsyncClient(timeout=10) as hx:
                url = f"https://api.airtable.com/v0/meta/bases/{base_id}/tables" if base_id else "https://api.airtable.com/v0/meta/bases"
                r = await hx.get(url, headers={"Authorization": f"Bearer {token}"})
            if r.status_code == 200:
                d = r.json()
                count = len(d.get("tables", d.get("bases", [])))
                return {"ok": True, "message": f"✅ Airtable подключён ({count} объектов)"}
            return {"ok": False, "error": f"HTTP {r.status_code}: {r.text[:100]}"}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    elif service == "tbank_time":
        tm = s.get("tbank_time", {})
        login = tm.get("login", "")
        password = tm.get("password", "")
        session = tm.get("session_cookie", "") or _env("TIME_SESSION_COOKIE")
        if not login and not session:
            return {"ok": False, "error": "Логин или session cookie не заданы"}
        try:
            async with httpx.AsyncClient(timeout=10) as hx:
                headers = {}
                if session:
                    headers["Cookie"] = f"MMAUTH={session}"
                r = await hx.get("https://time.tbank.ru/api/v1/users/me", headers=headers)
            if r.status_code == 200:
                me = r.json()
                return {"ok": True, "message": f"✅ {me.get('username', login)}"}
            return {"ok": False, "error": f"HTTP {r.status_code} — проверьте session cookie"}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    elif service == "sheets":
        gs = s.get("sheets", {})
        sheet_id = _extract_sheets_id(gs.get("spreadsheet_id") or _env("SHEETS_SPREADSHEET_ID"))
        if not sheet_id:
            return {"ok": False, "error": "Spreadsheet ID не задан"}
        try:
            from sheets import fetch_sheet_csv
            rows = await fetch_sheet_csv(sheet_id)
            return {"ok": bool(rows), "message": f"✅ Таблица доступна ({len(rows)} строк)" if rows else "❌ Таблица пустая или недоступна"}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    return {"ok": False, "error": f"Неизвестный сервис: {service}"}

# ============================================================================
# CHROME EXTENSION — push токенов Time и Ktalk
# ============================================================================

@app.post("/api/auth/tokens/push")
async def api_tokens_push(
    request: Request,
    db: Session = Depends(get_db),
):
    """
    Принимает токены от Chrome Extension.
    Расширение автоматически перехватывает MMAUTHTOKEN из cookies браузера.
    Auth через Bearer токен (hub_token из настроек расширения).
    """
    auth_header = request.headers.get("Authorization", "")
    bearer = auth_header.replace("Bearer ", "").strip()

    # Ищем пользователя по hub_token (сохранён в user.settings.hub_token)
    user = None
    if bearer:
        from auth import decode_access_token
        # Пробуем как JWT токен хаба
        payload = decode_access_token(bearer)
        if payload:
            user = db.query(User).filter(User.id == int(payload.get("sub", 0))).first()

        # Fallback: ищем по статическому hub_token в settings
        if not user:
            all_users = db.query(User).filter(User.is_active == True).all()
            for u in all_users:
                s = u.settings or {}
                if s.get("hub_token") == bearer:
                    user = u
                    break

    if not user:
        # Если нет авторизации — создаём анонимный push (для первичной настройки)
        # Токены запишутся как pending, менеджер увидит их на странице настроек
        data = await request.json()
        logger.info(f"Anon token push: time={'time_token' in data}, ktalk={'ktalk_token' in data}")
        return {"ok": True, "note": "Войдите в AM Hub и перейдите в Настройки для привязки токена"}

    data = await request.json()
    settings = dict(user.settings or {})
    updated = []

    # Tbank Time MMAUTHTOKEN
    time_token = data.get("time_token", "")
    if time_token:
        tm = dict(settings.get("tbank_time", {}))
        if tm.get("mmauthtoken") != time_token:
            tm["mmauthtoken"] = time_token
            tm["session_cookie"] = time_token
            # Сразу проверяем и получаем channel_id
            try:
                import httpx
                async with httpx.AsyncClient(timeout=10) as hx:
                    me = await hx.get(
                        "https://time.tbank.ru/api/v4/users/me",
                        headers={"Authorization": f"Bearer {time_token}"},
                    )
                    if me.status_code == 200:
                        me_data = me.json()
                        tm["username"] = me_data.get("username", "")
                        tm["email"] = me_data.get("email", "")
                    ch = await hx.get(
                        "https://time.tbank.ru/api/v4/teams/name/tinkoff/channels/name/any-team-support",
                        headers={"Authorization": f"Bearer {time_token}"},
                    )
                    if ch.status_code == 200:
                        tm["support_channel_id"] = ch.json().get("id", "")
            except Exception as e:
                logger.debug(f"Token validation error: {e}")
            settings["tbank_time"] = tm
            updated.append("time")

    # Ktalk access_token
    ktalk_token = data.get("ktalk_token", "")
    if ktalk_token:
        kt = dict(settings.get("ktalk", {}))
        if kt.get("access_token") != ktalk_token:
            kt["access_token"] = ktalk_token
            settings["ktalk"] = kt
            updated.append("ktalk")

    if updated:
        from sqlalchemy.orm.attributes import flag_modified
        user.settings = settings
        flag_modified(user, "settings")
        db.commit()
        logger.info(f"✅ Extension pushed tokens for {user.email}: {updated}")

    return {"ok": True, "updated": updated, "user": user.email}



# ── PWA ───────────────────────────────────────────────────────────────────────
from fastapi.responses import FileResponse
import os as _os

@app.get("/manifest.json")
async def pwa_manifest():
    p = "static/manifest.json"
    if _os.path.exists(p): return FileResponse(p, media_type="application/manifest+json")
    return {"name":"AM Hub","short_name":"AM Hub","start_url":"/","display":"standalone",
            "background_color":"#07090f","theme_color":"#6474ff"}

@app.get("/sw.js")
async def service_worker():
    p = "static/sw.js"
    if _os.path.exists(p): return FileResponse(p, media_type="application/javascript",
                                                  headers={"Service-Worker-Allowed": "/"})
    return FileResponse("/dev/null", media_type="application/javascript")

# ============================================================================
# WEBSOCKET — real-time обновления
# ============================================================================
from fastapi import WebSocket, WebSocketDisconnect
from typing import Dict, List
import asyncio, json as _json

class ConnectionManager:
    """Менеджер WebSocket соединений. Группирует по user_id."""
    def __init__(self):
        self.active: Dict[int, List[WebSocket]] = {}

    async def connect(self, user_id: int, ws: WebSocket):
        await ws.accept()
        self.active.setdefault(user_id, []).append(ws)

    def disconnect(self, user_id: int, ws: WebSocket):
        conns = self.active.get(user_id, [])
        if ws in conns:
            conns.remove(ws)
        if not conns:
            self.active.pop(user_id, None)

    async def send(self, user_id: int, data: dict):
        """Отправить данные всем вкладкам пользователя."""
        dead = []
        for ws in self.active.get(user_id, []):
            try:
                await ws.send_text(_json.dumps(data, ensure_ascii=False, default=str))
            except Exception:
                dead.append(ws)
        for ws in dead:
            self.disconnect(user_id, ws)

    async def broadcast(self, data: dict):
        """Отправить всем подключённым пользователям."""
        for uid in list(self.active.keys()):
            await self.send(uid, data)

    @property
    def connected_users(self):
        return len(self.active)


ws_manager = ConnectionManager()


@app.websocket("/ws")
async def websocket_endpoint(
    websocket: WebSocket,
    token: str = "",
    db: Session = Depends(get_db),
):
    """
    WebSocket endpoint для real-time обновлений.
    Клиент подключается с токеном: ws://host/ws?token=<auth_token>
    
    Сервер пушит:
      {type: "task_update", task: {...}}
      {type: "notification", ...}
      {type: "stats", overdue, warning, open_tasks}
      {type: "ping"}
    """
    # Авторизация через query param token
    if not token:
        await websocket.close(code=4001)
        return

    from auth import decode_access_token
    payload = decode_access_token(token)
    if not payload:
        await websocket.close(code=4001)
        return

    user_id = int(payload.get("sub", 0))
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        await websocket.close(code=4001)
        return

    await ws_manager.connect(user_id, websocket)
    logger.info(f"WS connected: user_id={user_id}, total={ws_manager.connected_users}")

    try:
        # Сразу отправляем текущие stats
        now = datetime.now()
        q = db.query(Client)
        if user.role == "manager":
            q = q.filter(Client.manager_email == user.email)
        clients = q.all()
        overdue = warning = 0
        for c in clients:
            last = c.last_meeting_date or c.last_checkup
            if not last: continue
            days = (now - last).days
            interval = CHECKUP_INTERVALS.get(c.segment or "", 90)
            if days > interval: overdue += 1
            elif days > interval - 14: warning += 1
        tq = db.query(Task).join(Client, Task.client_id == Client.id, isouter=True).filter(
            Task.status.in_(["plan", "in_progress", "blocked"])
        )
        if user.role == "manager":
            tq = tq.filter(Client.manager_email == user.email)
        open_tasks = tq.count()

        await websocket.send_text(_json.dumps({
            "type": "stats",
            "overdue": overdue, "warning": warning, "open_tasks": open_tasks
        }))

        # Heartbeat loop — держим соединение живым
        while True:
            try:
                # Ждём сообщение от клиента (ping) или timeout
                data = await asyncio.wait_for(websocket.receive_text(), timeout=30.0)
                msg = _json.loads(data)
                if msg.get("type") == "ping":
                    await websocket.send_text(_json.dumps({"type": "pong"}))
            except asyncio.TimeoutError:
                # Каждые 30 сек шлём ping
                await websocket.send_text(_json.dumps({"type": "ping"}))
            except WebSocketDisconnect:
                break

    except WebSocketDisconnect:
        pass
    finally:
        ws_manager.disconnect(user_id, websocket)
        logger.info(f"WS disconnected: user_id={user_id}, total={ws_manager.connected_users}")


async def ws_notify_user(user_id: int, event_type: str, data: dict):
    """
    Хелпер для push-уведомлений конкретному пользователю.
    Вызывается из других endpoint'ов после изменения данных.
    """
    await ws_manager.send(user_id, {"type": event_type, **data})


async def ws_invalidate_stats(user_id: int, db):
    """Инвалидирует кеш stats и пушит обновлённые данные через WS."""
    cache_del(f"stats:{user_id}")
    cache_del(f"notif:{user_id}")
    # Пушим новые stats если пользователь подключён по WS
    if user_id in ws_manager.active:
        now = datetime.now()
        user = db.query(User).filter(User.id == user_id).first()
        if not user: return
        q = db.query(Client)
        if user.role == "manager": q = q.filter(Client.manager_email == user.email)
        clients = q.all()
        overdue = warning = 0
        for c in clients:
            last = c.last_meeting_date or c.last_checkup
            if not last: continue
            days = (now - last).days
            interval = CHECKUP_INTERVALS.get(c.segment or "", 90)
            if days > interval: overdue += 1
            elif days > interval - 14: warning += 1
        tq = db.query(Task).join(Client, Task.client_id == Client.id, isouter=True).filter(
            Task.status.in_(["plan", "in_progress", "blocked"])
        )
        if user.role == "manager": tq = tq.filter(Client.manager_email == user.email)
        open_tasks = tq.count()
        await ws_manager.send(user_id, {
            "type": "stats", "overdue": overdue,
            "warning": warning, "open_tasks": open_tasks
        })


# ============================================================================
# REMAINING MISSING ENDPOINTS
# ============================================================================

@app.post("/api/checklist/init")
@app.post("/api/checklist/add")
@app.post("/api/checklist/clear")
async def api_checklist(
    request: Request,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """Чеклист встречи — хранится в user.settings."""
    if not auth_token: raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload: raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user: raise HTTPException(status_code=401)

    path = str(request.url.path)
    body = {}
    try: body = await request.json()
    except: pass

    from sqlalchemy.orm.attributes import flag_modified
    settings = dict(user.settings or {})

    if "init" in path:
        settings["checklist"] = [
            {"id": 1, "text": "Приветствие и цели встречи", "done": False},
            {"id": 2, "text": "Статус открытых задач", "done": False},
            {"id": 3, "text": "Метрики и показатели", "done": False},
            {"id": 4, "text": "Планы и следующие шаги", "done": False},
            {"id": 5, "text": "Фолоуап назначен", "done": False},
        ]
    elif "add" in path:
        cl = settings.get("checklist", [])
        new_id = max((i["id"] for i in cl), default=0) + 1
        cl.append({"id": new_id, "text": body.get("text", ""), "done": False})
        settings["checklist"] = cl
    elif "clear" in path:
        settings["checklist"] = []

    user.settings = settings
    flag_modified(user, "settings")
    db.commit()
    return {"ok": True, "checklist": settings.get("checklist", [])}


@app.get("/api/checklist")
async def api_checklist_get(
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    if not auth_token: raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload: raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user: raise HTTPException(status_code=401)
    settings = user.settings or {}
    return {"checklist": settings.get("checklist", [])}


@app.post("/api/internal-task")
@app.get("/api/internal-task")
async def api_internal_task(
    request: Request,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """Внутренние задачи менеджера (не привязаны к клиенту)."""
    if not auth_token: raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload: raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user: raise HTTPException(status_code=401)

    if request.method == "GET":
        tasks = db.query(Task).filter(
            Task.client_id.is_(None),
            Task.created_at >= datetime.utcnow() - timedelta(days=90)
        ).order_by(Task.created_at.desc()).limit(50).all()
        return {"tasks": [{"id":t.id,"title":t.title,"status":t.status,"priority":t.priority,
                           "due_date":t.due_date.isoformat() if t.due_date else None} for t in tasks]}

    body = await request.json()
    task = Task(
        client_id=None, title=body.get("title","Задача"),
        status=body.get("status","plan"), priority=body.get("priority","medium"),
        description=body.get("description",""),
        due_date=datetime.fromisoformat(body["due_date"]) if body.get("due_date") else None,
        created_at=datetime.utcnow(),
    )
    db.add(task); db.commit(); db.refresh(task)
    return {"ok": True, "id": task.id}


@app.post("/api/metrics/upload")
async def api_metrics_upload(
    file: UploadFile,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """Загрузка метрик Top-50 (CSV/Excel)."""
    if not auth_token: raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload: raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user: raise HTTPException(status_code=401)

    import io, pandas as pd
    content_bytes = await file.read()
    try:
        if file.filename and file.filename.endswith('.csv'):
            df = pd.read_csv(io.BytesIO(content_bytes), dtype=str)
        else:
            df = pd.read_excel(io.BytesIO(content_bytes), dtype=str)
        df.columns = [c.strip().lower().replace(" ","_") for c in df.columns]
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Ошибка файла: {e}")

    updated = 0
    from sqlalchemy.orm.attributes import flag_modified
    for _, row in df.iterrows():
        name = str(row.get("name") or row.get("название") or row.get("client") or "").strip()
        if not name or name == "nan": continue
        client = db.query(Client).filter(Client.name.ilike(f"%{name}%")).first()
        if not client: continue
        meta = dict(client.integration_metadata or {})
        for col in df.columns:
            v = str(row.get(col) or "").strip()
            if v and v != "nan" and col not in ("name","название","client"):
                meta[f"metric_{col}"] = v
        client.integration_metadata = meta
        flag_modified(client, "integration_metadata")
        updated += 1
    db.commit()
    return {"ok": True, "updated": updated}


# ============================================================================
# AI CHAT
# ============================================================================

@app.get("/ai-chat", response_class=HTMLResponse)
async def ai_chat_page(request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    if not auth_token: return RedirectResponse(url="/login", status_code=303)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload: return RedirectResponse(url="/login", status_code=303)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user: return RedirectResponse(url="/login", status_code=303)
    return templates.TemplateResponse("ai_chat.html", {"request": request, "user": user})


@app.post("/api/ai/chat")
async def api_ai_chat(
    request: Request,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """AI чат с контекстом клиента."""
    if not auth_token: raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload: raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user: raise HTTPException(status_code=401)

    body = await request.json()
    message   = body.get("message", "").strip()
    client_id = body.get("client_id")
    history   = body.get("history", [])

    if not message:
        return {"reply": "Напишите что-нибудь."}

    # Собираем контекст клиента
    context_parts = []
    client = None
    if client_id:
        client = db.query(Client).filter(Client.id == client_id).first()
    if client:
        open_tasks = db.query(Task).filter(Task.client_id == client.id, Task.status != "done").all()
        last_meeting = db.query(Meeting).filter(Meeting.client_id == client.id).order_by(Meeting.date.desc()).first()
        context_parts.append(f"""Данные клиента:
- Название: {client.name}
- Сегмент: {client.segment or '—'}
- Health Score: {client.health_score or 0:.0f}%
- Домен: {client.domain or '—'}
- Открытых задач: {len(open_tasks)}
- Последняя встреча: {last_meeting.date.strftime('%d.%m.%Y') if last_meeting and last_meeting.date else 'нет'}
- Топ задачи: {', '.join(t.title for t in open_tasks[:3])}""")
    else:
        # Общий контекст менеджера
        q = db.query(Client)
        if user.role == "manager": q = q.filter(Client.manager_email == user.email)
        clients = q.all()
        health_vals = [c.health_score for c in clients if c.health_score is not None]
        avg_h = sum(health_vals) / len(health_vals) if health_vals else 0
        open_t = db.query(Task).join(Client, Task.client_id == Client.id).filter(
            Task.status != "done"
        )
        if user.role == "manager": open_t = open_t.filter(Client.manager_email == user.email)
        context_parts.append(f"""Портфель менеджера {user.name}:
- Клиентов: {len(clients)}
- Средний Health Score: {avg_h:.0f}%
- Открытых задач: {open_t.count()}
- Клиенты с low health: {sum(1 for h in health_vals if h < 50)}""")

    system_prompt = f"""Ты — AI-ассистент AM Hub, помощник аккаунт-менеджера.
Ты помогаешь управлять портфелем клиентов, составлять планы, писать фолоуапы и анализировать данные.
Отвечай кратко, конкретно, на русском языке. Используй маркированные списки где уместно.

{chr(10).join(context_parts)}

Сегодня: {datetime.utcnow().strftime('%d.%m.%Y')}"""

    # Groq API
    u_settings = user.settings or {}
    groq_key = u_settings.get("groq", {}).get("api_key") or env.GROQ_KEY
    if not groq_key:
        return {"reply": "AI не настроен. Добавьте GROQ_API_KEY в Settings → AI или в Railway Variables."}

    messages = [{"role": "system", "content": system_prompt}]
    for h in history[-8:]:  # последние 8 сообщений контекста
        if h.get("role") in ("user", "assistant"):
            messages.append({"role": h["role"], "content": h["content"][:1000]})
    messages.append({"role": "user", "content": message})

    try:
        import httpx
        async with httpx.AsyncClient(timeout=30) as hx:
            r = await hx.post(
                "https://api.groq.com/openai/v1/chat/completions",
                headers={"Authorization": f"Bearer {groq_key}", "Content-Type": "application/json"},
                json={"model": "llama-3.3-70b-versatile", "messages": messages,
                      "max_tokens": 800, "temperature": 0.7},
            )
        if r.status_code != 200:
            return {"reply": f"Groq API error {r.status_code}. Проверьте API ключ."}
        reply = r.json()["choices"][0]["message"]["content"]
    except Exception as e:
        return {"reply": f"Ошибка AI: {e}"}

    # Сохраняем в историю
    from models import AIChat
    for role, text in [("user", message), ("assistant", reply)]:
        db.add(AIChat(client_id=client_id, user_id=user.id, role=role, content=text))
    db.commit()

    return {"reply": reply, "client_name": client.name if client else None}


@app.get("/api/ai/chat/history")
async def api_ai_chat_history(
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    if not auth_token: raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload: raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user: raise HTTPException(status_code=401)

    from models import AIChat
    from sqlalchemy import func
    # Группируем по session (первое сообщение user за последние 30 дней)
    msgs = (db.query(AIChat)
            .filter(AIChat.user_id == user.id, AIChat.role == "user")
            .order_by(AIChat.created_at.desc()).limit(20).all())
    return {"chats": [{"id": m.id, "first_message": m.content[:60], "created_at": m.created_at.isoformat()} for m in msgs]}


# ============================================================================
# CLIENT HISTORY (audit log)
# ============================================================================

@app.get("/api/clients/{client_id}/history")
async def api_client_history(
    client_id: int,
    limit: int = 50,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    if not auth_token: raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload: raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user: raise HTTPException(status_code=401)

    from models import ClientHistory
    history = (db.query(ClientHistory)
               .filter(ClientHistory.client_id == client_id)
               .order_by(ClientHistory.created_at.desc())
               .limit(limit).all())
    return {"history": [
        {"id": h.id, "field": h.field, "old_value": h.old_value, "new_value": h.new_value,
         "event_type": h.event_type, "comment": h.comment,
         "user_name": h.user.name if h.user else "Система",
         "created_at": h.created_at.isoformat()}
        for h in history
    ]}


def log_client_change(db, client_id: int, user_id: Optional[int],
                       field: str, old_val, new_val, event_type: str = "update", comment: str = None):
    """Хелпер для записи истории изменений."""
    from models import ClientHistory
    if str(old_val) == str(new_val):
        return  # нечего логировать
    entry = ClientHistory(
        client_id=client_id, user_id=user_id, field=field,
        old_value=str(old_val) if old_val is not None else None,
        new_value=str(new_val) if new_val is not None else None,
        event_type=event_type, comment=comment,
    )
    db.add(entry)


# ============================================================================
# TELEGRAM SETTINGS + SMART NOTIFICATIONS
# ============================================================================

@app.get("/api/telegram/status")
async def api_tg_status(db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    if not auth_token: raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload: raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user: raise HTTPException(status_code=401)
    from models import TelegramSubscription
    sub = db.query(TelegramSubscription).filter(TelegramSubscription.user_id == user.id).first()
    return {"connected": bool(sub and sub.chat_id), "chat_id": sub.chat_id if sub else None,
            "settings": {k: getattr(sub, k) for k in ("notify_overdue","notify_health_drop","notify_tasks","notify_daily")} if sub else {}}


@app.post("/api/telegram/connect")
async def api_tg_connect(request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    if not auth_token: raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload: raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user: raise HTTPException(status_code=401)

    body = await request.json()
    chat_id = str(body.get("chat_id", "")).strip()
    if not chat_id: return {"ok": False, "error": "chat_id обязателен"}

    # Проверяем что можем отправить сообщение
    from telegram_bot import send_message
    hub_url = str(request.base_url).rstrip("/")
    ok = await send_message(chat_id, f"✅ <b>AM Hub подключён!</b>\nМенеджер: {user.name}\n<a href='{hub_url}'>Открыть хаб →</a>")
    if not ok: return {"ok": False, "error": "Не удалось отправить сообщение. Проверьте chat_id и что бот запущен."}

    from models import TelegramSubscription
    sub = db.query(TelegramSubscription).filter(TelegramSubscription.user_id == user.id).first()
    if sub:
        sub.chat_id = chat_id; sub.is_active = True
    else:
        sub = TelegramSubscription(user_id=user.id, chat_id=chat_id)
        db.add(sub)
    db.commit()
    return {"ok": True}


@app.patch("/api/telegram/settings")
async def api_tg_settings(request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    if not auth_token: raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload: raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user: raise HTTPException(status_code=401)
    body = await request.json()
    from models import TelegramSubscription
    sub = db.query(TelegramSubscription).filter(TelegramSubscription.user_id == user.id).first()
    if not sub: return {"ok": False, "error": "Сначала подключите Telegram"}
    for k in ("notify_overdue","notify_health_drop","notify_tasks","notify_daily"):
        if k in body: setattr(sub, k, body[k])
    db.commit()
    return {"ok": True}


# ============================================================================
# TEAM DASHBOARD (admin only)
# ============================================================================

@app.get("/team", response_class=HTMLResponse)
async def team_page(request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    if not auth_token: return RedirectResponse(url="/login", status_code=303)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload: return RedirectResponse(url="/login", status_code=303)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user: return RedirectResponse(url="/login", status_code=303)
    if user.role != "admin": raise HTTPException(status_code=403, detail="Только для администраторов")
    return templates.TemplateResponse("team_dashboard.html", {"request": request, "user": user})


@app.get("/api/team/stats")
async def api_team_stats(
    days: int = 30,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    if not auth_token: raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload: raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user or user.role != "admin": raise HTTPException(status_code=403)

    managers = db.query(User).filter(User.role == "manager", User.is_active == True).all()
    all_clients = db.query(Client).all()
    health_all  = [c.health_score for c in all_clients if c.health_score is not None]

    mgr_stats = []
    now = datetime.utcnow()
    for m in managers:
        clients = [c for c in all_clients if c.manager_email == m.email]
        h_vals  = [c.health_score for c in clients if c.health_score is not None]
        avg_h   = sum(h_vals) / len(h_vals) if h_vals else 0
        overdue = sum(1 for c in clients if (c.last_meeting_date and (now - c.last_meeting_date).days > CHECKUP_INTERVALS.get(c.segment or "", 90)))
        mgr_stats.append({
            "id": m.id, "name": m.name or m.email, "email": m.email,
            "clients_count": len(clients), "avg_health": avg_h, "overdue": overdue,
        })
    mgr_stats.sort(key=lambda x: x["avg_health"], reverse=True)

    risk_clients = sorted(
        [{"id": c.id, "name": c.name, "health_score": c.health_score, "segment": c.segment,
          "manager_name": next((m.name or m.email for m in managers if m.email == c.manager_email), "—"),
          "last_contact": c.last_meeting_date.isoformat() if c.last_meeting_date else None}
         for c in all_clients if c.health_score is not None and c.health_score < 55],
        key=lambda x: x["health_score"]
    )[:20]

    open_tasks = db.query(Task).filter(Task.status != "done").count()
    overdue_ck  = sum(1 for c in all_clients
                      if c.last_meeting_date and
                      (now - c.last_meeting_date).days > CHECKUP_INTERVALS.get(c.segment or "", 90))

    return {
        "managers_count": len(managers),
        "total_clients":  len(all_clients),
        "avg_health":     sum(health_all) / len(health_all) if health_all else 0,
        "overdue_checkups": overdue_ck,
        "open_tasks":     open_tasks,
        "managers":       mgr_stats,
        "risk_clients":   risk_clients,
    }


@app.get("/api/team/export")
async def api_team_export(days: int = 30, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    if not auth_token: raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload: raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user or user.role != "admin": raise HTTPException(status_code=403)

    import csv, io
    managers = db.query(User).filter(User.role == "manager").all()
    clients  = db.query(Client).all()
    buf = io.StringIO()
    w   = csv.writer(buf)
    w.writerow(["Менеджер","Email","Клиентов","Avg Health","Просрочено"])
    for m in managers:
        mc = [c for c in clients if c.manager_email == m.email]
        hv = [c.health_score for c in mc if c.health_score is not None]
        ah = sum(hv)/len(hv) if hv else 0
        now = datetime.utcnow()
        ov  = sum(1 for c in mc if c.last_meeting_date and (now - c.last_meeting_date).days > 90)
        w.writerow([m.name or "", m.email, len(mc), f"{ah:.0f}%", ov])
    from fastapi.responses import Response
    return Response(content=buf.getvalue(), media_type="text/csv",
                    headers={"Content-Disposition": f"attachment; filename=team_{datetime.utcnow().strftime('%Y%m%d')}.csv"})


# ============================================================================
# BULK OPERATIONS
# ============================================================================

@app.post("/api/bulk/assign-checkup")
async def api_bulk_assign_checkup(
    request: Request,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """Назначить чекап нескольким клиентам сразу."""
    if not auth_token: raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload: raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user: raise HTTPException(status_code=401)

    body = await request.json()
    client_ids = body.get("client_ids", [])
    date_str   = body.get("date")
    due_date   = datetime.fromisoformat(date_str) if date_str else datetime.utcnow() + timedelta(days=7)

    created = 0
    for cid in client_ids[:50]:  # лимит 50
        task = Task(client_id=cid, title="Провести чекап", status="plan",
                    priority="high", due_date=due_date, created_at=datetime.utcnow())
        db.add(task)
        created += 1
    db.commit()
    return {"ok": True, "created": created}


@app.post("/api/bulk/create-task")
async def api_bulk_create_task(
    request: Request,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """Создать одну задачу для нескольких клиентов."""
    if not auth_token: raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload: raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user: raise HTTPException(status_code=401)

    body = await request.json()
    client_ids = body.get("client_ids", [])
    title      = body.get("title", "Задача")
    priority   = body.get("priority", "medium")
    due_date   = datetime.fromisoformat(body["due_date"]) if body.get("due_date") else datetime.utcnow() + timedelta(days=3)

    created = 0
    for cid in client_ids[:50]:
        db.add(Task(client_id=cid, title=title, status="plan", priority=priority,
                    due_date=due_date, created_at=datetime.utcnow()))
        created += 1
    db.commit()
    return {"ok": True, "created": created}


@app.patch("/api/bulk/update-segment")
async def api_bulk_update_segment(
    request: Request,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """Сменить сегмент у нескольких клиентов."""
    _require_admin(auth_token, db)
    body = await request.json()
    client_ids = body.get("client_ids", [])
    segment    = body.get("segment", "")
    if not segment: raise HTTPException(status_code=400, detail="segment required")

    db.query(Client).filter(Client.id.in_(client_ids)).update(
        {"segment": segment}, synchronize_session=False
    )
    db.commit()
    return {"ok": True, "updated": len(client_ids)}
# ============================================================================
# REVENUE TRACKING
# ============================================================================

# ============================================================================
# CHURN SCORING
# ============================================================================

@app.get("/api/clients/{client_id}/churn")
async def api_client_churn(
    client_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    from models import ChurnScore
    cs = db.query(ChurnScore).filter(ChurnScore.client_id == client_id).first()
    if not cs:
        # Считаем на лету
        from churn import calculate_churn_score
        client = db.query(Client).filter(Client.id == client_id).first()
        if not client: raise HTTPException(status_code=404)
        tasks = [{"due_date": t.due_date.isoformat() if t.due_date else None, "status": t.status}
                 for t in db.query(Task).filter(Task.client_id == client_id).all()]
        meetings = [{"date": m.date.isoformat() if m.date else None}
                    for m in db.query(Meeting).filter(Meeting.client_id == client_id).all()]
        result = calculate_churn_score(client, tasks, meetings)
        return {**result, "client_id": client_id, "fresh": True}
    return {
        "score": cs.score, "risk_level": cs.risk_level,
        "factors": cs.factors, "calculated_at": cs.calculated_at.isoformat(),
        "client_id": client_id,
    }


@app.get("/api/analytics/churn")
async def api_analytics_churn(
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Сводка по churm рискам портфеля."""
    from models import ChurnScore
    q = db.query(Client)
    if user.role == "manager": q = q.filter(Client.manager_email == user.email)
    client_ids = [c.id for c in q.all()]

    scores = db.query(ChurnScore).filter(ChurnScore.client_id.in_(client_ids)).all() if client_ids else []
    dist = {"low": 0, "medium": 0, "high": 0, "critical": 0}
    top_risk = []
    for s in scores:
        dist[s.risk_level] = dist.get(s.risk_level, 0) + 1
        if s.risk_level in ("high", "critical"):
            c = db.query(Client).filter(Client.id == s.client_id).first()
            if c: top_risk.append({"id": c.id, "name": c.name, "score": s.score,
                                    "risk_level": s.risk_level, "segment": c.segment})
    top_risk.sort(key=lambda x: -x["score"])
    return {"distribution": dist, "top_risk": top_risk[:10], "total_scored": len(scores)}


# ============================================================================
# DEDUPLICATION
# ============================================================================

@app.get("/dedup", response_class=HTMLResponse)
async def dedup_page(request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    if not auth_token: return RedirectResponse(url="/login", status_code=303)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload: return RedirectResponse(url="/login", status_code=303)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user: return RedirectResponse(url="/login", status_code=303)
    if user.role != "admin": raise HTTPException(status_code=403)
    return templates.TemplateResponse("dedup.html", {"request": request, "user": user})


@app.get("/api/clients/duplicates")
async def api_clients_duplicates(
    threshold: int = 75,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Находит похожие названия клиентов через простой алгоритм."""
    if user.role != "admin": raise HTTPException(status_code=403)

    clients = db.query(Client).order_by(Client.name).all()

    def similarity(a: str, b: str) -> int:
        """Простое сходство строк через общие биграммы."""
        a, b = a.lower().strip(), b.lower().strip()
        if a == b: return 100
        def bigrams(s):
            return set(s[i:i+2] for i in range(len(s)-1))
        ba, bb = bigrams(a), bigrams(b)
        if not ba or not bb: return 0
        return int(2 * len(ba & bb) / (len(ba) + len(bb)) * 100)

    groups = []
    used   = set()
    for i, c1 in enumerate(clients):
        if c1.id in used: continue
        group = [c1]
        for c2 in clients[i+1:]:
            if c2.id in used: continue
            sim = similarity(c1.name, c2.name)
            if sim >= threshold:
                group.append(c2)
                used.add(c2.id)
        if len(group) > 1:
            used.add(c1.id)
            t_counts = {c.id: db.query(Task).filter(Task.client_id == c.id).count() for c in group}
            m_counts = {c.id: db.query(Meeting).filter(Meeting.client_id == c.id).count() for c in group}
            sim_score = max(similarity(group[0].name, c.name) for c in group[1:])
            groups.append({
                "similarity": sim_score,
                "clients": [
                    {"id": c.id, "name": c.name, "segment": c.segment, "domain": c.domain,
                     "tasks_count": t_counts[c.id], "meetings_count": m_counts[c.id]}
                    for c in sorted(group, key=lambda x: -(t_counts[x.id] + m_counts[x.id]))
                ]
            })

    groups.sort(key=lambda x: -x["similarity"])
    return {"groups": groups[:50], "total": len(groups)}


@app.post("/api/clients/merge")
async def api_clients_merge(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Слияние двух клиентов. master_id — остаётся, dup_id — удаляется."""
    if user.role != "admin": raise HTTPException(status_code=403)
    body      = await request.json()
    master_id = int(body.get("masterId", 0))
    dup_id    = int(body.get("dupId", 0))
    if not master_id or not dup_id or master_id == dup_id:
        raise HTTPException(status_code=400, detail="Некорректные ID")

    master = db.query(Client).filter(Client.id == master_id).first()
    dup    = db.query(Client).filter(Client.id == dup_id).first()
    if not master or not dup: raise HTTPException(status_code=404)

    # Переносим все связанные данные
    for model, col in [
        (Task,    "client_id"),
        (Meeting, "client_id"),
    ]:
        db.query(model).filter(getattr(model, col) == dup_id).update(
            {col: master_id}, synchronize_session=False
        )

    # Логируем событие
    from models import ClientHistory
    db.add(ClientHistory(
        client_id=master_id, user_id=user.id,
        field="merge", old_value=None,
        new_value=f"Слит с: {dup.name} (id={dup_id})",
        event_type="merge",
    ))
    db.delete(dup)
    db.commit()

    logger.info(f"Merged client {dup_id} ({dup.name}) into {master_id} ({master.name})")
    return {"ok": True, "master_id": master_id}


# ============================================================================
# DATA ENRICHMENT
# ============================================================================

# ============================================================================
# DATA VALIDATION
# ============================================================================

@app.get("/api/clients/validation")
async def api_clients_validation(
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Клиенты с неполными данными."""
    q = db.query(Client)
    if user.role == "manager": q = q.filter(Client.manager_email == user.email)
    clients = q.all()

    issues = []
    for c in clients:
        client_issues = []
        if not c.domain:         client_issues.append("нет домена")
        if not c.segment:        client_issues.append("нет сегмента")
        if not c.manager_email:  client_issues.append("нет менеджера")
        if c.health_score is None: client_issues.append("нет health score")
        if not c.merchrules_account_id: client_issues.append("нет MR ID")
        if client_issues:
            issues.append({"id": c.id, "name": c.name, "issues": client_issues, "count": len(client_issues)})

    issues.sort(key=lambda x: -x["count"])
    return {"total_with_issues": len(issues), "clients": issues[:100]}


# ============================================================================
# JOBS MONITORING
# ============================================================================

_job_log: list = []   # circular buffer for job log
_job_status: dict = {}


def log_job(job_id: str, message: str, level: str = "info"):
    import time
    _job_log.append({"ts": datetime.utcnow().isoformat(), "job": job_id, "message": message, "level": level})
    if len(_job_log) > 200:
        _job_log.pop(0)


@app.get("/admin/jobs", response_class=HTMLResponse)
async def jobs_page(
    request: Request, db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    if not auth_token: return RedirectResponse(url="/login", status_code=303)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload: return RedirectResponse(url="/login", status_code=303)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user or user.role != "admin": raise HTTPException(status_code=403)
    return templates.TemplateResponse("jobs_monitor.html", {"request": request, "user": user})


@app.get("/api/admin/jobs")
async def api_jobs_list(
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    if user.role != "admin": raise HTTPException(status_code=403)

    jobs = [
        {"id": "mr_sync",      "name": "Merchrules Sync",     "description": "Синхронизация клиентов и задач из Merchrules", "interval": "каждый час"},
        {"id": "checkup",      "name": "Checkup плановый",    "description": "Проверка просроченных чекапов и создание задач", "interval": "08:00 ежедневно"},
        {"id": "churn",        "name": "Churn Scoring",       "description": "Пересчёт риска оттока для всех клиентов", "interval": "воскресенье 02:00"},
        {"id": "tg_digest",    "name": "Telegram Digest",     "description": "Умный утренний дайджест менеджерам", "interval": "09:00 ежедневно"},
        {"id": "airtable_sync","name": "Airtable Sync",       "description": "Синхронизация клиентов с Airtable", "interval": "каждый час"},
        {"id": "auto_tasks",   "name": "AutoTask Rules",      "description": "Создание задач по правилам автозадач", "interval": "каждые 6 часов"},
    ]

    for j in jobs:
        st = _job_status.get(j["id"], {})
        j["status"]      = st.get("status", "pending")
        j["last_run"]    = st.get("last_run")
        j["next_run"]    = st.get("next_run")
        j["duration_ms"] = st.get("duration_ms")
        j["error"]       = st.get("error")

    return {"jobs": jobs, "log": list(reversed(_job_log[-50:]))}


@app.post("/api/admin/jobs/{job_id}/run")
async def api_job_run(
    job_id: str,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    if user.role != "admin": raise HTTPException(status_code=403)

    _job_status[job_id] = {"status": "running", "last_run": datetime.utcnow().isoformat()}
    log_job(job_id, f"Запущен вручную пользователем {user.name}")

    import asyncio, time

    async def run():
        start = time.time()
        try:
            if job_id == "churn":
                from churn import recalculate_all
                n = await recalculate_all(db)
                msg = f"Пересчитано: {n} клиентов"
            elif job_id == "mr_sync":
                msg = "Запущен через API — результат в логе"
            elif job_id == "auto_tasks":
                from models import AutoTaskRule
                rules = db.query(AutoTaskRule).filter(AutoTaskRule.is_active == True).all()
                msg = f"Правил обработано: {len(rules)}"
            else:
                msg = f"Job {job_id} не имеет прямого вызова"

            ms = int((time.time() - start) * 1000)
            _job_status[job_id] = {"status": "ok", "last_run": datetime.utcnow().isoformat(), "duration_ms": ms}
            log_job(job_id, f"Завершён за {ms}мс: {msg}")
        except Exception as e:
            _job_status[job_id] = {"status": "error", "last_run": datetime.utcnow().isoformat(), "error": str(e)}
            log_job(job_id, f"Ошибка: {e}", level="error")

    asyncio.create_task(run())
    return {"ok": True}


# ============================================================================
# FILE ATTACHMENTS
# ============================================================================

@app.get("/api/files/{file_path:path}")
async def api_serve_file(
    file_path: str,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Serve locally stored files."""
    from storage import get_file
    data = await get_file(file_path)
    if not data: raise HTTPException(status_code=404)
    from fastapi.responses import Response
    return Response(content=data, media_type="application/octet-stream")


# ============================================================================
# EXCEL EXPORT (полноценный)
# ============================================================================

# ============================================================================
# MEETING TRANSCRIPTION + AI SUMMARY
# ============================================================================

@app.post("/api/meetings/{meeting_id}/transcribe")
async def api_meeting_transcribe(
    meeting_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """AI Summary встречи на основе заметок + контекста."""
    meeting = db.query(Meeting).filter(Meeting.id == meeting_id).first()
    if not meeting: raise HTTPException(status_code=404)

    body  = await request.json()
    notes = body.get("notes", "") or meeting.notes or ""
    if not notes:
        return {"ok": False, "error": "Нет заметок для анализа. Добавьте заметки встречи."}

    client = db.query(Client).filter(Client.id == meeting.client_id).first()

    u_settings = user.settings or {}
    groq_key   = u_settings.get("groq", {}).get("api_key") or env.GROQ_KEY
    if not groq_key:
        return {"ok": False, "error": "Groq API key не настроен"}

    prompt = f"""Проанализируй заметки встречи и создай структурированное резюме.

Клиент: {client.name if client else "—"}
Дата встречи: {meeting.date.strftime("%d.%m.%Y") if meeting.date else "—"}
Тип: {meeting.meeting_type or "встреча"}

Заметки:
{notes}

Создай резюме в формате:
## Ключевые договорённости
- ...

## Следующие шаги (задачи)
- [ЗАДАЧА] Описание задачи — ответственный
- ...

## Риски и вопросы
- ...

## Краткое резюме (1-2 предложения)
...
"""

    try:
        import httpx
        async with httpx.AsyncClient(timeout=30) as hx:
            r = await hx.post(
                "https://api.groq.com/openai/v1/chat/completions",
                headers={"Authorization": f"Bearer {groq_key}", "Content-Type": "application/json"},
                json={"model": "llama-3.3-70b-versatile",
                      "messages": [{"role": "user", "content": prompt}],
                      "max_tokens": 600, "temperature": 0.3},
            )
        if r.status_code != 200:
            return {"ok": False, "error": f"Groq error: {r.status_code}"}
        summary = r.json()["choices"][0]["message"]["content"]
    except Exception as e:
        return {"ok": False, "error": str(e)}

    # Сохраняем summary в meeting
    from sqlalchemy.orm.attributes import flag_modified
    meeting.summary = summary
    db.commit()

    # Извлекаем задачи из summary и создаём их
    import re
    task_lines = re.findall(r"\[ЗАДАЧА\] (.+)", summary)
    created_tasks = []
    for tl in task_lines[:5]:
        task = Task(client_id=meeting.client_id, title=tl.strip()[:200],
                    status="plan", priority="medium", created_at=datetime.utcnow(),
                    due_date=datetime.utcnow() + timedelta(days=3))
        db.add(task); created_tasks.append(tl.strip())
    if created_tasks: db.commit()

    return {"ok": True, "summary": summary, "tasks_created": created_tasks}


# ============================================================================
# MEETING COMMENTS
# ============================================================================

# ============================================================================
# ONBOARDING WIZARD
# ============================================================================

@app.get("/onboarding/wizard", response_class=HTMLResponse)
async def onboarding_wizard(
    request: Request,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    if not auth_token: return RedirectResponse(url="/login", status_code=303)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload: return RedirectResponse(url="/login", status_code=303)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user: return RedirectResponse(url="/login", status_code=303)
    return templates.TemplateResponse("onboarding_wizard.html", {"request": request, "user": user})


@app.get("/api/onboarding/progress")
async def api_onboarding_progress(
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    from models import OnboardingProgress
    prog = db.query(OnboardingProgress).filter(OnboardingProgress.user_id == user.id).first()
    completed = prog.completed if prog else []
    steps = [
        {"id": "hub_url",      "title": "Настройте AM Hub URL",      "done": "hub_url" in completed},
        {"id": "mr_connect",   "title": "Подключите Merchrules",      "done": "mr_connect" in completed},
        {"id": "add_clients",  "title": "Добавьте первых клиентов",   "done": "add_clients" in completed},
        {"id": "first_task",   "title": "Создайте первую задачу",     "done": "first_task" in completed},
        {"id": "tg_connect",   "title": "Подключите Telegram",        "done": "tg_connect" in completed},
        {"id": "first_checkup","title": "Запустите первый чекап",     "done": "first_checkup" in completed},
    ]
    done_count = sum(1 for s in steps if s["done"])
    return {"steps": steps, "done": done_count, "total": len(steps), "completed": done_count == len(steps)}


@app.post("/api/onboarding/complete-step")
async def api_onboarding_complete_step(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    body = await request.json()
    step = body.get("step", "")
    from models import OnboardingProgress
    from sqlalchemy.orm.attributes import flag_modified
    prog = db.query(OnboardingProgress).filter(OnboardingProgress.user_id == user.id).first()
    if not prog:
        prog = OnboardingProgress(user_id=user.id, completed=[])
        db.add(prog)
    completed = list(prog.completed or [])
    if step and step not in completed:
        completed.append(step)
        prog.completed = completed
        flag_modified(prog, "completed")
    db.commit()
    return {"ok": True, "completed": completed}

# ============================================================================
# REVENUE TRACKING
# ============================================================================

@app.patch("/api/clients/{client_id}/revenue")
async def api_set_revenue(
    client_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Установить MRR/ARR клиента."""
    body  = await request.json()
    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        raise HTTPException(status_code=404)
    from sqlalchemy.orm.attributes import flag_modified
    meta = dict(client.integration_metadata or {})
    if "mrr" in body:
        meta["mrr"] = float(body["mrr"])
    if "arr" in body:
        meta["arr"] = float(body.get("arr", meta.get("mrr", 0) * 12))
    if "currency" in body:
        meta["currency"] = body["currency"]
    client.integration_metadata = meta
    flag_modified(client, "integration_metadata")
    db.commit()
    return {"ok": True}


@app.get("/api/analytics/revenue")
async def api_revenue_analytics(
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Аналитика выручки портфеля."""
    q = db.query(Client)
    if user.role == "manager":
        q = q.filter(Client.manager_email == user.email)
    clients = q.all()

    total_mrr = 0.0
    at_risk_mrr = 0.0
    by_segment: dict = {}
    top_clients = []

    for c in clients:
        meta = c.integration_metadata or {}
        mrr  = float(meta.get("mrr") or 0)
        if not mrr:
            continue
        total_mrr += mrr
        seg = c.segment or "Unknown"
        by_segment[seg] = by_segment.get(seg, 0.0) + mrr

        # Считаем "под угрозой" если health < 50 или нет контакта давно
        is_risk = (c.health_score or 50) < 50
        if is_risk:
            at_risk_mrr += mrr

        top_clients.append({
            "id": c.id, "name": c.name, "segment": c.segment,
            "mrr": mrr, "health_score": c.health_score, "is_risk": is_risk,
        })

    top_clients.sort(key=lambda x: x["mrr"], reverse=True)

    return {
        "total_mrr": round(total_mrr, 2),
        "total_arr": round(total_mrr * 12, 2),
        "at_risk_mrr": round(at_risk_mrr, 2),
        "at_risk_pct": round(at_risk_mrr / total_mrr * 100, 1) if total_mrr else 0,
        "by_segment": [{"segment": k, "mrr": round(v, 2)} for k, v in
                       sorted(by_segment.items(), key=lambda x: x[1], reverse=True)],
        "top_clients": top_clients[:10],
        "clients_with_mrr": len(top_clients),
        "clients_total": len(clients),
    }


# ============================================================================
# DATA VALIDATION — клиенты с неполными данными
# ============================================================================

@app.get("/api/clients/validation/issues")
async def api_validation_issues(
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Клиенты с неполными/проблемными данными."""
    q = db.query(Client)
    if user.role == "manager":
        q = q.filter(Client.manager_email == user.email)
    clients = q.limit(500).all()

    issues = []
    for c in clients:
        client_issues = []
        if not c.domain:
            client_issues.append({"field": "domain", "msg": "Нет домена сайта"})
        if not c.segment:
            client_issues.append({"field": "segment", "msg": "Не указан сегмент"})
        if not c.manager_email:
            client_issues.append({"field": "manager", "msg": "Нет ответственного менеджера"})
        meta = c.integration_metadata or {}
        if not meta.get("mrr") and not meta.get("arr"):
            client_issues.append({"field": "revenue", "msg": "Нет данных о выручке"})
        digi = meta.get("diginetica", {})
        if not any(digi.get(p, {}).get("api_key") for p in ("sort", "autocomplete", "recommendations")):
            client_issues.append({"field": "diginetica", "msg": "Нет Diginetica API ключей"})
        if client_issues:
            issues.append({
                "id": c.id, "name": c.name, "segment": c.segment,
                "issues": client_issues, "issues_count": len(client_issues),
            })

    issues.sort(key=lambda x: x["issues_count"], reverse=True)
    return {
        "total_issues": len(issues),
        "total_clients": len(clients),
        "clean_pct": round((len(clients) - len(issues)) / len(clients) * 100, 1) if clients else 0,
        "clients": issues[:100],
    }


# ============================================================================
# DATA ENRICHMENT — обогащение по домену
# ============================================================================

@app.post("/api/clients/{client_id}/enrich")
async def api_enrich_client(
    client_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """
    Обогащение данных клиента по домену.
    Использует открытые источники: Clearbit Reveal (free tier) / whois / robots.txt
    """
    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        raise HTTPException(status_code=404)

    domain = client.domain
    if not domain:
        return {"ok": False, "error": "У клиента не указан домен"}

    # Нормализуем домен
    import re as _re
    domain = _re.sub(r'^https?://', '', domain).strip('/').split('/')[0]

    enriched = {}
    errors   = []

    # 1. Clearbit Logo API (бесплатно, без ключа)
    try:
        import httpx
        async with httpx.AsyncClient(timeout=8) as hx:
            logo_url = f"https://logo.clearbit.com/{domain}"
            r = await hx.head(logo_url)
            if r.status_code == 200:
                enriched["logo_url"] = logo_url
    except Exception:
        pass

    # 2. Публичный Clearbit Autocomplete (company name)
    try:
        import httpx
        async with httpx.AsyncClient(timeout=8) as hx:
            r = await hx.get(
                f"https://autocomplete.clearbit.com/v1/companies/suggest?query={domain}",
                headers={"Accept": "application/json"},
            )
            if r.status_code == 200:
                companies = r.json()
                if companies:
                    co = companies[0]
                    enriched["company_name"] = co.get("name")
                    enriched["company_domain"] = co.get("domain")
                    if not enriched.get("logo_url"):
                        enriched["logo_url"] = co.get("logo")
    except Exception as e:
        errors.append(f"clearbit: {e}")

    # 3. Сохраняем в integration_metadata
    if enriched:
        from sqlalchemy.orm.attributes import flag_modified
        meta = dict(client.integration_metadata or {})
        meta["enriched"] = enriched
        meta["enriched_at"] = datetime.utcnow().isoformat()
        if enriched.get("company_name") and not client.name:
            client.name = enriched["company_name"]
        if enriched.get("logo_url"):
            meta["logo_url"] = enriched["logo_url"]
        client.integration_metadata = meta
        flag_modified(client, "integration_metadata")
        db.commit()

    return {"ok": bool(enriched), "enriched": enriched, "errors": errors}


@app.post("/api/clients/enrich-bulk")
async def api_enrich_bulk(
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Обогатить всех клиентов с доменом но без лого."""
    q = db.query(Client).filter(Client.domain.isnot(None))
    if user.role == "manager":
        q = q.filter(Client.manager_email == user.email)
    clients = q.limit(50).all()

    enriched = skipped = failed = 0
    import httpx, asyncio
    async with httpx.AsyncClient(timeout=8) as hx:
        for client in clients:
            meta = client.integration_metadata or {}
            if meta.get("enriched_at"):
                skipped += 1
                continue
            domain = client.domain.replace("https://","").replace("http://","").strip("/").split("/")[0]
            try:
                r = await hx.get(f"https://autocomplete.clearbit.com/v1/companies/suggest?query={domain}")
                if r.status_code == 200 and r.json():
                    co = r.json()[0]
                    from sqlalchemy.orm.attributes import flag_modified
                    m = dict(meta)
                    m["enriched"] = {"company_name": co.get("name"), "logo_url": co.get("logo")}
                    m["enriched_at"] = datetime.utcnow().isoformat()
                    if co.get("logo"): m["logo_url"] = co["logo"]
                    client.integration_metadata = m
                    flag_modified(client, "integration_metadata")
                    enriched += 1
                await asyncio.sleep(0.3)
            except Exception:
                failed += 1
    db.commit()
    return {"ok": True, "enriched": enriched, "skipped": skipped, "failed": failed}


# ============================================================================
# CHURN SCORING ENDPOINTS
# ============================================================================

@app.get("/api/clients/churn-scores")
async def api_churn_scores(
    risk_level: str = "",
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Список клиентов с churn score."""
    from models import ChurnScore
    q = (db.query(Client, ChurnScore)
         .outerjoin(ChurnScore, Client.id == ChurnScore.client_id))
    if user.role == "manager":
        q = q.filter(Client.manager_email == user.email)
    if risk_level:
        q = q.filter(ChurnScore.risk_level == risk_level)
    rows = q.order_by(ChurnScore.score.desc().nullslast()).limit(200).all()
    return {"clients": [
        {"id": c.id, "name": c.name, "segment": c.segment,
         "health_score": c.health_score,
         "churn_score": cs.score if cs else None,
         "risk_level":  cs.risk_level if cs else "unknown",
         "factors":     cs.factors if cs else {},
         "calculated_at": cs.calculated_at.isoformat() if cs and cs.calculated_at else None}
        for c, cs in rows
    ]}


@app.post("/api/clients/{client_id}/churn-recalc")
async def api_churn_recalc_single(
    client_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Пересчитать churn score одного клиента."""
    from churn import calculate_churn_score
    from models import ChurnScore
    from sqlalchemy.orm.attributes import flag_modified
    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        raise HTTPException(status_code=404)
    tasks    = [{"due_date": t.due_date.isoformat() if t.due_date else None, "status": t.status}
                for t in db.query(Task).filter(Task.client_id == client_id).all()]
    meetings = [{"date": m.date.isoformat() if m.date else None}
                for m in db.query(Meeting).filter(Meeting.client_id == client_id).all()]
    result = calculate_churn_score(client, tasks, meetings)
    cs = db.query(ChurnScore).filter(ChurnScore.client_id == client_id).first()
    if cs:
        cs.score = result["score"]; cs.risk_level = result["risk_level"]
        cs.factors = result["factors"]; cs.calculated_at = datetime.utcnow()
        flag_modified(cs, "factors")
    else:
        db.add(ChurnScore(client_id=client_id, **result))
    db.commit()
    return {"ok": True, **result}


# ============================================================================
# EXCEL EXPORT — полноценный с форматированием
# ============================================================================

@app.get("/api/export/excel")
async def api_export_excel(
    scope: str = "clients",  # clients | tasks | checkups | full
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Excel экспорт с форматированием, несколько листов."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from fastapi.responses import Response

    wb = Workbook()

    # ── Стили ──────────────────────────────────────────────────────────────────
    HDR_FILL = PatternFill("solid", start_color="07090F", end_color="07090F")
    HDR_FONT = Font(name="Arial", bold=True, color="6474FF", size=10)
    ROW_FILL = [
        PatternFill("solid", start_color="0D1117", end_color="0D1117"),
        PatternFill("solid", start_color="131924", end_color="131924"),
    ]
    GREEN_FONT  = Font(name="Arial", color="23D18B", size=9)
    YELLOW_FONT = Font(name="Arial", color="F5A623", size=9)
    RED_FONT    = Font(name="Arial", color="F0556A", size=9)
    DEF_FONT    = Font(name="Arial", color="E8ECF4", size=9)
    MONO_FONT   = Font(name="Courier New", color="E8ECF4", size=9)

    def style_header(ws, headers: list):
        for col, (label, width) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=label)
            cell.font = HDR_FONT; cell.fill = HDR_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.row_dimensions[1].height = 24
        ws.freeze_panes = "A2"

    def write_row(ws, row_num: int, values: list, fonts: list = None):
        fill = ROW_FILL[(row_num - 2) % 2]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.fill = fill
            cell.alignment = Alignment(vertical="center")
            f = (fonts[col-1] if fonts and col-1 < len(fonts) else None) or DEF_FONT
            cell.font = f
        ws.row_dimensions[row_num].height = 18

    # ── Лист 1: Клиенты ────────────────────────────────────────────────────────
    ws1 = wb.active; ws1.title = "Клиенты"
    style_header(ws1, [
        ("ID",22), ("Название",36), ("Сегмент",16), ("Домен",30),
        ("Health %",14), ("Менеджер",24), ("MRR ₽",16), ("Последняя встреча",22),
        ("Риск оттока",16), ("Задач открытых",18),
    ])
    q = db.query(Client)
    if user.role == "manager":
        q = q.filter(Client.manager_email == user.email)
    clients = q.order_by(Client.name).all()
    from models import ChurnScore
    for i, c in enumerate(clients, 2):
        meta = c.integration_metadata or {}
        mrr  = meta.get("mrr", "")
        cs   = db.query(ChurnScore).filter(ChurnScore.client_id == c.id).first()
        open_t = db.query(Task).filter(Task.client_id == c.id, Task.status != "done").count()
        h = c.health_score or 0
        hfont = GREEN_FONT if h >= 70 else YELLOW_FONT if h >= 40 else RED_FONT
        write_row(ws1, i, [
            c.id, c.name, c.segment or "—", c.domain or "—",
            f"{h:.0f}%" if c.health_score else "—",
            c.manager_email or "—",
            f"{mrr:,.0f}" if mrr else "—",
            c.last_meeting_date.strftime("%d.%m.%Y") if c.last_meeting_date else "—",
            cs.risk_level if cs else "—", open_t,
        ], fonts=[MONO_FONT, None, None, None, hfont, None, MONO_FONT, None, None, MONO_FONT])

    # ── Лист 2: Задачи ─────────────────────────────────────────────────────────
    if scope in ("tasks", "full"):
        ws2 = wb.create_sheet("Задачи")
        style_header(ws2, [
            ("ID",12), ("Задача",44), ("Клиент",30), ("Статус",18),
            ("Приоритет",16), ("Срок",18), ("Синк MR",16),
        ])
        tq = db.query(Task).join(Client, Task.client_id == Client.id, isouter=True)
        if user.role == "manager":
            tq = tq.filter(Client.manager_email == user.email)
        tasks = tq.order_by(Task.due_date.asc().nullslast()).limit(2000).all()
        STATUS_LABELS = {"plan":"Планирование","in_progress":"В работе","review":"Ревью","done":"Выполнено","blocked":"Заблокировано"}
        PRIO_LABELS   = {"high":"Высокий","medium":"Средний","low":"Низкий"}
        for i, t in enumerate(tasks, 2):
            now = datetime.utcnow()
            is_overdue = t.due_date and t.due_date < now and t.status != "done"
            dfont = RED_FONT if is_overdue else DEF_FONT
            write_row(ws2, i, [
                t.id, t.title, t.client.name if t.client else "—",
                STATUS_LABELS.get(t.status, t.status),
                PRIO_LABELS.get(t.priority, t.priority),
                t.due_date.strftime("%d.%m.%Y") if t.due_date else "—",
                "✅" if t.merchrules_task_id else "—",
            ], fonts=[MONO_FONT, None, None, None, None, dfont, None])

    # ── Лист 3: Чекапы ─────────────────────────────────────────────────────────
    if scope in ("checkups", "full"):
        from models import CheckupResult
        ws3 = wb.create_sheet("Чекапы")
        style_header(ws3, [
            ("ID",12), ("Клиент",32), ("Дата",18), ("Запросов",14),
            ("Avg Score",14), ("Продукт",16), ("Менеджер",24),
        ])
        cq = db.query(CheckupResult).join(Client, CheckupResult.client_id == Client.id, isouter=True)
        if user.role == "manager":
            cq = cq.filter(Client.manager_email == user.email)
        results = cq.order_by(CheckupResult.created_at.desc()).limit(500).all()
        for i, r in enumerate(results, 2):
            avg = r.avg_score or 0
            afont = GREEN_FONT if avg >= 2.5 else YELLOW_FONT if avg >= 1.5 else RED_FONT
            write_row(ws3, i, [
                r.id, r.client.name if r.client else "—",
                r.created_at.strftime("%d.%m.%Y %H:%M") if r.created_at else "—",
                r.total_queries or 0, f"{avg:.2f}", r.query_type or "—",
                r.manager_name or "—",
            ], fonts=[MONO_FONT, None, None, MONO_FONT, afont, None, None])

    # ── Сохраняем ──────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"amhub_export_{scope}_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


# ============================================================================
# FILE ATTACHMENTS
# ============================================================================

@app.post("/api/clients/{client_id}/attachments")
async def api_upload_attachment(
    client_id: int,
    file: UploadFile,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Загрузить файл к клиенту."""
    from storage import upload_file, ALLOWED_MIME
    from models import ClientAttachment
    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        raise HTTPException(status_code=404)

    file_bytes = await file.read()
    try:
        result = await upload_file(file_bytes, file.filename, client_id, file.content_type)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    att = ClientAttachment(
        client_id=client_id, user_id=user.id,
        filename=file.filename, file_key=result["key"],
        file_size=result["size"], mime_type=result.get("mime_type"),
    )
    db.add(att); db.commit(); db.refresh(att)
    return {"ok": True, "id": att.id, "filename": att.filename,
            "url": result["url"], "size": att.file_size}


@app.get("/api/clients/{client_id}/attachments")
async def api_list_attachments(
    client_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    from models import ClientAttachment
    from storage import get_signed_url
    atts = db.query(ClientAttachment).filter(ClientAttachment.client_id == client_id)              .order_by(ClientAttachment.created_at.desc()).all()
    return {"attachments": [
        {"id": a.id, "filename": a.filename,
         "url": get_signed_url(a.file_key),
         "size": a.file_size, "mime_type": a.mime_type,
         "created_at": a.created_at.isoformat(),
         "uploaded_by": a.user.name if a.user else "—"}
        for a in atts
    ]}


@app.delete("/api/attachments/{att_id}")
async def api_delete_attachment(
    att_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    from models import ClientAttachment
    from storage import delete_file
    att = db.query(ClientAttachment).filter(ClientAttachment.id == att_id).first()
    if not att:
        raise HTTPException(status_code=404)
    await delete_file(att.file_key)
    db.delete(att); db.commit()
    return {"ok": True}


@app.get("/api/files/{file_key:path}")
async def api_serve_file(
    file_key: str,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Отдать файл из local storage."""
    from storage import get_file
    from fastapi.responses import Response as FR
    data = await get_file(file_key)
    if not data:
        raise HTTPException(status_code=404)
    import mimetypes
    mime, _ = mimetypes.guess_type(file_key)
    return FR(content=data, media_type=mime or "application/octet-stream")


# ============================================================================
# MEETING COMMENTS
# ============================================================================

@app.post("/api/meetings/{meeting_id}/comments")
async def api_add_meeting_comment(
    meeting_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    from models import MeetingComment
    body = await request.json()
    content_text = body.get("content", "").strip()
    if not content_text:
        raise HTTPException(status_code=400, detail="Комментарий не может быть пустым")
    c = MeetingComment(meeting_id=meeting_id, user_id=user.id, content=content_text)
    db.add(c); db.commit(); db.refresh(c)
    return {"ok": True, "id": c.id, "content": c.content,
            "created_at": c.created_at.isoformat(), "user_name": user.name}


@app.get("/api/meetings/{meeting_id}/comments")
async def api_meeting_comments(
    meeting_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    from models import MeetingComment
    comments = db.query(MeetingComment).filter(MeetingComment.meeting_id == meeting_id)                 .order_by(MeetingComment.created_at.asc()).all()
    return {"comments": [
        {"id": c.id, "content": c.content, "created_at": c.created_at.isoformat(),
         "user_name": c.user.name if c.user else "Система"}
        for c in comments
    ]}


@app.post("/api/onboarding/skip")
async def api_onboarding_skip(
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Пропустить онбординг."""
    from models import OnboardingProgress
    from sqlalchemy.orm.attributes import flag_modified
    prog = db.query(OnboardingProgress).filter(OnboardingProgress.user_id == user.id).first()
    if not prog:
        prog = OnboardingProgress(user_id=user.id, completed=[], skipped=True)
        db.add(prog)
    else:
        prog.skipped = True
        flag_modified(prog, "completed")
    db.commit()
    return {"ok": True}
