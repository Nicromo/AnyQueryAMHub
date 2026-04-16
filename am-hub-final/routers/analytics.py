"""
Auto-extracted router — do not edit the @app registrations manually.
"""
from typing import Optional, List
from datetime import datetime, timedelta
import os
import json
import logging

from fastapi import APIRouter, Request, Depends, HTTPException, Query, Cookie, Form, status
from fastapi.responses import JSONResponse, HTMLResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from sqlalchemy import text

from database import get_db, SessionLocal
from models import (
    Client, Task, Meeting, CheckUp, User, SyncLog, AuditLog,
    Notification, QBR, AccountPlan, ClientNote, TaskComment,
    FollowupTemplate, VoiceNote,
)
from auth import (
    authenticate_user, create_user, create_access_token,
    verify_password, hash_password, log_audit, decode_access_token,
)
from error_handlers import log_error, handle_db_error

logger = logging.getLogger(__name__)
templates = Jinja2Templates(directory="templates")

router = APIRouter()

def _env(key: str, default: str = "") -> str:
    return os.environ.get(key, default)

def _env_bool(key: str) -> bool:
    return bool(os.environ.get(key, ""))

@router.get("/api/analytics/overview")
async def api_analytics_overview(
    days: int = 30,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    if not auth_token: raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload: raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user: raise HTTPException(status_code=401)

    since = datetime.utcnow() - timedelta(days=days)
    q = db.query(Client)
    if user.role == "manager": q = q.filter(Client.manager_email == user.email)
    clients = q.all()

    health_vals = [c.health_score for c in clients if c.health_score is not None]
    avg_health  = sum(health_vals) / len(health_vals) if health_vals else 0

    # Сегменты
    from collections import Counter
    seg_counter = Counter(c.segment or "Unknown" for c in clients)
    segments = [{"segment": k, "count": v} for k, v in seg_counter.most_common()]

    # Health distribution
    health_good = sum(1 for h in health_vals if h >= 70)
    health_warn = sum(1 for h in health_vals if 40 <= h < 70)
    health_bad  = sum(1 for h in health_vals if h < 40)

    # Tasks
    cids = [c.id for c in clients]
    open_tasks    = db.query(Task).filter(Task.client_id.in_(cids), Task.status != "done").count() if cids else 0
    overdue_tasks = db.query(Task).filter(
        Task.client_id.in_(cids), Task.status != "done",
        Task.due_date < datetime.utcnow()
    ).count() if cids else 0

    # Meetings + followups за период
    meetings_count  = db.query(Meeting).filter(Meeting.client_id.in_(cids), Meeting.date >= since).count() if cids else 0
    followups_count = 0  # TODO: followup model

    # Risk clients
    risk_clients = sorted(
        [{"id": c.id, "name": c.name, "segment": c.segment, "health_score": c.health_score}
         for c in clients if c.health_score is not None and c.health_score < 60],
        key=lambda x: x["health_score"]
    )[:8]

    # Active clients (by meetings + tasks)
    active = []
    for c in clients:
        m_cnt = db.query(Meeting).filter(Meeting.client_id == c.id, Meeting.date >= since).count()
        t_cnt = db.query(Task).filter(Task.client_id == c.id, Task.created_at >= since).count()
        if m_cnt + t_cnt > 0:
            active.append({"id": c.id, "name": c.name, "activity_score": m_cnt * 3 + t_cnt})
    active.sort(key=lambda x: x["activity_score"], reverse=True)

    return {
        "total_clients": len(clients),
        "avg_health": avg_health,
        "open_tasks": open_tasks,
        "overdue_tasks": overdue_tasks,
        "meetings_count": meetings_count,
        "followups_count": followups_count,
        "segments": segments,
        "health_good": health_good,
        "health_warn": health_warn,
        "health_bad": health_bad,
        "risk_clients": risk_clients,
        "active_clients": active[:8],
    }



@router.get("/api/analytics/health-trend")
async def api_analytics_health_trend(
    days: int = 30,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    if not auth_token: raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload: raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user: raise HTTPException(status_code=401)

    # Генерируем точки по неделям на основе текущих данных (без истории)
    q = db.query(Client)
    if user.role == "manager": q = q.filter(Client.manager_email == user.email)
    clients = q.all()
    health_vals = [c.health_score for c in clients if c.health_score is not None]
    avg = sum(health_vals) / len(health_vals) if health_vals else 0

    # Симулируем тренд за период (заглушка до появления audit log)
    import random
    n_points = min(days // 7, 12)
    labels, values = [], []
    base = avg
    for i in range(n_points, 0, -1):
        d = datetime.utcnow() - timedelta(weeks=i)
        labels.append(d.strftime("%d.%m"))
        # Небольшой шум вокруг текущего значения
        values.append(round(max(0, min(100, base + random.uniform(-5, 5))), 1))
    labels.append("Сейчас")
    values.append(round(avg, 1))

    return {"labels": labels, "values": values}



@router.get("/api/analytics/tasks-stats")
async def api_analytics_tasks_stats(
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    if not auth_token: raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload: raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user: raise HTTPException(status_code=401)

    q = db.query(Client)
    if user.role == "manager": q = q.filter(Client.manager_email == user.email)
    cids = [c.id for c in q.all()]

    from collections import Counter
    tasks = db.query(Task).filter(Task.client_id.in_(cids)).all() if cids else []
    by_status = dict(Counter(t.status or "plan" for t in tasks))

    return {"by_status": by_status, "total": len(tasks)}



@router.get("/api/analytics/activity")
async def api_analytics_activity(
    days: int = 30,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    if not auth_token: raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload: raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user: raise HTTPException(status_code=401)

    q = db.query(Client)
    if user.role == "manager": q = q.filter(Client.manager_email == user.email)
    cids = [c.id for c in q.all()]

    # По неделям
    n_weeks = min(days // 7, 8)
    labels, meetings_data, tasks_data = [], [], []
    for i in range(n_weeks, 0, -1):
        week_start = datetime.utcnow() - timedelta(weeks=i)
        week_end   = datetime.utcnow() - timedelta(weeks=i-1)
        label = week_start.strftime("Нед %d.%m")
        m_cnt = db.query(Meeting).filter(Meeting.client_id.in_(cids), Meeting.date >= week_start, Meeting.date < week_end).count() if cids else 0
        t_cnt = db.query(Task).filter(Task.client_id.in_(cids), Task.created_at >= week_start, Task.created_at < week_end).count() if cids else 0
        labels.append(label); meetings_data.append(m_cnt); tasks_data.append(t_cnt)

    return {"labels": labels, "meetings": meetings_data, "tasks": tasks_data}



@router.get("/api/analytics/export")
async def api_analytics_export(
    days: int = 30,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """Экспорт аналитики в CSV."""
    if not auth_token: raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload: raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user: raise HTTPException(status_code=401)

    q = db.query(Client)
    if user.role == "manager": q = q.filter(Client.manager_email == user.email)
    clients = q.order_by(Client.name).all()

    import csv, io
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["ID","Название","Сегмент","Health Score","Последняя встреча","Открытых задач"])
    for c in clients:
        open_t = db.query(Task).filter(Task.client_id == c.id, Task.status != "done").count()
        w.writerow([c.id, c.name, c.segment, f"{c.health_score:.0f}%" if c.health_score else "—",
                    c.last_meeting_date.strftime("%d.%m.%Y") if c.last_meeting_date else "—", open_t])

    from fastapi.responses import Response
    return Response(content=buf.getvalue(), media_type="text/csv",
                    headers={"Content-Disposition": f"attachment; filename=analytics_{datetime.utcnow().strftime('%Y%m%d')}.csv"})


# ── Auto-task rules ─────────────────────────────────────────────────────────


@router.get("/api/analytics")
async def api_analytics(db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Данные для аналитики."""
    if not auth_token:
        return {}
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        return {}
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        return {}

    q = db.query(Client)
    if user.role == "manager":
        q = q.filter(Client.manager_email == user.email)
    clients = q.all()

    # Segments
    seg_counts = {}
    for c in clients:
        seg = c.segment or "other"
        seg_counts[seg] = seg_counts.get(seg, 0) + 1

    # Health distribution
    health_buckets = {"0-25": 0, "25-50": 0, "50-75": 0, "75-100": 0}
    for c in clients:
        score = (c.health_score or 0) * 100
        if score < 25:
            health_buckets["0-25"] += 1
        elif score < 50:
            health_buckets["25-50"] += 1
        elif score < 75:
            health_buckets["50-75"] += 1
        else:
            health_buckets["75-100"] += 1

    # Tasks by status
    task_q = db.query(Task).join(Client, Task.client_id == Client.id, isouter=True)
    if user.role == "manager":
        task_q = task_q.filter(Client.manager_email == user.email)
    all_tasks = task_q.all()
    task_status_counts = {}
    for t in all_tasks:
        s = t.status or "plan"
        task_status_counts[s] = task_status_counts.get(s, 0) + 1

    # Meetings per month (last 6 months)
    meetings_per_month = {}
    meeting_q = db.query(Meeting).join(Client, Meeting.client_id == Client.id, isouter=True)
    if user.role == "manager":
        meeting_q = meeting_q.filter(Client.manager_email == user.email)
    all_meetings = meeting_q.filter(Meeting.date != None).order_by(Meeting.date.desc()).all()
    for m in all_meetings:
        if m.date:
            key = m.date.strftime("%Y-%m")
            meetings_per_month[key] = meetings_per_month.get(key, 0) + 1

    return {
        "total_clients": len(clients),
        "segments": seg_counts,
        "health_distribution": health_buckets,
        "task_status": task_status_counts,
        "total_tasks": len(all_tasks),
        "meetings_per_month": dict(sorted(meetings_per_month.items(), reverse=True)[:6]),
        "avg_health": round(sum((c.health_score or 0) for c in clients) / max(len(clients), 1) * 100, 1),
    }


# ============================================================================
# BULK ACTIONS

@router.post("/api/bulk/checkups")
async def api_bulk_checkups(request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Массовое назначение чекапов."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)
    data = await request.json()
    client_ids = data.get("client_ids", [])
    date_str = data.get("date")
    meeting_date = datetime.fromisoformat(date_str) if date_str else datetime.now()
    created = 0
    for cid in client_ids:
        client = db.query(Client).filter(Client.id == cid).first()
        if client:
            m = Meeting(client_id=cid, date=meeting_date, type="checkup", source="internal", title="Чекап")
            db.add(m)
            client.last_meeting_date = meeting_date
            client.needs_checkup = False
            created += 1
    db.commit()
    return {"ok": True, "created": created}



@router.post("/api/bulk/segment")
async def api_bulk_segment(request: Request, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Массовое изменение сегмента."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)
    data = await request.json()
    client_ids = data.get("client_ids", [])
    segment = data.get("segment", "")
    updated = 0
    for cid in client_ids:
        client = db.query(Client).filter(Client.id == cid).first()
        if client:
            client.segment = segment
            updated += 1
    db.commit()
    return {"ok": True, "updated": updated}


# ============================================================================
# EXPORT

@router.get("/api/export/client/{client_id}")
async def api_export_client(client_id: int, fmt: str = "json", db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Экспорт отчёта по клиенту (JSON/CSV)."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)

    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        raise HTTPException(status_code=404)

    tasks = db.query(Task).filter(Task.client_id == client_id).all()
    meetings = db.query(Meeting).filter(Meeting.client_id == client_id).order_by(Meeting.date.desc()).all()
    notes = db.query(ClientNote).filter(ClientNote.client_id == client_id).all()

    data = {
        "client": {"id": client.id, "name": client.name, "segment": client.segment, "health_score": client.health_score, "manager_email": client.manager_email},
        "tasks": [{"id": t.id, "title": t.title, "status": t.status, "priority": t.priority, "due_date": t.due_date.isoformat() if t.due_date else None} for t in tasks],
        "meetings": [{"id": m.id, "title": m.title or m.type, "date": m.date.isoformat() if m.date else None, "type": m.type} for m in meetings],
        "notes": [{"id": n.id, "content": n.content, "pinned": n.is_pinned} for n in notes],
        "exported_at": datetime.utcnow().isoformat(),
    }

    if fmt == "csv":
        import io, csv
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["type", "id", "title", "date", "details"])
        for t in tasks:
            writer.writerow(["task", t.id, t.title, t.due_date.isoformat() if t.due_date else "", t.status])
        for m in meetings:
            writer.writerow(["meeting", m.id, m.title or m.type, m.date.isoformat() if m.date else "", m.type])
        for n in notes:
            writer.writerow(["note", n.id, n.content[:50], "", "pinned" if n.is_pinned else ""])
        from fastapi.responses import PlainTextResponse
        return PlainTextResponse(content=output.getvalue(), headers={"Content-Disposition": f"attachment; filename=client_{client_id}.csv"})

    return data


# ============================================================================
# AI RECOMMENDATIONS & AUTO-QBR

@router.get("/api/churn/risk")
async def api_churn_risk(db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Прогнозирование оттока: rule-based scoring."""
    if not auth_token:
        return {"clients": []}
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        return {"clients": []}
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        return {"clients": []}

    q = db.query(Client)
    if user.role == "manager":
        q = q.filter(Client.manager_email == user.email)
    clients = q.all()
    now = datetime.now()
    results = []

    for c in clients:
        score = 0
        reasons = []
        interval = CHECKUP_INTERVALS.get(c.segment or "", 90)
        last = c.last_meeting_date or c.last_checkup

        # Фактор 1: Нет контакта > 2x интервала
        if last and (now - last).days > interval * 2:
            score += 40
            reasons.append(f"Нет контакта {(now-last).days} дн. (норма: {interval})")

        # Фактор 2: Low health score
        if c.health_score and c.health_score < 0.3:
            score += 30
            reasons.append(f"Низкий health score: {c.health_score:.0%}")

        # Фактор 3: Blocked tasks
        blocked = db.query(Task).filter(Task.client_id == c.id, Task.status == "blocked").count()
        if blocked > 0:
            score += 15
            reasons.append(f"{blocked} заблокированных задач")

        # Фактор 4: Нет задач вообще
        total_tasks = db.query(Task).filter(Task.client_id == c.id).count()
        if total_tasks == 0:
            score += 15
            reasons.append("Нет задач")

        risk = "low"
        if score >= 60:
            risk = "critical"
        elif score >= 30:
            risk = "medium"

        results.append({"id": c.id, "name": c.name, "segment": c.segment, "risk": risk, "score": score, "reasons": reasons})

    results.sort(key=lambda x: x["score"], reverse=True)
    return {"clients": results}


# ============================================================================
# AI AUTO-QBR PAGE

@router.get("/api/stats")
async def api_stats(db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    """Быстрая статистика для sidebar — вызывается на каждой странице."""
    if not auth_token:
        return {"overdue": 0, "warning": 0, "open_tasks": 0}
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        return {"overdue": 0, "warning": 0, "open_tasks": 0}
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user:
        return {"overdue": 0, "warning": 0, "open_tasks": 0}

    # Кеш 90 сек — вызывается на каждой странице от 18 менеджеров
    ck = f"stats:{user.id}"
    cached = cache_get(ck)
    if cached:
        return cached

    now = datetime.now()
    q = db.query(Client)
    if user.role == "manager":
        q = q.filter(Client.manager_email == user.email)
    clients = q.all()

    overdue = warning = 0
    for c in clients:
        last = c.last_meeting_date or c.last_checkup
        if not last:
            continue
        days = (now - last).days
        from models import CHECKUP_INTERVALS
        interval = CHECKUP_INTERVALS.get(c.segment or "", 90)
        if days > interval:
            overdue += 1
        elif days > interval - 14:
            warning += 1

    tq = db.query(Task).join(Client, Task.client_id == Client.id, isouter=True).filter(
        Task.status.in_(["plan", "in_progress", "blocked"])
    )
    if user.role == "manager":
        tq = tq.filter(Client.manager_email == user.email)
    open_tasks = tq.count()

    result = {"overdue": overdue, "warning": warning, "open_tasks": open_tasks}
    cache_set(ck, result, ttl=90)
    return result



@router.get("/api/team/kpi")
async def api_team_kpi(
    period_days: int = 30,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """KPI всей команды — только для admin."""
    if not auth_token:
        raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload:
        raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user or user.role != "admin":
        raise HTTPException(status_code=403)

    since = datetime.utcnow() - timedelta(days=period_days)
    managers = db.query(User).filter(User.role == "manager", User.is_active == True).all()
    result = []

    for m in managers:
        tasks_closed = db.query(Task).join(Client, Task.client_id == Client.id, isouter=True).filter(
            Client.manager_email == m.email,
            Task.status == "done", Task.confirmed_at >= since,
        ).count()
        meetings_held = db.query(Meeting).join(Client, Meeting.client_id == Client.id, isouter=True).filter(
            Client.manager_email == m.email,
            Meeting.date >= since, Meeting.date <= datetime.utcnow(),
        ).count()
        followups_sent = db.query(Meeting).join(Client, Meeting.client_id == Client.id, isouter=True).filter(
            Client.manager_email == m.email,
            Meeting.followup_status == "sent", Meeting.followup_sent_at >= since,
        ).count()
        clients_count = db.query(Client).filter(Client.manager_email == m.email).count()

        result.append({
            "manager": m.email,
            "name": f"{m.first_name or ''} {m.last_name or ''}".strip() or m.email,
            "tasks_closed": tasks_closed,
            "meetings_held": meetings_held,
            "followups_sent": followups_sent,
            "clients": clients_count,
        })

    result.sort(key=lambda x: x["tasks_closed"] + x["meetings_held"], reverse=True)
    return {"period_days": period_days, "managers": result}


# ============================================================================
# AIRTABLE WEBHOOK (входящие изменения из Airtable)
# ============================================================================

# ============================================================================
# ОНБОРДИНГ ПАРТНЁРА: чеклист
# ============================================================================

ONBOARDING_CHECKLIST = [
    {"id": 1, "title": "Провести вводную встречу (1 ч)", "type": "meeting", "day": 0},
    {"id": 2, "title": "Отправить welcome-фолоуап", "type": "followup", "day": 0},
    {"id": 3, "title": "Добавить в карточку клиента", "type": "admin", "day": 1},
    {"id": 4, "title": "Касание в Ktalk (день 3)", "type": "ktalk", "day": 3},
    {"id": 5, "title": "Проверить первые шаги партнёра", "type": "check", "day": 7},
    {"id": 6, "title": "Касание в Ktalk (день 7)", "type": "ktalk", "day": 7},
    {"id": 7, "title": "Первый чекап (2 нед)", "type": "meeting", "day": 14},
    {"id": 8, "title": "Касание в Ktalk (день 14)", "type": "ktalk", "day": 14},
    {"id": 9, "title": "Проверить health score", "type": "check", "day": 30},
    {"id": 10, "title": "Закрыть онбординг, перевести в активные", "type": "admin", "day": 30},
]



@router.get("/api/team/stats")
async def api_team_stats(
    days: int = 30,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    if not auth_token: raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload: raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user or user.role != "admin": raise HTTPException(status_code=403)

    managers = db.query(User).filter(User.role == "manager", User.is_active == True).all()
    all_clients = db.query(Client).all()
    health_all  = [c.health_score for c in all_clients if c.health_score is not None]

    mgr_stats = []
    now = datetime.utcnow()
    for m in managers:
        clients = [c for c in all_clients if c.manager_email == m.email]
        h_vals  = [c.health_score for c in clients if c.health_score is not None]
        avg_h   = sum(h_vals) / len(h_vals) if h_vals else 0
        overdue = sum(1 for c in clients if (c.last_meeting_date and (now - c.last_meeting_date).days > CHECKUP_INTERVALS.get(c.segment or "", 90)))
        mgr_stats.append({
            "id": m.id, "name": m.name or m.email, "email": m.email,
            "clients_count": len(clients), "avg_health": avg_h, "overdue": overdue,
        })
    mgr_stats.sort(key=lambda x: x["avg_health"], reverse=True)

    risk_clients = sorted(
        [{"id": c.id, "name": c.name, "health_score": c.health_score, "segment": c.segment,
          "manager_name": next((m.name or m.email for m in managers if m.email == c.manager_email), "—"),
          "last_contact": c.last_meeting_date.isoformat() if c.last_meeting_date else None}
         for c in all_clients if c.health_score is not None and c.health_score < 55],
        key=lambda x: x["health_score"]
    )[:20]

    open_tasks = db.query(Task).filter(Task.status != "done").count()
    overdue_ck  = sum(1 for c in all_clients
                      if c.last_meeting_date and
                      (now - c.last_meeting_date).days > CHECKUP_INTERVALS.get(c.segment or "", 90))

    return {
        "managers_count": len(managers),
        "total_clients":  len(all_clients),
        "avg_health":     sum(health_all) / len(health_all) if health_all else 0,
        "overdue_checkups": overdue_ck,
        "open_tasks":     open_tasks,
        "managers":       mgr_stats,
        "risk_clients":   risk_clients,
    }



@router.get("/api/team/export")
async def api_team_export(days: int = 30, db: Session = Depends(get_db), auth_token: Optional[str] = Cookie(None)):
    if not auth_token: raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload: raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user or user.role != "admin": raise HTTPException(status_code=403)

    import csv, io
    managers = db.query(User).filter(User.role == "manager").all()
    clients  = db.query(Client).all()
    buf = io.StringIO()
    w   = csv.writer(buf)
    w.writerow(["Менеджер","Email","Клиентов","Avg Health","Просрочено"])
    for m in managers:
        mc = [c for c in clients if c.manager_email == m.email]
        hv = [c.health_score for c in mc if c.health_score is not None]
        ah = sum(hv)/len(hv) if hv else 0
        now = datetime.utcnow()
        ov  = sum(1 for c in mc if c.last_meeting_date and (now - c.last_meeting_date).days > 90)
        w.writerow([m.name or "", m.email, len(mc), f"{ah:.0f}%", ov])
    from fastapi.responses import Response
    return Response(content=buf.getvalue(), media_type="text/csv",
                    headers={"Content-Disposition": f"attachment; filename=team_{datetime.utcnow().strftime('%Y%m%d')}.csv"})


# ============================================================================
# BULK OPERATIONS

@router.post("/api/bulk/assign-checkup")
async def api_bulk_assign_checkup(
    request: Request,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """Назначить чекап нескольким клиентам сразу."""
    if not auth_token: raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload: raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user: raise HTTPException(status_code=401)

    body = await request.json()
    client_ids = body.get("client_ids", [])
    date_str   = body.get("date")
    due_date   = datetime.fromisoformat(date_str) if date_str else datetime.utcnow() + timedelta(days=7)

    created = 0
    for cid in client_ids[:50]:  # лимит 50
        task = Task(client_id=cid, title="Провести чекап", status="plan",
                    priority="high", due_date=due_date, created_at=datetime.utcnow())
        db.add(task)
        created += 1
    db.commit()
    return {"ok": True, "created": created}



@router.post("/api/bulk/create-task")
async def api_bulk_create_task(
    request: Request,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """Создать одну задачу для нескольких клиентов."""
    if not auth_token: raise HTTPException(status_code=401)
    from auth import decode_access_token
    payload = decode_access_token(auth_token)
    if not payload: raise HTTPException(status_code=401)
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user: raise HTTPException(status_code=401)

    body = await request.json()
    client_ids = body.get("client_ids", [])
    title      = body.get("title", "Задача")
    priority   = body.get("priority", "medium")
    due_date   = datetime.fromisoformat(body["due_date"]) if body.get("due_date") else datetime.utcnow() + timedelta(days=3)

    created = 0
    for cid in client_ids[:50]:
        db.add(Task(client_id=cid, title=title, status="plan", priority=priority,
                    due_date=due_date, created_at=datetime.utcnow()))
        created += 1
    db.commit()
    return {"ok": True, "created": created}



@router.patch("/api/bulk/update-segment")
async def api_bulk_update_segment(
    request: Request,
    db: Session = Depends(get_db),
    auth_token: Optional[str] = Cookie(None),
):
    """Сменить сегмент у нескольких клиентов."""
    _require_admin(auth_token, db)
    body = await request.json()
    client_ids = body.get("client_ids", [])
    segment    = body.get("segment", "")
    if not segment: raise HTTPException(status_code=400, detail="segment required")

    db.query(Client).filter(Client.id.in_(client_ids)).update(
        {"segment": segment}, synchronize_session=False
    )
    db.commit()
    return {"ok": True, "updated": len(client_ids)}
# ============================================================================
# REVENUE TRACKING
# ============================================================================

# ============================================================================
# CHURN SCORING

@router.get("/api/analytics/churn")
async def api_analytics_churn(
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Сводка по churm рискам портфеля."""
    from models import ChurnScore
    q = db.query(Client)
    if user.role == "manager": q = q.filter(Client.manager_email == user.email)
    client_ids = [c.id for c in q.all()]

    scores = db.query(ChurnScore).filter(ChurnScore.client_id.in_(client_ids)).all() if client_ids else []
    dist = {"low": 0, "medium": 0, "high": 0, "critical": 0}
    top_risk = []
    for s in scores:
        dist[s.risk_level] = dist.get(s.risk_level, 0) + 1
        if s.risk_level in ("high", "critical"):
            c = db.query(Client).filter(Client.id == s.client_id).first()
            if c: top_risk.append({"id": c.id, "name": c.name, "score": s.score,
                                    "risk_level": s.risk_level, "segment": c.segment})
    top_risk.sort(key=lambda x: -x["score"])
    return {"distribution": dist, "top_risk": top_risk[:10], "total_scored": len(scores)}


# ============================================================================
# DEDUPLICATION

@router.get("/api/analytics/revenue")
async def api_revenue_analytics(
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Аналитика выручки портфеля."""
    q = db.query(Client)
    if user.role == "manager":
        q = q.filter(Client.manager_email == user.email)
    clients = q.all()

    total_mrr = 0.0
    at_risk_mrr = 0.0
    by_segment: dict = {}
    top_clients = []

    for c in clients:
        meta = c.integration_metadata or {}
        mrr  = float(meta.get("mrr") or 0)
        if not mrr:
            continue
        total_mrr += mrr
        seg = c.segment or "Unknown"
        by_segment[seg] = by_segment.get(seg, 0.0) + mrr

        # Считаем "под угрозой" если health < 50 или нет контакта давно
        is_risk = (c.health_score or 50) < 50
        if is_risk:
            at_risk_mrr += mrr

        top_clients.append({
            "id": c.id, "name": c.name, "segment": c.segment,
            "mrr": mrr, "health_score": c.health_score, "is_risk": is_risk,
        })

    top_clients.sort(key=lambda x: x["mrr"], reverse=True)

    return {
        "total_mrr": round(total_mrr, 2),
        "total_arr": round(total_mrr * 12, 2),
        "at_risk_mrr": round(at_risk_mrr, 2),
        "at_risk_pct": round(at_risk_mrr / total_mrr * 100, 1) if total_mrr else 0,
        "by_segment": [{"segment": k, "mrr": round(v, 2)} for k, v in
                       sorted(by_segment.items(), key=lambda x: x[1], reverse=True)],
        "top_clients": top_clients[:10],
        "clients_with_mrr": len(top_clients),
        "clients_total": len(clients),
    }


# ============================================================================
# DATA VALIDATION — клиенты с неполными данными

@router.get("/api/export/excel")
async def api_export_excel(
    scope: str = "clients",  # clients | tasks | checkups | full
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Excel экспорт с форматированием, несколько листов."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from fastapi.responses import Response

    wb = Workbook()

    # ── Стили ──────────────────────────────────────────────────────────────────
    HDR_FILL = PatternFill("solid", start_color="07090F", end_color="07090F")
    HDR_FONT = Font(name="Arial", bold=True, color="6474FF", size=10)
    ROW_FILL = [
        PatternFill("solid", start_color="0D1117", end_color="0D1117"),
        PatternFill("solid", start_color="131924", end_color="131924"),
    ]
    GREEN_FONT  = Font(name="Arial", color="23D18B", size=9)
    YELLOW_FONT = Font(name="Arial", color="F5A623", size=9)
    RED_FONT    = Font(name="Arial", color="F0556A", size=9)
    DEF_FONT    = Font(name="Arial", color="E8ECF4", size=9)
    MONO_FONT   = Font(name="Courier New", color="E8ECF4", size=9)

    def style_header(ws, headers: list):
        for col, (label, width) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=label)
            cell.font = HDR_FONT; cell.fill = HDR_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.row_dimensions[1].height = 24
        ws.freeze_panes = "A2"

    def write_row(ws, row_num: int, values: list, fonts: list = None):
        fill = ROW_FILL[(row_num - 2) % 2]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.fill = fill
            cell.alignment = Alignment(vertical="center")
            f = (fonts[col-1] if fonts and col-1 < len(fonts) else None) or DEF_FONT
            cell.font = f
        ws.row_dimensions[row_num].height = 18

    # ── Лист 1: Клиенты ────────────────────────────────────────────────────────
    ws1 = wb.active; ws1.title = "Клиенты"
    style_header(ws1, [
        ("ID",22), ("Название",36), ("Сегмент",16), ("Домен",30),
        ("Health %",14), ("Менеджер",24), ("MRR ₽",16), ("Последняя встреча",22),
        ("Риск оттока",16), ("Задач открытых",18),
    ])
    q = db.query(Client)
    if user.role == "manager":
        q = q.filter(Client.manager_email == user.email)
    clients = q.order_by(Client.name).all()
    from models import ChurnScore
    for i, c in enumerate(clients, 2):
        meta = c.integration_metadata or {}
        mrr  = meta.get("mrr", "")
        cs   = db.query(ChurnScore).filter(ChurnScore.client_id == c.id).first()
        open_t = db.query(Task).filter(Task.client_id == c.id, Task.status != "done").count()
        h = c.health_score or 0
        hfont = GREEN_FONT if h >= 70 else YELLOW_FONT if h >= 40 else RED_FONT
        write_row(ws1, i, [
            c.id, c.name, c.segment or "—", c.domain or "—",
            f"{h:.0f}%" if c.health_score else "—",
            c.manager_email or "—",
            f"{mrr:,.0f}" if mrr else "—",
            c.last_meeting_date.strftime("%d.%m.%Y") if c.last_meeting_date else "—",
            cs.risk_level if cs else "—", open_t,
        ], fonts=[MONO_FONT, None, None, None, hfont, None, MONO_FONT, None, None, MONO_FONT])

    # ── Лист 2: Задачи ─────────────────────────────────────────────────────────
    if scope in ("tasks", "full"):
        ws2 = wb.create_sheet("Задачи")
        style_header(ws2, [
            ("ID",12), ("Задача",44), ("Клиент",30), ("Статус",18),
            ("Приоритет",16), ("Срок",18), ("Синк MR",16),
        ])
        tq = db.query(Task).join(Client, Task.client_id == Client.id, isouter=True)
        if user.role == "manager":
            tq = tq.filter(Client.manager_email == user.email)
        tasks = tq.order_by(Task.due_date.asc().nullslast()).limit(2000).all()
        STATUS_LABELS = {"plan":"Планирование","in_progress":"В работе","review":"Ревью","done":"Выполнено","blocked":"Заблокировано"}
        PRIO_LABELS   = {"high":"Высокий","medium":"Средний","low":"Низкий"}
        for i, t in enumerate(tasks, 2):
            now = datetime.utcnow()
            is_overdue = t.due_date and t.due_date < now and t.status != "done"
            dfont = RED_FONT if is_overdue else DEF_FONT
            write_row(ws2, i, [
                t.id, t.title, t.client.name if t.client else "—",
                STATUS_LABELS.get(t.status, t.status),
                PRIO_LABELS.get(t.priority, t.priority),
                t.due_date.strftime("%d.%m.%Y") if t.due_date else "—",
                "✅" if t.merchrules_task_id else "—",
            ], fonts=[MONO_FONT, None, None, None, None, dfont, None])

    # ── Лист 3: Чекапы ─────────────────────────────────────────────────────────
    if scope in ("checkups", "full"):
        from models import CheckupResult
        ws3 = wb.create_sheet("Чекапы")
        style_header(ws3, [
            ("ID",12), ("Клиент",32), ("Дата",18), ("Запросов",14),
            ("Avg Score",14), ("Продукт",16), ("Менеджер",24),
        ])
        cq = db.query(CheckupResult).join(Client, CheckupResult.client_id == Client.id, isouter=True)
        if user.role == "manager":
            cq = cq.filter(Client.manager_email == user.email)
        results = cq.order_by(CheckupResult.created_at.desc()).limit(500).all()
        for i, r in enumerate(results, 2):
            avg = r.avg_score or 0
            afont = GREEN_FONT if avg >= 2.5 else YELLOW_FONT if avg >= 1.5 else RED_FONT
            write_row(ws3, i, [
                r.id, r.client.name if r.client else "—",
                r.created_at.strftime("%d.%m.%Y %H:%M") if r.created_at else "—",
                r.total_queries or 0, f"{avg:.2f}", r.query_type or "—",
                r.manager_name or "—",
            ], fonts=[MONO_FONT, None, None, MONO_FONT, afont, None, None])

    # ── Сохраняем ──────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"amhub_export_{scope}_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


# ============================================================================
# FILE ATTACHMENTS

